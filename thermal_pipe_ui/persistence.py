##########################################################################
# __author__         = "Tim Kayser"
# __date__           = "09.03.2026"
# __version__        = "2.1"
# __maintainer__     = "Tim Kayser"
# __email__          = "kaysert@purdue.edu"
# __status__         = "Open Beta"
# __copyright__      = "Copyright 2026"
# __credits__        = ["Tim Kayser"]
# __license__        = "GPL"
##########################################################################
"""Configuration persistence, ledger, and export mixin for the thermal pipe application."""

from __future__ import annotations

import csv
import io
import json
import logging
import shutil
import traceback
import zipfile
from datetime import datetime
from pathlib import Path
from typing import Any, Dict

import numpy as np

from PyQt6.QtCore import Qt
from PyQt6.QtWidgets import QFileDialog, QMessageBox, QTableWidgetItem

from .window import HAS_MPL, INSULATION_MATERIALS, INSULATION_TEMP_PROPS, PIPE_MATERIALS, PIPE_TEMP_PROPS


class PersistenceMixin:
    def _all_pipe_defaults(self) -> Dict[str, Dict[str, Any]]:
        merged = dict(self.PIPE_DEFAULTS)
        merged.update(self._custom_pipe_presets)
        return merged

    def _refresh_pipe_default_combo(self):
        if not hasattr(self, "pipe_default"):
            return
        current = self.pipe_default.currentText()
        names = list(self.PIPE_DEFAULTS.keys()) + sorted(
            [n for n in self._custom_pipe_presets.keys() if n not in self.PIPE_DEFAULTS]
        )
        self.pipe_default.blockSignals(True)
        self.pipe_default.clear()
        self.pipe_default.addItems(names)
        if current in names:
            self.pipe_default.setCurrentText(current)
        elif names:
            self.pipe_default.setCurrentIndex(0)
        self.pipe_default.blockSignals(False)

    def _refresh_pipe_preset_list(self):
        if not hasattr(self, "pipe_preset_list"):
            return
        self.pipe_preset_list.clear()
        for name in list(self.PIPE_DEFAULTS.keys()) + sorted(
            [n for n in self._custom_pipe_presets.keys() if n not in self.PIPE_DEFAULTS]
        ):
            self.pipe_preset_list.addItem(name)

    def _load_custom_pipe_presets(self):
        self._custom_pipe_presets = {}
        if not self._pipe_preset_path.exists():
            return
        try:
            raw = json.loads(self._pipe_preset_path.read_text(encoding="utf-8"))
            if isinstance(raw, dict):
                for name, cfg in raw.items():
                    if not isinstance(name, str) or not isinstance(cfg, dict):
                        continue
                    if name in self.PIPE_DEFAULTS:
                        continue
                    self._custom_pipe_presets[name] = cfg
        except Exception as exc:
            logging.warning("Failed to load pipe presets from %s: %s", self._pipe_preset_path, exc)

    def _save_custom_pipe_presets(self):
        try:
            self._pipe_preset_path.write_text(
                json.dumps(self._custom_pipe_presets, indent=2, sort_keys=True),
                encoding="utf-8",
            )
        except Exception as exc:
            QMessageBox.warning(self, "Preset Library", f"Failed to save presets: {exc}")

    def _capture_current_pipe_preset(self) -> Dict[str, Any]:
        L_si = float(self._length_from_display(self.in_L.value()))
        n_elbows = int(self.in_elbows.value())
        elbow_positions_frac = self._parse_elbow_positions(
            self.in_elbow_positions.text(),
            n_elbows,
            L_si,
        )
        n_tmass = int(self.in_tmass_count.value())
        tmass_positions_frac = self._parse_fractional_positions(
            self.in_tmass_positions.text(),
            n_tmass,
            L_si,
        )
        elbow_positions_text = ", ".join(f"{100.0 * v:.1f}%" for v in elbow_positions_frac) if n_elbows > 0 else ""
        tmass_positions_text = ", ".join(f"{100.0 * v:.1f}%" for v in tmass_positions_frac) if n_tmass > 0 else ""
        return {
            "pipe_material": self.pipe_material.currentText(),
            "insulation": bool(self.chk_use_insulation.isChecked()),
            "ins_material": self.ins_material.currentText(),
            "L": L_si,
            "Di": float(self._diam_from_display(self.in_Di.value())),
            "t_wall": float(self._diam_from_display(self.in_t_wall.value())),
            "t_ins": float(self._diam_from_display(self.in_t_ins.value())),
            "Tamb": float(self._temp_from_display(self.in_ambient.value())),
            "Tin": float(self._temp_from_display(self.in_Tin.value())),
            "Tin_ramp_s": float(self.in_tin_ramp.value()),
            "Tin_ramp_model": str(self.in_tin_ramp_model.currentData() or "logistic"),
            "p": float(self._pressure_from_display(self.in_p.value())),
            "m_dot": float(self._mdot_from_display(self.in_mdot.value())),
            "n_elbows": n_elbows,
            "elbow_sif": float(self.in_elbow_sif.value()),
            "elbow_positions": elbow_positions_text,
            "n_tmass": n_tmass,
            "tmass_factor": float(self.in_tmass_factor.value()),
            "tmass_positions": tmass_positions_text,
            "tmass_spread_pct": float(self.in_tmass_spread.value()),
            "tmass_deadleg_len_m": float(self._length_from_display(self.in_tmass_deadleg_len.value())),
            "tmass_deadleg_d_ratio": float(self.in_tmass_deadleg_d_ratio.value()),
        }

    def _save_current_pipe_preset(self):
        name = (self.new_preset_name.text() if hasattr(self, "new_preset_name") else "").strip()
        if not name:
            QMessageBox.warning(self, "Preset Library", "Preset name is required.")
            return
        if name in self.PIPE_DEFAULTS:
            QMessageBox.warning(self, "Preset Library", "Choose a new name (cannot overwrite built-in presets).")
            return
        self._custom_pipe_presets[name] = self._capture_current_pipe_preset()
        self._save_custom_pipe_presets()
        self._refresh_pipe_default_combo()
        self._refresh_pipe_preset_list()
        self.pipe_default.setCurrentText(name)
        self.new_preset_name.clear()
        QMessageBox.information(self, "Preset Library", f"Saved preset: {name}")

    def _selected_pipe_preset_name(self) -> str | None:
        if not hasattr(self, "pipe_preset_list"):
            return None
        item = self.pipe_preset_list.currentItem()
        if item is None:
            return None
        name = item.text().strip()
        return name or None

    def _apply_selected_pipe_preset(self):
        name = self._selected_pipe_preset_name()
        if not name:
            return
        self.pipe_default.setCurrentText(name)
        self._apply_pipe_default()

    def _delete_selected_pipe_preset(self):
        name = self._selected_pipe_preset_name()
        if not name:
            return
        if name in self.PIPE_DEFAULTS:
            QMessageBox.warning(self, "Preset Library", "Built-in presets cannot be deleted.")
            return
        if name in self._custom_pipe_presets:
            del self._custom_pipe_presets[name]
            self._save_custom_pipe_presets()
            self._refresh_pipe_default_combo()
            self._refresh_pipe_preset_list()
            QMessageBox.information(self, "Preset Library", f"Deleted preset: {name}")

    def _choose_dir(self):
        path = QFileDialog.getExistingDirectory(self, "Select save folder")
        if path:
            self._save_dir = Path(path)
            self.dir_label.setText(self._save_dir.name or str(self._save_dir))
            self.dir_label.setToolTip(str(self._save_dir))
        else:
            self._save_dir = None
            self.dir_label.setText("(auto)")
            self.dir_label.setToolTip("")
        self._refresh_ledger_preview()

    def _choose_ledger_file(self):
        path, _ = QFileDialog.getSaveFileName(
            self,
            "Select or create run ledger",
            str((self._save_dir or Path.cwd()) / "run_history.csv"),
            "CSV files (*.csv);;Excel Workbook (*.xlsx)",
        )
        if path:
            self._ledger_path = Path(path)
            self.ledger_label.setText(self._ledger_path.name)
            self.ledger_label.setToolTip(str(self._ledger_path))
        else:
            self._ledger_path = None
            self.ledger_label.setText("(auto)")
            self.ledger_label.setToolTip(str(self._resolve_ledger_path()))
        self._refresh_ledger_preview()

    def _resolve_ledger_path(self) -> Path:
        if self._ledger_path is not None:
            return self._ledger_path
        base = self._save_dir if self._save_dir is not None else Path.cwd()
        return base / "run_history.csv"

    def _refresh_readme_view(self):
        if not hasattr(self, "readme_view"):
            return
        self.lbl_readme_path.setText(str(self._readme_path))
        if not self._readme_path.exists():
            self.readme_view.setPlainText(f"README not found at: {self._readme_path}")
            return
        try:
            text = self._readme_path.read_text(encoding="utf-8")
        except Exception as exc:
            self.readme_view.setPlainText(f"Failed to read README: {exc}")
            return
        try:
            self.readme_view.setMarkdown(text)
        except Exception:
            self.readme_view.setPlainText(text)

    @staticmethod
    def _merge_ledger_headers(existing: list[str], incoming: list[str]) -> list[str]:
        merged: list[str] = [str(h).strip() for h in existing if str(h).strip()]
        for key in incoming:
            k = str(key).strip()
            if k and k not in merged:
                merged.append(k)
        return merged

    @staticmethod
    def _normalize_legacy_ledger_row(values: list[Any], header_len: int) -> list[Any]:
        # Legacy schema lacked asset_id/branch_id; newer rows may already contain both fields.
        if not values:
            return ["", "", ""]
        if len(values) >= header_len + 2:
            return [values[0], values[1], values[2]] + values[3:]
        return [values[0], "", ""] + values[1:]

    def _load_csv_ledger_rows(self, ledger_path: Path) -> tuple[list[str], list[dict[str, Any]]]:
        with ledger_path.open("r", encoding="utf-8", newline="") as f:
            raw = list(csv.reader(f))
        if not raw:
            return [], []

        header_base = [str(v).strip() for v in raw[0] if str(v).strip()]
        if not header_base:
            return [], []

        legacy_schema = (
            header_base[0] == "timestamp_local"
            and "asset_id" not in header_base
            and "branch_id" not in header_base
        )
        header_len_base = len(header_base)
        headers = [header_base[0], "asset_id", "branch_id"] + header_base[1:] if legacy_schema else list(header_base)

        rows: list[dict[str, Any]] = []
        extra_headers: list[str] = []
        for raw_row in raw[1:]:
            vals: list[Any] = list(raw_row)
            if not any(str(v).strip() for v in vals):
                continue
            if legacy_schema:
                vals = self._normalize_legacy_ledger_row(vals, header_len_base)
            if len(vals) < len(headers):
                vals.extend([""] * (len(headers) - len(vals)))

            row_map: dict[str, Any] = {h: vals[i] if i < len(vals) else "" for i, h in enumerate(headers)}
            if len(vals) > len(headers):
                for k, extra in enumerate(vals[len(headers):], start=1):
                    extra_name = f"extra_{k}"
                    if extra_name not in extra_headers:
                        extra_headers.append(extra_name)
                    row_map[extra_name] = extra
            rows.append(row_map)

        headers = headers + [h for h in extra_headers if h not in headers]
        return headers, rows

    @staticmethod
    def _write_csv_ledger_rows(ledger_path: Path, headers: list[str], rows: list[dict[str, Any]]) -> None:
        with ledger_path.open("w", encoding="utf-8", newline="") as f:
            writer = csv.DictWriter(f, fieldnames=headers)
            writer.writeheader()
            for row in rows:
                writer.writerow({h: row.get(h, "") for h in headers})

    def _load_xlsx_ledger_rows(self, ledger_path: Path) -> tuple[list[str], list[dict[str, Any]]]:
        try:
            from openpyxl import load_workbook  # type: ignore
        except Exception as exc:  # pragma: no cover - optional dependency
            raise RuntimeError(f"openpyxl required for XLSX support: {exc}") from exc

        wb = load_workbook(ledger_path)
        ws = wb.active
        data = list(ws.iter_rows(values_only=True))
        if not data:
            return [], []

        header_base = [str(v).strip() if v is not None else "" for v in data[0]]
        while header_base and not header_base[-1]:
            header_base.pop()
        if not header_base:
            return [], []

        legacy_schema = (
            header_base[0] == "timestamp_local"
            and "asset_id" not in header_base
            and "branch_id" not in header_base
        )
        header_len_base = len(header_base)
        headers = [header_base[0], "asset_id", "branch_id"] + header_base[1:] if legacy_schema else list(header_base)

        rows: list[dict[str, Any]] = []
        extra_headers: list[str] = []
        for raw_row in data[1:]:
            vals: list[Any] = [v for v in raw_row]
            while vals and vals[-1] is None:
                vals.pop()
            if not any(v is not None and str(v).strip() for v in vals):
                continue
            if legacy_schema:
                vals = self._normalize_legacy_ledger_row(vals, header_len_base)
            if len(vals) < len(headers):
                vals.extend([""] * (len(headers) - len(vals)))

            row_map: dict[str, Any] = {h: vals[i] if i < len(vals) else "" for i, h in enumerate(headers)}
            if len(vals) > len(headers):
                for k, extra in enumerate(vals[len(headers):], start=1):
                    extra_name = f"extra_{k}"
                    if extra_name not in extra_headers:
                        extra_headers.append(extra_name)
                    row_map[extra_name] = extra
            rows.append(row_map)

        headers = headers + [h for h in extra_headers if h not in headers]
        return headers, rows

    @staticmethod
    def _write_xlsx_ledger_rows(ledger_path: Path, headers: list[str], rows: list[dict[str, Any]]) -> None:
        try:
            from openpyxl import Workbook  # type: ignore
        except Exception as exc:  # pragma: no cover - optional dependency
            raise RuntimeError(f"openpyxl required for XLSX support: {exc}") from exc

        wb = Workbook()
        ws = wb.active
        ws.title = "run_history"
        ws.append(headers)
        for row in rows:
            ws.append([row.get(h, "") for h in headers])
        wb.save(ledger_path)

    def _refresh_ledger_preview(self):
        if not hasattr(self, "ledger_table"):
            return
        ledger_path = self._resolve_ledger_path()
        self.lbl_ledger_status.setText(f"Ledger: {ledger_path.name}")
        self.lbl_ledger_status.setToolTip(str(ledger_path))
        self.ledger_table.clear()
        self.ledger_table.setRowCount(0)
        self.ledger_table.setColumnCount(0)

        if not ledger_path.exists():
            self.ledger_table.setRowCount(1)
            self.ledger_table.setColumnCount(1)
            self.ledger_table.setHorizontalHeaderLabels(["Info"])
            self.ledger_table.setItem(0, 0, QTableWidgetItem("Ledger file does not exist yet."))
            return

        try:
            if ledger_path.suffix.lower() == ".xlsx":
                headers, rows = self._load_xlsx_ledger_rows(ledger_path)
            else:
                headers, rows = self._load_csv_ledger_rows(ledger_path)
        except Exception as exc:
            self.ledger_table.setRowCount(1)
            self.ledger_table.setColumnCount(1)
            self.ledger_table.setHorizontalHeaderLabels(["Info"])
            self.ledger_table.setItem(0, 0, QTableWidgetItem(str(exc)))
            return

        self._ledger_rows = rows
        self._ledger_fieldnames = headers
        self.lbl_ledger_status.setText(f"Ledger: {ledger_path.name} ({len(rows)} rows)")
        display_headers = self._select_ledger_display_headers(headers)
        self.ledger_table.setColumnCount(len(display_headers))
        self.ledger_table.setHorizontalHeaderLabels(display_headers)
        self.ledger_table.setRowCount(len(rows))
        for i, row in enumerate(rows):
            for j, h in enumerate(display_headers):
                self.ledger_table.setItem(i, j, QTableWidgetItem(self._format_ledger_cell(h, row.get(h, ""))))
        self._set_ledger_column_widths(display_headers)

    @staticmethod
    def _select_ledger_display_headers(headers: list[str]) -> list[str]:
        preferred = [
            "timestamp_local",
            "asset_id",
            "branch_id",
            "mode",
            "target_metric",
            "stop_reason",
            "reached_target",
            "m_dot_kg_s",
            "Tg_out_final_K",
            "time_to_target_s",
            "max_vm_elbow_mpa",
            "t_max_vm_s",
            "x_max_vm_m",
            "warnings",
        ]
        out = [h for h in preferred if h in headers]
        if out:
            return out
        return headers[: min(12, len(headers))]

    def _set_ledger_column_widths(self, headers: list[str]):
        if not hasattr(self, "ledger_table"):
            return
        for j, h in enumerate(headers):
            key = str(h).strip().lower()
            if key in ("timestamp_local", "mode", "stop_reason", "asset_id", "branch_id"):
                w = 170
            elif "warnings" in key:
                w = 320
            elif key.endswith("_id"):
                w = 140
            else:
                w = 130
            self.ledger_table.setColumnWidth(j, w)

    def _append_current_config_to_ledger(self):
        try:
            _ = self._collect_spec()
        except Exception:
            pass

        class _Result:
            pass

        r = _Result()
        r.stop_reason = "manual_config_append"
        r.reached_Tg_target = False
        r.Tg_outlet_final = float(self._last_stats.get("Tg_outlet_final_K", np.nan))
        self._append_run_ledger(r)
        self._refresh_ledger_preview()

    def _delete_selected_ledger_rows(self):
        ledger_path = self._resolve_ledger_path()
        if not ledger_path.exists():
            return
        selected = sorted({idx.row() for idx in self.ledger_table.selectionModel().selectedRows()})
        if not selected:
            return
        try:
            if ledger_path.suffix.lower() == ".xlsx":
                headers, rows = self._load_xlsx_ledger_rows(ledger_path)
                keep = [row for i, row in enumerate(rows) if i not in set(selected)]
                self._write_xlsx_ledger_rows(ledger_path, headers, keep)
            else:
                headers, rows = self._load_csv_ledger_rows(ledger_path)
                keep = [row for i, row in enumerate(rows) if i not in set(selected)]
                self._write_csv_ledger_rows(ledger_path, headers, keep)
        except Exception as exc:
            QMessageBox.warning(self, "Ledger", f"Could not edit ledger: {exc}")
            return
        self._refresh_ledger_preview()

    def _append_run_ledger(self, result):
        now = datetime.now()
        ledger_path = self._ledger_path
        if ledger_path is None:
            base = self._save_dir if self._save_dir is not None else Path.cwd()
            ledger_path = base / "run_history.csv"
        stats = self._last_stats if self._last_stats else {}
        opt_summary = getattr(result, "opt_summary", None)

        row = {
            "timestamp_local": now.strftime("%Y-%m-%d %H:%M:%S"),
            "asset_id": str(self._last_run_si.get("asset_id", self.asset_id.text().strip())),
            "branch_id": str(self._last_run_si.get("branch_id", self.branch_id.text().strip())),
            "material": self._last_pipe_material,
            "mode": self.mode.currentText(),
            "stop_reason": str(result.stop_reason),
            "reached_target": int(bool(result.reached_Tg_target)),
            "target_metric": str(self._last_run_si.get("target_metric", "gas_outlet")),
            "target_temp_K": float(self._last_run_si.get("target", np.nan)),
            "target_outlet_final_K": float(getattr(result, "target_outlet_final", np.nan)),
            "L_m": float(self._last_run_si.get("L", 0.0)),
            "Di_m": float(self._last_run_si.get("Di", 0.0)),
            "t_wall_m": float(self._last_run_si.get("t_wall", 0.0)),
            "t_ins_m": float(self._last_run_si.get("t_ins", 0.0)),
            "p_Pa": float(self._last_run_si.get("p", 0.0)),
            "Tin_K": float(self._last_run_si.get("Tin", 0.0)),
            "Tamb_K": float(self._last_run_si.get("Tamb", 0.0)),
            "m_dot_kg_s": float(self._last_run_si.get("m_dot", 0.0)),
            "Nx": int(self._last_run_si.get("Nx", 0)),
            "Nr_wall": int(self._last_run_si.get("nr_wall", 0)),
            "inlet_ramp_s": float(self._last_run_si.get("Tin_ramp_s", 0.0)),
            "inlet_ramp_model": str(self._last_run_si.get("Tin_ramp_model", "logistic")),
            "axial_restraint": float(self._last_run_si.get("axial_restraint", 0.0)),
            "elbows": int(self._last_geom.get("n_elbows", 0)),
            "elbow_sif": float(self._last_geom.get("elbow_sif", 1.0)),
            "thermal_masses": int(self._last_geom.get("n_tmass", 0)),
            "thermal_mass_factor": float(self._last_geom.get("tmass_factor", 0.0)),
            "thermal_mass_spread_pctL": float(self._last_geom.get("tmass_spread_frac", 0.0)) * 100.0,
            "thermal_mass_deadleg_len_m": float(self._last_geom.get("tmass_deadleg_len_m", 0.0)),
            "thermal_mass_deadleg_d_ratio": float(self._last_geom.get("tmass_deadleg_d_ratio", 1.0)),
            "Tg_out_final_K": float(result.Tg_outlet_final),
            "Tw_out_final_K": float(stats.get("Tw_outlet_final_K", np.nan)),
            "Tw_outer_surface_out_final_K": float(stats.get("Tw_outer_surface_outlet_K", np.nan)),
            "Ti_out_final_K": float(stats.get("Ti_outlet_final_K", np.nan)),
            "Tin_eff_final_K": float(stats.get("Tin_eff_final_K", np.nan)),
            "time_to_target_s": stats.get("time_to_target_s"),
            "max_vm_mpa": float(stats.get("sigma_max_mpa", np.nan)),
            "max_vm_elbow_mpa": float(stats.get("sigma_max_elbow_mpa", np.nan)),
            "max_vm_thermal_mpa": float(stats.get("sigma_max_thermal_mpa", np.nan)),
            "max_vm_thermal_elbow_mpa": float(stats.get("sigma_max_thermal_elbow_mpa", np.nan)),
            "t_max_vm_s": float(stats.get("t_sigma_max", np.nan)),
            "x_max_vm_m": float(stats.get("x_sigma_max_m", np.nan)),
            "max_wall_dT_K": float(stats.get("max_wall_dT_K", np.nan)),
            "warnings": " | ".join(self._last_warnings),
        }
        if isinstance(opt_summary, dict):
            row.update(
                {
                    "opt_mode": str(opt_summary.get("mode", "")),
                    "opt_selected_mdot_kg_s": float(opt_summary.get("m_dot_kg_s", np.nan)),
                    "opt_t_hit_final_s": opt_summary.get("time_to_target_final_s"),
                    "opt_sigma_final_mpa": float(opt_summary.get("sigma_final_mpa", np.nan)),
                    "opt_stress_limit_mpa": float(opt_summary.get("stress_limit_mpa", np.nan)),
                    "opt_heatup_target_s": float(opt_summary.get("heatup_target_s", np.nan)),
                    "opt_heatup_tol_s": float(opt_summary.get("heatup_tol_s", np.nan)),
                    "opt_meets_stress_limit": int(bool(opt_summary.get("meets_stress_limit", False))),
                    "opt_meets_heatup_tolerance": int(bool(opt_summary.get("meets_heatup_tolerance", False))),
                    "opt_target_reached_final": int(bool(opt_summary.get("target_reached_final", False))),
                }
            )
        row = {k: self._round_ledger_value(k, v) for k, v in row.items()}

        incoming_headers = list(row.keys())
        if ledger_path.suffix.lower() == ".xlsx":
            try:
                headers_existing, rows_existing = self._load_xlsx_ledger_rows(ledger_path) if ledger_path.exists() else ([], [])
                headers = self._merge_ledger_headers(headers_existing, incoming_headers)
                rows_existing.append(row)
                self._write_xlsx_ledger_rows(ledger_path, headers, rows_existing)
                self._refresh_ledger_preview()
                return
            except Exception as exc:
                logging.warning("XLSX ledger append failed; falling back to CSV ledger. Reason: %s", exc)
                ledger_path = ledger_path.with_suffix(".csv")

        headers_existing, rows_existing = self._load_csv_ledger_rows(ledger_path) if ledger_path.exists() else ([], [])
        headers = self._merge_ledger_headers(headers_existing, incoming_headers)
        rows_existing.append(row)
        self._write_csv_ledger_rows(ledger_path, headers, rows_existing)
        self._refresh_ledger_preview()

    def _export_bundle_clicked(self):
        if self._last_result is None:
            QMessageBox.information(self, "Export", "No run available yet. Run a simulation first.")
            return
        default_name = f"thermal_pipe_bundle_{datetime.now().strftime('%Y%m%d_%H%M%S')}.zip"
        base = self._save_dir if self._save_dir is not None else Path.cwd()
        path, _ = QFileDialog.getSaveFileName(self, "Export run bundle", str(base / default_name), "Zip archive (*.zip)")
        if not path:
            return
        zpath = Path(path)
        if zpath.suffix.lower() != ".zip":
            zpath = zpath.with_suffix(".zip")

        result = self._last_result
        try:
            with zipfile.ZipFile(zpath, "w", compression=zipfile.ZIP_DEFLATED) as zf:
                manifest = {
                    "created_local": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    "units": self._units,
                    "status": self.status.text(),
                    "has_saved_outdir": bool(result.outdir),
                }

                run_dir = result.outdir if result.outdir is not None else None
                if run_dir is not None:
                    run_dir = Path(run_dir)
                if run_dir is not None and run_dir.exists():
                    for fp in run_dir.rglob("*"):
                        if fp.is_file():
                            zf.write(fp, arcname=f"run/{fp.relative_to(run_dir)}")
                    manifest["run_source"] = str(run_dir)
                else:
                    buf = io.BytesIO()
                    np.savez_compressed(
                        buf,
                        x=np.asarray(result.x, dtype=float),
                        times=np.asarray(result.times, dtype=float),
                        Tg=np.asarray(result.Tg_hist, dtype=float),
                        Tw=np.asarray(result.Tw_hist, dtype=float),
                        Ti=np.asarray(result.Ti_hist, dtype=float),
                    )
                    zf.writestr("run/fields.npz", buf.getvalue())

                setup = {
                    "run_si": self._last_run_si,
                    "geom": self._last_geom,
                    "mech": self._last_mech,
                    "pipe_material_selected": self._last_pipe_material,
                    "mode_ui": self.mode.currentText(),
                }
                zf.writestr("session/current_setup.json", json.dumps(setup, indent=2))
                zf.writestr("session/stats.json", json.dumps(self._last_stats, indent=2, default=float))
                zf.writestr("session/warnings.txt", "\n".join(self._last_warnings) if self._last_warnings else "")

                opt_summary = getattr(result, "opt_summary", None)
                if isinstance(opt_summary, dict):
                    zf.writestr("session/optimization_summary.json", json.dumps(opt_summary, indent=2, default=float))
                    manifest["optimization_mode"] = str(opt_summary.get("mode", ""))

                if HAS_MPL and hasattr(self, "results_fig"):
                    pbuf = io.BytesIO()
                    self.results_fig.savefig(pbuf, format="png", dpi=180)
                    zf.writestr("session/results_plot.png", pbuf.getvalue())

                ledger_path = self._resolve_ledger_path()
                if ledger_path.exists() and ledger_path.is_file():
                    zf.write(ledger_path, arcname=f"ledger/{ledger_path.name}")
                if self._pipe_preset_path.exists():
                    zf.write(self._pipe_preset_path, arcname=f"library/{self._pipe_preset_path.name}")
                zf.writestr(
                    "library/materials_snapshot.json",
                    json.dumps(
                        {
                            "pipe_materials": PIPE_MATERIALS,
                            "insulation_materials": INSULATION_MATERIALS,
                            "pipe_temp_props": PIPE_TEMP_PROPS,
                            "insulation_temp_props": INSULATION_TEMP_PROPS,
                        },
                        indent=2,
                    ),
                )
                if self._readme_path.exists():
                    zf.write(self._readme_path, arcname="docs/README.md")

                zf.writestr("manifest.json", json.dumps(manifest, indent=2))
        except Exception as exc:
            QMessageBox.critical(self, "Export failed", f"Could not create bundle:\n{exc}")
            return

        QMessageBox.information(self, "Export", f"Bundle written to:\n{zpath}")
