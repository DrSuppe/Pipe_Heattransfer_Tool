##########################################################################
# __author__         = "Tim Kayser"
# __date__           = "27.02.2026"
# __version__        = "2.0"
# __maintainer__     = "Tim Kayser"
# __email__          = "kaysert@purdue.edu"
# __status__         = "Open Beta"
# __copyright__      = "Copyright 2026"
# __credits__        = ["Tim Kayser"]
# __license__        = "GPL"
##########################################################################
"""PyQt6 desktop application for configuring, running, and reviewing simulations."""

from __future__ import annotations

import csv
import io
import json
import logging
import shutil
import sys
import traceback
import zipfile
from dataclasses import replace
from datetime import datetime
from pathlib import Path
from typing import Any, Dict

import numpy as np

from sim_controller import HardwareConfig, RunInputs, RunSpec, run_once
from Pipe_Sim_V4 import HAS_NUMBA

try:
    from PyQt6.QtCore import QObject, QSize, Qt, QThread, QTimer, pyqtSignal, pyqtSlot
    from PyQt6.QtWidgets import (
        QAbstractItemView,
        QAbstractScrollArea,
        QApplication,
        QCheckBox,
        QComboBox,
        QDoubleSpinBox,
        QFileDialog,
        QFormLayout,
        QFrame,
        QGroupBox,
        QHBoxLayout,
        QDialog,
        QLabel,
        QLineEdit,
        QListWidget,
        QMainWindow,
        QMessageBox,
        QPlainTextEdit,
        QProgressBar,
        QPushButton,
        QHeaderView,
        QTableWidget,
        QTableWidgetItem,
        QTextBrowser,
        QScrollArea,
        QSlider,
        QSpinBox,
        QSizePolicy,
        QSplitter,
        QTabWidget,
        QVBoxLayout,
        QWidget,
    )
except Exception as exc:  # pragma: no cover - import guard
    raise SystemExit(
        "PyQt6 is required to run this UI.\n"
        "Install with: pip install PyQt6\n"
        f"Import error: {exc}"
    )

try:
    from matplotlib.backends.backend_qtagg import FigureCanvasQTAgg as FigureCanvas
    from matplotlib.figure import Figure

    HAS_MPL = True
except Exception:
    HAS_MPL = False
    FigureCanvas = None  # type: ignore[assignment]
    Figure = None  # type: ignore[assignment]


PIPE_MATERIALS: Dict[str, Dict[str, float]] = {
    "Carbon Steel": {
        "rho_w": 7850.0,
        "cp_w": 486.0,
        "k_w": 45.0,
        "E": 210.0e9,
        "alpha": 12.0e-6,
        "nu": 0.29,
        "Sy": 250.0e6,
        "eps_default": 0.80,
    },
    "Stainless 316": {
        "rho_w": 8000.0,
        "cp_w": 500.0,
        "k_w": 16.3,
        "E": 193.0e9,
        "alpha": 16.0e-6,
        "nu": 0.30,
        "Sy": 290.0e6,
        "eps_default": 0.70,
    },
    "Copper": {
        "rho_w": 8960.0,
        "cp_w": 385.0,
        "k_w": 401.0,
        "E": 110.0e9,
        "alpha": 16.5e-6,
        "nu": 0.34,
        "Sy": 70.0e6,
        "eps_default": 0.35,
    },
    "Inconel 625": {
        "rho_w": 8440.0,
        "cp_w": 435.0,
        "k_w": 11.0,
        "E": 205.0e9,
        "alpha": 12.8e-6,
        "nu": 0.29,
        "Sy": 460.0e6,
        "eps_default": 0.65,
    },
    "Haynes 282": {
        "rho_w": 8220.0,
        "cp_w": 435.0,
        "k_w": 11.3,
        "E": 217.0e9,
        "alpha": 14.1e-6,
        "nu": 0.30,
        "Sy": 620.0e6,
        "eps_default": 0.65,
    },
    "Hastelloy X": {
        "rho_w": 8220.0,
        "cp_w": 460.0,
        "k_w": 9.2,
        "E": 205.0e9,
        "alpha": 13.2e-6,
        "nu": 0.31,
        "Sy": 355.0e6,
        "eps_default": 0.62,
    },
}

INSULATION_MATERIALS: Dict[str, Dict[str, float]] = {
    "Mineral Wool": {"rho_i": 128.0, "cp_i": 840.0, "k_i": 0.045},
    "Calcium Silicate": {"rho_i": 220.0, "cp_i": 900.0, "k_i": 0.06},
    "Aerogel Blanket": {"rho_i": 150.0, "cp_i": 1000.0, "k_i": 0.02},
}

PIPE_TEMP_PROPS: Dict[str, Dict[str, list[float]]] = {
    # Representative engineering curves for first-estimate work; replace with project-specific datasheets when available.
    "Carbon Steel": {"T": [300.0, 500.0, 700.0, 900.0, 1100.0], "cp": [470.0, 520.0, 580.0, 640.0, 690.0], "k": [47.0, 44.0, 40.0, 36.0, 32.0]},
    "Stainless 316": {"T": [300.0, 500.0, 700.0, 900.0, 1100.0], "cp": [500.0, 530.0, 560.0, 600.0, 640.0], "k": [14.5, 16.0, 17.8, 19.5, 21.2]},
    "Copper": {"T": [300.0, 500.0, 700.0, 900.0, 1100.0], "cp": [385.0, 400.0, 420.0, 440.0, 460.0], "k": [401.0, 378.0, 350.0, 320.0, 285.0]},
    "Inconel 625": {"T": [300.0, 500.0, 700.0, 900.0, 1100.0], "cp": [430.0, 470.0, 510.0, 560.0, 610.0], "k": [9.8, 11.0, 13.0, 15.8, 18.5]},
    "Haynes 282": {"T": [300.0, 500.0, 700.0, 900.0, 1100.0], "cp": [435.0, 475.0, 520.0, 575.0, 635.0], "k": [11.3, 12.8, 14.8, 17.2, 20.2]},
    "Hastelloy X": {"T": [300.0, 500.0, 700.0, 900.0, 1100.0], "cp": [460.0, 495.0, 530.0, 585.0, 640.0], "k": [9.2, 10.6, 12.4, 15.0, 18.0]},
}

INSULATION_TEMP_PROPS: Dict[str, Dict[str, list[float]]] = {
    "Mineral Wool": {"T": [300.0, 500.0, 700.0, 900.0, 1100.0], "cp": [840.0, 900.0, 960.0, 1020.0, 1080.0], "k": [0.045, 0.060, 0.080, 0.105, 0.135]},
    "Calcium Silicate": {"T": [300.0, 500.0, 700.0, 900.0, 1100.0], "cp": [900.0, 950.0, 1000.0, 1060.0, 1120.0], "k": [0.060, 0.075, 0.090, 0.108, 0.128]},
    "Aerogel Blanket": {"T": [300.0, 500.0, 700.0, 900.0, 1100.0], "cp": [1000.0, 1030.0, 1060.0, 1100.0, 1140.0], "k": [0.020, 0.024, 0.029, 0.035, 0.043]},
}

TARGET_METRIC_CHOICES: list[tuple[str, str]] = [
    ("Gas outlet Tg", "gas_outlet"),
    ("Wall inner outlet Tw(in)", "wall_inner_outlet"),
    ("Wall outer outlet Tw(out)", "wall_outer_outlet"),
    ("Insulation outlet Ti", "insulation_outlet"),
]
TARGET_METRIC_LABELS: Dict[str, str] = {key: label for label, key in TARGET_METRIC_CHOICES}


FT2M = 0.3048
IN2M = 0.0254
M2FT = 1.0 / FT2M
M2IN = 1.0 / IN2M
PSI2PA = 6894.757293168
PA2PSI = 1.0 / PSI2PA
KG_S__TO__LBM_S = 2.20462262185
LBM_S__TO__KG_S = 1.0 / KG_S__TO__LBM_S
MPA_TO_KSI = 0.1450377377
KSI_TO_MPA = 1.0 / MPA_TO_KSI


def _sanitize_target_metric(metric: str) -> str:
    key = str(metric or "").strip().lower()
    valid = {k for _label, k in TARGET_METRIC_CHOICES}
    return key if key in valid else "gas_outlet"


def _wall_fraction_from_geom(geom: Dict[str, Any]) -> float:
    di = max(float(geom.get("Di", 0.13)), 1.0e-9)
    t_wall = max(float(geom.get("t_wall", 0.018)), 1.0e-9)
    t_ins = max(float(geom.get("t_ins", 0.0)), 0.0)
    k_w = max(float(geom.get("k_w", 15.0)), 1.0e-12)
    k_i = max(float(geom.get("k_i", 0.05)), 1.0e-12)
    ri = 0.5 * di
    ro = ri + t_wall
    r_ins = ro + t_ins
    r_wall = np.log(max(ro / ri, 1.0 + 1.0e-12)) / (2.0 * np.pi * k_w)
    if t_ins > 1.0e-12:
        r_insu = np.log(max(r_ins / ro, 1.0 + 1.0e-12)) / (2.0 * np.pi * k_i)
    else:
        r_insu = 0.0
    return float(r_wall / max(r_wall + r_insu, 1.0e-12))


def _outlet_series_for_metric(
    tg_hist: np.ndarray,
    tw_hist: np.ndarray,
    ti_hist: np.ndarray,
    metric: str,
    geom: Dict[str, Any],
) -> np.ndarray:
    key = _sanitize_target_metric(metric)
    tg_out = np.asarray(tg_hist[:, -1], dtype=float)
    tw_out = np.asarray(tw_hist[:, -1], dtype=float)
    ti_out = np.asarray(ti_hist[:, -1], dtype=float)
    if key == "gas_outlet":
        return tg_out
    if key == "wall_inner_outlet":
        return tw_out
    if key == "insulation_outlet":
        return ti_out
    wall_frac = _wall_fraction_from_geom(geom)
    return tw_out - wall_frac * (tw_out - ti_out)


def K_to_F(value_k: float) -> float:
    return (value_k - 273.15) * 9.0 / 5.0 + 32.0


def F_to_K(value_f: float) -> float:
    return (value_f - 32.0) * 5.0 / 9.0 + 273.15


def _target_crossing_time_series(times: np.ndarray, values: np.ndarray, target: float) -> float | None:
    if values.size == 0 or times.size == 0:
        return None
    mode_le = bool(target <= float(values[0]))
    cond = values <= target if mode_le else values >= target
    hit = np.where(cond)[0]
    if hit.size == 0:
        return None
    i = int(hit[0])
    if i == 0:
        return float(times[0])
    t0, t1 = float(times[i - 1]), float(times[i])
    y0, y1 = float(values[i - 1]), float(values[i])
    if abs(y1 - y0) <= 1.0e-12:
        return t1
    frac = (target - y0) / (y1 - y0)
    frac = max(0.0, min(1.0, frac))
    return t0 + frac * (t1 - t0)


def _estimate_vm_total_max_mpa(
    tw_si: np.ndarray,
    ti_si: np.ndarray,
    *,
    length_m: float,
    geom: Dict[str, Any],
    run_si: Dict[str, Any],
    mech: Dict[str, Any],
    include_pressure: bool,
    nr_wall: int,
) -> float:
    nr = int(max(3, nr_wall))
    E = float(mech.get("E", 210.0e9))
    alpha = float(mech.get("alpha", 12.0e-6))
    nu = float(mech.get("nu", 0.30))

    Di_m = float(geom.get("Di", 0.1))
    t_wall_m = float(geom.get("t_wall", 0.01))
    t_ins_m = float(geom.get("t_ins", 0.0))
    k_w = float(geom.get("k_w", 15.0))
    k_i = float(geom.get("k_i", 0.04))
    n_elbows = int(geom.get("n_elbows", 0))
    elbow_sif = float(geom.get("elbow_sif", 1.0))
    elbow_positions_frac = list(geom.get("elbow_positions_frac", []))

    p_si = float(run_si.get("p", 0.0))
    Tamb = float(run_si.get("Tamb", 300.0))
    axial_restraint = float(run_si.get("axial_restraint", 0.0))

    ri = max(0.5 * Di_m, 1.0e-9)
    ro = ri + max(t_wall_m, 1.0e-9)
    nx = max(1, tw_si.shape[1])
    dx_si = max(length_m, 1.0e-12) / max(1, nx - 1)

    R_wall = np.log(max(ro / ri, 1.0 + 1.0e-12)) / (2.0 * np.pi * max(k_w, 1.0e-12) * max(dx_si, 1.0e-12))
    if t_ins_m > 1.0e-12:
        r_ins_o = ro + t_ins_m
        R_ins = np.log(max(r_ins_o / ro, 1.0 + 1.0e-12)) / (2.0 * np.pi * max(k_i, 1.0e-12) * max(dx_si, 1.0e-12))
    else:
        R_ins = 0.0
    wall_frac = float(R_wall / max(R_wall + R_ins, 1.0e-12))

    deltaT_wall_si = np.abs(tw_si - ti_si) * wall_frac
    xi = np.linspace(0.0, 1.0, nr, dtype=float)
    r = ri + xi * max(t_wall_m, 1.0e-9)
    grad_term = (0.5 - xi)[:, None, None] * deltaT_wall_si[None, :, :]
    temp_r = tw_si[None, :, :] + grad_term

    thermo_coeff = E * alpha / max(1.0e-9, (1.0 - nu))
    sigma_theta_th = thermo_coeff * (tw_si[None, :, :] - temp_r)
    sigma_z_th = -axial_restraint * thermo_coeff * (tw_si - Tamb)[None, :, :]
    sigma_r_th = np.zeros_like(sigma_theta_th)

    if include_pressure:
        denom = max(ro * ro - ri * ri, 1.0e-12)
        r2 = np.maximum(r * r, 1.0e-12)[:, None, None]
        coeff = p_si * ri * ri / denom
        sigma_r_p = coeff * (1.0 - (ro * ro) / r2)
        sigma_theta_p = coeff * (1.0 + (ro * ro) / r2)
        sigma_z_p = np.full_like(sigma_r_p, coeff)
    else:
        sigma_r_p = np.zeros_like(temp_r)
        sigma_theta_p = np.zeros_like(temp_r)
        sigma_z_p = np.zeros_like(temp_r)

    sigma_theta = sigma_theta_th + sigma_theta_p
    sigma_r = sigma_r_th + sigma_r_p
    sigma_z = sigma_z_th + sigma_z_p
    sigma_vm_total = np.sqrt(
        0.5
        * (
            (sigma_theta - sigma_z) ** 2
            + (sigma_z - sigma_r) ** 2
            + (sigma_r - sigma_theta) ** 2
        )
    )
    vm_total_mpa = np.max(sigma_vm_total, axis=0) / 1.0e6

    elbow_profile = np.ones(nx, dtype=float)
    if n_elbows > 0 and nx > 2:
        if len(elbow_positions_frac) != n_elbows:
            elbow_positions_frac = [float(v) for v in np.linspace(0.15, 0.85, n_elbows)]
        spread = max(1, int(0.03 * nx))
        for frac in elbow_positions_frac:
            j0 = int(round(max(0.0, min(1.0, float(frac))) * (nx - 1)))
            for dj in range(-3 * spread, 3 * spread + 1):
                j = j0 + dj
                if j < 0 or j >= nx:
                    continue
                w = max(0.0, 1.0 - abs(dj) / (3.0 * spread + 1.0e-9))
                elbow_profile[j] = max(elbow_profile[j], 1.0 + (elbow_sif - 1.0) * w)

    vm_total_elbow_mpa = vm_total_mpa * elbow_profile[None, :]
    return float(np.nanmax(vm_total_elbow_mpa))


class LogEmitter(QObject):
    message = pyqtSignal(str)


class QtLogHandler(logging.Handler):
    def __init__(self, emitter: LogEmitter):
        super().__init__()
        self.emitter = emitter
        self._thermal_pipe_qt_handler = True

    def emit(self, record):
        self.emitter.message.emit(self.format(record))


class CompactTableWidget(QTableWidget):
    """Table widget with conservative size hints to avoid parent window auto-expansion."""

    def minimumSizeHint(self):
        return QSize(280, 180)

    def sizeHint(self):
        return QSize(560, 360)


class UserCancelledError(RuntimeError):
    """Signal user-requested cancellation without surfacing an error dialog."""


class SimulationWorker(QObject):
    finished = pyqtSignal(object)
    failed = pyqtSignal(str)
    cancelled = pyqtSignal()
    snapshot = pyqtSignal(float, object, object, object)

    def __init__(self, spec: RunSpec, optimization: Dict[str, Any] | None = None):
        super().__init__()
        self.spec = spec
        self.optimization = optimization or None
        self._cancel_requested = False

    def request_cancel(self):
        self._cancel_requested = True

    def _abort_requested(self) -> bool:
        return bool(self._cancel_requested)

    def _check_cancel(self):
        if self._cancel_requested:
            raise UserCancelledError("Simulation cancelled by user.")

    @pyqtSlot()
    def run(self):
        try:
            if self.optimization:
                result = self._run_optimization()
            else:
                spec = replace(
                    self.spec,
                    snapshot_callback=self._emit_snapshot,
                    abort_callback=self._abort_requested,
                )
                result = run_once(spec)
            if str(getattr(result, "stop_reason", "")) == "aborted_by_user":
                self.finished.emit(result)
                return
            self.finished.emit(result)
        except UserCancelledError:
            self.cancelled.emit()
        except Exception:
            self.failed.emit(traceback.format_exc())

    def _emit_snapshot(self, t_s: float, tg, tw, ti):
        self.snapshot.emit(float(t_s), tg, tw, ti)

    def _target_value(self, opt: Dict[str, Any]) -> float:
        return float(opt.get("target_temp_K", opt.get("target_Tg_K", np.nan)))

    def _target_metric(self, opt: Dict[str, Any]) -> str:
        return _sanitize_target_metric(str(opt.get("target_metric", "gas_outlet")))

    def _candidate_result(self, m_dot_kg_s: float, nx: int, opt: Dict[str, Any]) -> Dict[str, Any]:
        self._check_cancel()
        run = replace(
            self.spec.run,
            mode="target",
            m_dot=float(m_dot_kg_s),
            Tg_out_target=self._target_value(opt),
            t_end=float(opt.get("opt_t_end_s", self.spec.run.t_end)),
            stop_dir=None,
        )
        overrides = dict(self.spec.overrides or {})
        overrides.update(
            {
                "Nx": int(nx),
                "t_end": float(opt.get("opt_t_end_s", self.spec.run.t_end)),
                "save_frames": int(max(20, min(120, int(overrides.get("save_frames", 120))))),
                "progress": "none",
                "log_to_file": False,
                "write_trace_csv": False,
                "show_plots": False,
            }
        )
        spec = RunSpec(
            hardware=self.spec.hardware,
            run=run,
            overrides=overrides,
            save_dir=None,
            make_plots=False,
            save_results=False,
            snapshot_callback=None,
            abort_callback=self._abort_requested,
        )
        r = run_once(spec)
        if str(getattr(r, "stop_reason", "")) == "aborted_by_user":
            raise UserCancelledError("Simulation cancelled by user.")
        target_series = _outlet_series_for_metric(
            np.asarray(r.Tg_hist, dtype=float),
            np.asarray(r.Tw_hist, dtype=float),
            np.asarray(r.Ti_hist, dtype=float),
            self._target_metric(opt),
            opt.get("geom", {}),
        )
        t_hit = _target_crossing_time_series(np.asarray(r.times, dtype=float), target_series, self._target_value(opt))
        sigma_max = _estimate_vm_total_max_mpa(
            np.asarray(r.Tw_hist, dtype=float),
            np.asarray(r.Ti_hist, dtype=float),
            length_m=float(self.spec.hardware.L),
            geom=opt["geom"],
            run_si=opt["run_si"],
            mech=opt["mech"],
            include_pressure=bool(opt["include_pressure"]),
            nr_wall=int(opt["nr_wall"]),
        )
        return {
            "m_dot": float(m_dot_kg_s),
            "reached": bool(r.reached_Tg_target),
            "t_hit_s": t_hit,
            "sigma_max_mpa": float(sigma_max),
        }

    @staticmethod
    def _heatup_score(
        t_hit_s: float | None,
        sigma_mpa: float,
        target_t_s: float,
        tol_t_s: float,
        sigma_lim_mpa: float,
    ) -> tuple[float, float, float]:
        stress_violation = max(0.0, float(sigma_mpa) - float(sigma_lim_mpa))
        if t_hit_s is None:
            return (stress_violation, 1.0e9, 1.0e9)
        abs_err = abs(float(t_hit_s) - float(target_t_s))
        time_violation = max(0.0, abs_err - float(tol_t_s))
        return (stress_violation, time_violation, abs_err)

    def _full_nx_for_opt(self, opt: Dict[str, Any]) -> int:
        nx_from_overrides = None
        if isinstance(self.spec.overrides, dict):
            try:
                nx_from_overrides = int(self.spec.overrides.get("Nx", 0))
            except Exception:
                nx_from_overrides = None
        if nx_from_overrides is None or nx_from_overrides <= 0:
            nx_from_overrides = int(max(120, opt.get("search_nx", 120)))
        return int(max(80, nx_from_overrides))

    def _run_final_candidate(self, m_dot_kg_s: float, opt: Dict[str, Any]):
        self._check_cancel()
        run_final = replace(
            self.spec.run,
            mode="target",
            m_dot=float(m_dot_kg_s),
            Tg_out_target=self._target_value(opt),
            t_end=float(opt.get("opt_t_end_s", self.spec.run.t_end)),
            stop_dir=None,
        )
        overrides_final = dict(self.spec.overrides or {})
        overrides_final["t_end"] = float(run_final.t_end)
        final_spec = replace(
            self.spec,
            run=run_final,
            overrides=overrides_final,
            snapshot_callback=self._emit_snapshot,
            abort_callback=self._abort_requested,
        )
        result = run_once(final_spec)
        if str(getattr(result, "stop_reason", "")) == "aborted_by_user":
            raise UserCancelledError("Simulation cancelled by user.")

        target_series = _outlet_series_for_metric(
            np.asarray(result.Tg_hist, dtype=float),
            np.asarray(result.Tw_hist, dtype=float),
            np.asarray(result.Ti_hist, dtype=float),
            self._target_metric(opt),
            opt.get("geom", {}),
        )
        t_hit_final = _target_crossing_time_series(np.asarray(result.times, dtype=float), target_series, self._target_value(opt))
        sigma_final_mpa = _estimate_vm_total_max_mpa(
            np.asarray(result.Tw_hist, dtype=float),
            np.asarray(result.Ti_hist, dtype=float),
            length_m=float(self.spec.hardware.L),
            geom=opt["geom"],
            run_si=opt["run_si"],
            mech=opt["mech"],
            include_pressure=bool(opt["include_pressure"]),
            nr_wall=int(opt["nr_wall"]),
        )
        return result, t_hit_final, float(sigma_final_mpa)

    def _run_heatup_time_opt(self, opt: Dict[str, Any]) -> Dict[str, Any]:
        self._check_cancel()
        md_min = float(opt["mdot_min_kg_s"])
        md_max = float(opt["mdot_max_kg_s"])
        target_t = float(opt["heatup_target_s"])
        tol_t = float(opt["heatup_tol_s"])
        sigma_lim = float(opt["stress_limit_mpa"])
        nx_search = int(opt["search_nx"])
        coarse_points = int(max(5, opt.get("coarse_points", 9)))
        refine_iters = int(max(2, opt.get("refine_iters", 6)))

        cache: Dict[float, Dict[str, Any]] = {}

        def evaluate(m: float) -> Dict[str, Any]:
            m_key = float(np.round(m, 10))
            if m_key not in cache:
                cache[m_key] = self._candidate_result(m_key, nx_search, opt)
            return cache[m_key]

        def score(c: Dict[str, Any]) -> tuple[float, float, float]:
            t_hit = c["t_hit_s"]
            stress_violation = max(0.0, float(c["sigma_max_mpa"]) - sigma_lim)
            if t_hit is None:
                time_violation = 1.0e9
                abs_time_err = 1.0e9
            else:
                abs_time_err = abs(float(t_hit) - target_t)
                time_violation = max(0.0, abs_time_err - tol_t)
            return (stress_violation, time_violation, abs_time_err)

        grid = np.linspace(md_min, md_max, coarse_points)
        coarse = [evaluate(float(m)) for m in grid]
        best = min(coarse, key=score)
        best_score = score(best)

        idx_best = int(np.argmin([score(c) for c in coarse]))
        left = float(grid[max(0, idx_best - 1)])
        right = float(grid[min(len(grid) - 1, idx_best + 1)])
        if right <= left:
            left, right = md_min, md_max

        for _ in range(refine_iters):
            self._check_cancel()
            pts = np.linspace(left, right, 5)
            local = [evaluate(float(m)) for m in pts]
            local_best = min(local, key=score)
            local_score = score(local_best)
            if local_score < best_score:
                best, best_score = local_best, local_score

            if (
                local_best["t_hit_s"] is not None
                and abs(float(local_best["t_hit_s"]) - target_t) <= tol_t
                and float(local_best["sigma_max_mpa"]) <= sigma_lim
            ):
                best = local_best
                break

            i_local = int(np.argmin([score(c) for c in local]))
            l_idx = max(0, i_local - 1)
            r_idx = min(len(pts) - 1, i_local + 1)
            new_left = float(pts[l_idx])
            new_right = float(pts[r_idx])
            if new_right - new_left <= 1.0e-9:
                break
            left, right = new_left, new_right

        return best

    def _run_stress_limit_opt(self, opt: Dict[str, Any]) -> Dict[str, Any]:
        self._check_cancel()
        md_min = float(opt["mdot_min_kg_s"])
        md_max = float(opt["mdot_max_kg_s"])
        sigma_lim = float(opt["stress_limit_mpa"])
        sigma_tol = float(opt.get("stress_tol_mpa", 2.0))
        nx_search = int(opt["search_nx"])
        max_iter = int(max(4, opt.get("bisection_iters", 10)))

        def evaluate(m: float) -> Dict[str, Any]:
            return self._candidate_result(float(m), nx_search, opt)

        lo = evaluate(md_min)
        hi = evaluate(md_max)
        f_lo = float(lo["sigma_max_mpa"]) - sigma_lim
        f_hi = float(hi["sigma_max_mpa"]) - sigma_lim

        if f_lo <= 0.0 and f_hi <= 0.0:
            return hi
        if f_lo > 0.0 and f_hi > 0.0:
            return lo

        l_m, h_m = md_min, md_max
        c_best = lo if abs(f_lo) <= abs(f_hi) else hi
        for _ in range(max_iter):
            self._check_cancel()
            mid = 0.5 * (l_m + h_m)
            c_mid = evaluate(mid)
            f_mid = float(c_mid["sigma_max_mpa"]) - sigma_lim
            if abs(f_mid) < abs(float(c_best["sigma_max_mpa"]) - sigma_lim):
                c_best = c_mid
            if abs(f_mid) <= sigma_tol:
                c_best = c_mid
                break
            if f_lo * f_mid <= 0.0:
                h_m = mid
                hi = c_mid
                f_hi = f_mid
            else:
                l_m = mid
                lo = c_mid
                f_lo = f_mid
            if h_m - l_m <= 1.0e-9:
                break
        return c_best

    def _run_optimization(self):
        self._check_cancel()
        opt = dict(self.optimization or {})
        mode = str(opt.get("mode", ""))
        if mode not in ("heatup_time_opt", "stress_limit_opt"):
            raise ValueError(f"Unsupported optimization mode: {mode}")

        if mode == "heatup_time_opt":
            best = self._run_heatup_time_opt(opt)
        else:
            best = self._run_stress_limit_opt(opt)

        mdot_best = float(best["m_dot"])
        result, t_hit_final, sigma_final_mpa = self._run_final_candidate(mdot_best, opt)

        sigma_lim = float(opt["stress_limit_mpa"])
        nx_full = self._full_nx_for_opt(opt)
        correction_applied = False

        if mode == "heatup_time_opt":
            tgt = float(opt.get("heatup_target_s", np.nan))
            tol = float(opt.get("heatup_tol_s", np.nan))
            cur_score = self._heatup_score(t_hit_final, sigma_final_mpa, tgt, tol, sigma_lim)
            if cur_score[0] > 0.0 or cur_score[1] > 0.0:
                md_min = float(opt["mdot_min_kg_s"])
                md_max = float(opt["mdot_max_kg_s"])
                span = max(md_max - md_min, 1.0e-8)
                local_half = max(0.2 * span, 0.25 * mdot_best)
                left = max(md_min, mdot_best - local_half)
                right = min(md_max, mdot_best + local_half)
                probes = np.unique(
                    np.concatenate(
                        (
                            np.array([md_min, left, mdot_best, right, md_max], dtype=float),
                            np.linspace(left, right, 7),
                        )
                    )
                )
                candidates = [self._candidate_result(float(m), nx_full, opt) for m in probes]
                best_full = min(
                    candidates,
                    key=lambda c: self._heatup_score(
                        c.get("t_hit_s"),
                        float(c.get("sigma_max_mpa", np.inf)),
                        tgt,
                        tol,
                        sigma_lim,
                    ),
                )
                best_full_score = self._heatup_score(
                    best_full.get("t_hit_s"),
                    float(best_full.get("sigma_max_mpa", np.inf)),
                    tgt,
                    tol,
                    sigma_lim,
                )
                if best_full_score < cur_score and abs(float(best_full["m_dot"]) - mdot_best) > 1.0e-12:
                    mdot_best = float(best_full["m_dot"])
                    result, t_hit_final, sigma_final_mpa = self._run_final_candidate(mdot_best, opt)
                    correction_applied = True
        else:
            md_min = float(opt["mdot_min_kg_s"])
            md_max = float(opt["mdot_max_kg_s"])
            sigma_tol = float(opt.get("stress_tol_mpa", 2.0))
            if sigma_final_mpa > sigma_lim + sigma_tol:
                probes = np.linspace(md_min, max(md_min, mdot_best), 9)
                candidates = [self._candidate_result(float(m), nx_full, opt) for m in probes]
                passing = [c for c in candidates if float(c.get("sigma_max_mpa", np.inf)) <= sigma_lim + sigma_tol]
                if passing:
                    c_pick = max(passing, key=lambda c: float(c["m_dot"]))
                    if abs(float(c_pick["m_dot"]) - mdot_best) > 1.0e-12:
                        mdot_best = float(c_pick["m_dot"])
                        result, t_hit_final, sigma_final_mpa = self._run_final_candidate(mdot_best, opt)
                        correction_applied = True
            elif sigma_final_mpa <= sigma_lim:
                probes = np.linspace(min(mdot_best, md_max), md_max, 7)
                candidates = [self._candidate_result(float(m), nx_full, opt) for m in probes]
                passing = [c for c in candidates if float(c.get("sigma_max_mpa", np.inf)) <= sigma_lim + sigma_tol]
                if passing:
                    c_pick = max(passing, key=lambda c: float(c["m_dot"]))
                    if abs(float(c_pick["m_dot"]) - mdot_best) > 1.0e-12:
                        mdot_best = float(c_pick["m_dot"])
                        result, t_hit_final, sigma_final_mpa = self._run_final_candidate(mdot_best, opt)
                        correction_applied = True

        meets_stress = bool(sigma_final_mpa <= sigma_lim)
        meets_heatup = True
        if mode == "heatup_time_opt":
            tgt = float(opt.get("heatup_target_s", np.nan))
            tol = float(opt.get("heatup_tol_s", np.nan))
            meets_heatup = bool(t_hit_final is not None and abs(float(t_hit_final) - tgt) <= tol)

        logging.info(
            "Optimization summary: mode=%s, m_dot=%.5f kg/s, t_hit_final=%s, sigma_final=%.2f MPa, "
            "meets_stress=%s, meets_heatup=%s, correction=%s",
            mode,
            mdot_best,
            "n/a" if t_hit_final is None else f"{float(t_hit_final):.2f}s",
            sigma_final_mpa,
            meets_stress,
            meets_heatup,
            correction_applied,
        )
        result.opt_summary = {
            "mode": mode,
            "m_dot_kg_s": mdot_best,
            "target_metric": self._target_metric(opt),
            "target_temp_K": self._target_value(opt),
            "sigma_est_search_mpa": float(best["sigma_max_mpa"]),
            "sigma_final_mpa": float(sigma_final_mpa),
            "time_to_target_search_s": None if best["t_hit_s"] is None else float(best["t_hit_s"]),
            "time_to_target_final_s": None if t_hit_final is None else float(t_hit_final),
            "stress_limit_mpa": float(opt["stress_limit_mpa"]),
            "heatup_target_s": float(opt.get("heatup_target_s", np.nan)),
            "heatup_tol_s": float(opt.get("heatup_tol_s", np.nan)),
            "meets_stress_limit": meets_stress,
            "meets_heatup_tolerance": meets_heatup,
            "target_reached_final": bool(result.reached_Tg_target),
            "correction_applied": correction_applied,
            "search_nx": int(opt.get("search_nx", 0)),
            "full_nx": int(nx_full),
        }
        return result


class ThermalPipeWindow(QMainWindow):
    PRESETS: Dict[str, Dict[str, Any]] = {
        "Fast estimate": {
            "Nx": 450, "save_frames": 120, "dt_max": 4.0, "dt_min": 1e-4, "update_props_every": 10,
            "adv_scheme": "semi_lagrangian", "semi_lag_courant_max": 2.2, "use_float32": True,
        },
        "Balanced": {
            "Nx": 1300, "save_frames": 240, "dt_max": 2.5, "dt_min": 1e-4, "update_props_every": 5,
            "adv_scheme": "semi_lagrangian", "semi_lag_courant_max": 1.2, "use_float32": True,
        },
        "High fidelity": {
            "Nx": 2600, "save_frames": 320, "dt_max": 1.0, "dt_min": 1e-4, "update_props_every": 2,
            "adv_scheme": "upwind", "semi_lag_courant_max": 1.0, "use_float32": False,
        },
    }
    PIPE_DEFAULTS: Dict[str, Dict[str, Any]] = {
        "Custom / current": {},
        "Transfer Line (Insulated SS316)": {
            "pipe_material": "Stainless 316",
            "insulation": True,
            "ins_material": "Mineral Wool",
            "L": 65.0,
            "Di": 0.13,
            "t_wall": 0.018,
            "t_ins": 0.150,
            "p": 5.0e6,
            "Tin": 1000.0,
            "m_dot": 1.0,
            "Tamb": 300.0,
            "n_elbows": 0,
            "elbow_sif": 1.30,
        },
        "High-Temp Alloy (Inconel 625)": {
            "pipe_material": "Inconel 625",
            "insulation": True,
            "ins_material": "Calcium Silicate",
            "L": 45.0,
            "Di": 0.10,
            "t_wall": 0.015,
            "t_ins": 0.120,
            "p": 6.0e6,
            "Tin": 1100.0,
            "m_dot": 1.2,
            "Tamb": 320.0,
            "n_elbows": 2,
            "elbow_sif": 1.50,
            "elbow_positions": "22%, 78%",
        },
        "Bare Carbon Steel Startup": {
            "pipe_material": "Carbon Steel",
            "insulation": False,
            "L": 30.0,
            "Di": 0.09,
            "t_wall": 0.010,
            "t_ins": 0.0,
            "p": 2.0e6,
            "Tin": 700.0,
            "m_dot": 0.8,
            "Tamb": 295.0,
            "n_elbows": 1,
            "elbow_sif": 1.30,
            "elbow_positions": "60%",
        },
    }

    def __init__(self):
        super().__init__()
        self.setWindowTitle("Thermal Pipe Simulator")
        screen = QApplication.primaryScreen()
        if screen is not None:
            ag = screen.availableGeometry()
            w = max(1080, min(1500, ag.width() - 40))
            h = max(760, min(960, ag.height() - 60))
            self.resize(w, h)
        else:
            self.resize(1400, 900)

        self._save_dir: Path | None = None
        self._thread: QThread | None = None
        self._worker: SimulationWorker | None = None
        self._units = "SI"
        self._current_L = 65.0
        self._ambient_temp = 300.0
        self._last_mech = {"E": 210e9, "alpha": 12e-6, "nu": 0.29, "Sy": 250e6}
        self._last_pipe_material = "Carbon Steel"

        self._snap_t: list[float] = []
        self._snap_outlet: list[float] = []
        self._snap_tw_out: list[float] = []
        self._snap_ti_out: list[float] = []
        self._snap_inlet: list[float] = []
        self._snap_inlet_cell: list[float] = []
        self._snap_tg_rows: list[np.ndarray] = []
        self._snapshot_counter = 0

        self._play_times = np.array([], dtype=float)
        self._play_tg = np.empty((0, 0), dtype=float)
        self._play_tw = np.empty((0, 0), dtype=float)
        self._play_ti = np.empty((0, 0), dtype=float)
        self._play_tin_eff = np.array([], dtype=float)
        self._last_result = None
        self._last_geom: Dict[str, float] = {}
        self._last_run_si: Dict[str, Any] = {}
        self._last_stats: Dict[str, Any] = {}
        self._last_warnings: list[str] = []
        self._ledger_path: Path | None = None
        self._readme_path = Path(__file__).with_name("README.md")
        self._pipe_preset_path = Path(__file__).with_name("pipe_presets.json")
        self._custom_pipe_presets: Dict[str, Dict[str, Any]] = {}
        self._heat_im = None
        self._heat_cbar = None
        self._results_axes: list[Any] = []
        self._results_cbar_parent: Dict[Any, Any] = {}
        self._cancel_pending = False

        self._play_timer = QTimer(self)
        self._play_timer.setInterval(50)
        self._play_timer.timeout.connect(self._playback_tick)

        root = QWidget(self)
        self.setCentralWidget(root)
        root_layout = QVBoxLayout(root)
        root_layout.setContentsMargins(10, 10, 10, 10)
        root_layout.setSpacing(8)

        splitter = QSplitter(Qt.Orientation.Horizontal)
        self._left_panel = self._build_left_panel()
        self._left_panel.setMinimumWidth(350)
        self._left_panel.setMaximumWidth(760)
        splitter.addWidget(self._left_panel)
        splitter.addWidget(self._build_right_panel())
        splitter.setCollapsible(0, False)
        splitter.setStretchFactor(0, 0)
        splitter.setStretchFactor(1, 1)
        splitter.setSizes([460, 980])
        root_layout.addWidget(splitter, 1)

        root_layout.addLayout(self._build_run_bar())

        self._setup_logging()
        self._load_custom_pipe_presets()
        self._refresh_pipe_default_combo()
        self._apply_preset("Fast estimate")
        self._on_pipe_material_changed(self.pipe_material.currentText())
        self._sync_mode_widgets()
        self._sync_insulation_widgets()
        self._refresh_material_lists()
        self._refresh_temp_prop_materials()
        self._refresh_pipe_preset_list()
        self._apply_unit_ranges()
        self._apply_unit_labels()
        self._update_tmass_estimate_label()
        self._reset_live_views()
        self._refresh_readme_view()
        self._refresh_ledger_preview()

    def _build_left_panel(self):
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOff)
        scroll.horizontalScrollBar().setEnabled(False)
        scroll.setSizeAdjustPolicy(QAbstractScrollArea.SizeAdjustPolicy.AdjustIgnored)

        panel = QWidget()
        panel.setMinimumWidth(0)
        panel.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Preferred)
        layout = QVBoxLayout(panel)
        layout.setContentsMargins(8, 8, 8, 8)
        layout.setSpacing(8)

        layout.addWidget(self._build_scenario_group())
        layout.addWidget(self._build_material_group())
        layout.addWidget(self._build_pipe_group())
        layout.addWidget(self._build_flow_group())
        layout.addWidget(self._build_output_group())
        layout.addWidget(self._build_advanced_group())
        layout.addStretch(1)

        scroll.setWidget(panel)
        return scroll

    def _build_right_panel(self):
        panel = QWidget()
        panel.setMinimumWidth(0)
        panel.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding)
        layout = QVBoxLayout(panel)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(8)

        layout.addLayout(self._build_metrics_row())

        self.tabs = QTabWidget()
        self.tabs.setMinimumWidth(0)
        self.tabs.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding)
        self.tabs.addTab(self._build_live_tab(), "Live View")
        self.tabs.addTab(self._build_results_tab(), "Results")
        self.tabs.addTab(self._build_material_library_tab(), "Library")
        self.tabs.addTab(self._build_ledger_tab(), "Ledger")
        self.tabs.addTab(self._build_readme_tab(), "README")
        self.tabs.addTab(self._build_log_tab(), "Simulation Log")
        layout.addWidget(self.tabs, 1)
        return panel

    def _build_metrics_row(self):
        row = QHBoxLayout()
        row.setSpacing(10)
        self.lbl_runtime = QLabel("Sim time: 0.0 s")
        self.lbl_target_time = QLabel("Target time: --")
        self.lbl_target_time.setStyleSheet("color: #b8d3ff;")
        self.lbl_frames = QLabel("Frames: 0")
        self.lbl_live_readout = QLabel("Time: -- s || Tg_out -- | Tw_out -- | Ti_out -- | Tin_eff -- | Tg_in -- K")
        self.lbl_live_readout.setStyleSheet("color: #d0d0d0;")
        # Keep individual labels for compatibility with existing update paths.
        self.lbl_outlet = QLabel("Outlet Tg: -- K")
        self.lbl_wall_out = QLabel("Outlet Tw: -- K")
        self.lbl_ins_out = QLabel("Outlet Ti: -- K")
        self.lbl_inlet = QLabel("Tin_eff(bc): -- K")
        self.lbl_inlet_cell = QLabel("Tg_in(c0): -- K")
        self.lbl_mode_hint = QLabel("Workflow: configure -> run")
        self.lbl_mode_hint.setStyleSheet("color: #808080;")
        self.lbl_mode_hint.setMaximumWidth(210)
        for lbl in (
            self.lbl_runtime,
            self.lbl_target_time,
            self.lbl_frames,
            self.lbl_live_readout,
            self.lbl_mode_hint,
        ):
            lbl.setMinimumWidth(0)
            lbl.setSizePolicy(QSizePolicy.Policy.Ignored, QSizePolicy.Policy.Preferred)
        row.addWidget(self.lbl_runtime)
        row.addWidget(self.lbl_target_time)
        row.addWidget(self.lbl_frames)
        row.addWidget(self.lbl_live_readout, 1)
        row.addWidget(self.lbl_mode_hint)
        return row

    def _build_scenario_group(self):
        g = QGroupBox("1) Scenario")
        form = QFormLayout(g)

        self.preset = QComboBox()
        self.preset.addItems(list(self.PRESETS))
        self.preset.setCurrentText("Fast estimate")
        self.preset.currentTextChanged.connect(self._apply_preset)

        self.pipe_default = QComboBox()
        self.pipe_default.addItems(list(self.PIPE_DEFAULTS))
        btn_apply_pipe_default = QPushButton("Apply Pipe Default")
        btn_apply_pipe_default.clicked.connect(self._apply_pipe_default)

        self.units = QComboBox()
        self.units.addItems(["SI", "Imperial"])
        self.units.currentTextChanged.connect(self._on_units_changed)

        self.asset_id = QLineEdit("PIPE-001")
        self.branch_id = QLineEdit("main")

        self.mode = QComboBox()
        self.mode.addItems(
            [
                "Fixed time",
                "Outlet target",
                "Heatup-time optimize",
                "Stress-limit optimize",
            ]
        )
        self.mode.currentIndexChanged.connect(self._sync_mode_widgets)

        self.in_t_end = self._dbl(5000.0, 0.001, 1.0e8, 3, " s")
        self.in_target = self._dbl(950.0, 1.0, 5000.0, 2, " K")
        self.target_metric = QComboBox()
        for label, key in TARGET_METRIC_CHOICES:
            self.target_metric.addItem(label, key)
        self.target_metric.setCurrentIndex(0)
        self.stop_dir = QComboBox()
        self.stop_dir.addItems(["auto", "le", "ge"])
        self.in_heatup_target = self._dbl(1800.0, 1.0, 1.0e8, 1, " s")
        self.in_heatup_tol = self._dbl(20.0, 1.0, 1.0e5, 1, " s")
        self.in_stress_limit = self._dbl(250.0, 1.0, 1.0e5, 2, " MPa")

        form.addRow("Units", self.units)
        form.addRow("Performance mode", self.preset)
        form.addRow("Pipe default", self.pipe_default)
        form.addRow("", btn_apply_pipe_default)
        form.addRow("Asset ID", self.asset_id)
        form.addRow("Branch ID", self.branch_id)
        form.addRow("Run mode", self.mode)
        form.addRow("t_end / t_max", self.in_t_end)
        form.addRow("Outlet target", self.in_target)
        form.addRow("Target variable", self.target_metric)
        form.addRow("Stop direction", self.stop_dir)
        form.addRow("Heatup target time", self.in_heatup_target)
        form.addRow("Heatup tolerance", self.in_heatup_tol)
        form.addRow("Stress limit", self.in_stress_limit)
        return g

    def _build_material_group(self):
        g = QGroupBox("2) Materials / Ambient")
        form = QFormLayout(g)

        self.pipe_material = QComboBox()
        self.pipe_material.addItems(list(PIPE_MATERIALS))
        self.pipe_material.currentTextChanged.connect(self._on_pipe_material_changed)

        self.in_ambient = self._dbl(300.0, 100.0, 2000.0, 2, " K")
        self.in_eps = self._dbl(0.70, 0.0, 1.0, 3, "")

        self.chk_use_insulation = QCheckBox("Enable insulation")
        self.chk_use_insulation.setChecked(True)
        self.chk_use_insulation.toggled.connect(self._sync_insulation_widgets)

        self.ins_material = QComboBox()
        self.ins_material.addItems(list(INSULATION_MATERIALS))
        self.in_t_ins = self._dbl(0.15, 0.0, 2.0, 4, " m")
        self.chk_temp_dep_props = QCheckBox("Use temperature-dependent\ncp(T), k(T)")
        self.chk_temp_dep_props.setChecked(True)
        self.lbl_temp_dep_hint = QLabel(
            "Edit property tables in Library"
            "-> Temperature-Dependent Tables."
            "If disabled, scalar cp/k values are used."
        )
        self.lbl_temp_dep_hint.setWordWrap(True)
        self.lbl_temp_dep_hint.setStyleSheet("color: #8c8c8c;")

        form.addRow("Pipe material", self.pipe_material)
        form.addRow("Ambient temp", self.in_ambient)
        form.addRow("Surface emissivity", self.in_eps)
        form.addRow("", self.chk_use_insulation)
        form.addRow("Insulation material", self.ins_material)
        form.addRow("Insulation thickness", self.in_t_ins)
        form.addRow("", self.chk_temp_dep_props)
        form.addRow("", self.lbl_temp_dep_hint)
        return g

    def _build_pipe_group(self):
        g = QGroupBox("3) Pipe Geometry")
        form = QFormLayout(g)

        self.in_L = self._dbl(65.0, 0.01, 1e6, 3, " m")
        self.in_Di = self._dbl(0.13, 0.001, 1e3, 4, " m")
        self.in_t_wall = self._dbl(0.018, 0.0001, 1.0, 4, " m")
        self.in_elbows = QSpinBox()
        self.in_elbows.setRange(0, 40)
        self.in_elbows.setValue(0)
        self.in_elbow_sif = self._dbl(1.30, 1.00, 5.00, 2, "")
        self.in_elbow_positions = QLineEdit("")
        self.in_elbow_positions.setPlaceholderText("auto or e.g. 45, 50 (m/ft) or 20%, 60%")
        self.in_tmass_count = QSpinBox()
        self.in_tmass_count.setRange(0, 100)
        self.in_tmass_count.setValue(0)
        self.in_tmass_factor = self._dbl(0.0, 0.0, 50.0, 2, " x")
        self.in_tmass_positions = QLineEdit("")
        self.in_tmass_positions.setPlaceholderText("auto or e.g. 12, 35 (m/ft) or 20%, 60%")
        self.in_tmass_spread = self._dbl(3.0, 0.1, 50.0, 1, " %L")
        self.in_tmass_deadleg_len = self._dbl(2.0, 0.0, 1.0e6, 3, " m")
        self.in_tmass_deadleg_d_ratio = self._dbl(1.0, 0.1, 10.0, 2, " x")
        btn_apply_tmass_est = QPushButton("Apply rough mass factor")
        btn_apply_tmass_est.clicked.connect(self._apply_tmass_rough_estimate)
        self.lbl_tmass_help = QLabel(
            "Mass factor increases local wall thermal capacity around each attachment center. "
            "If positions are blank, centers auto-distribute from 15% to 85% of length."
        )
        self.lbl_tmass_help.setWordWrap(True)
        self.lbl_tmass_help.setStyleSheet("color: #8c8c8c;")
        self.lbl_tmass_est = QLabel("")
        self.lbl_tmass_est.setWordWrap(True)
        self.lbl_tmass_est.setStyleSheet("color: #8c8c8c;")
        self.in_tmass_deadleg_len.valueChanged.connect(self._update_tmass_estimate_label)
        self.in_tmass_deadleg_d_ratio.valueChanged.connect(self._update_tmass_estimate_label)
        self.in_tmass_spread.valueChanged.connect(self._update_tmass_estimate_label)
        self.in_L.valueChanged.connect(self._update_tmass_estimate_label)

        form.addRow("Length L", self.in_L)
        form.addRow("Inner diameter Di", self.in_Di)
        form.addRow("Wall thickness", self.in_t_wall)
        form.addRow("Number of elbows", self.in_elbows)
        form.addRow("Elbow SIF factor", self.in_elbow_sif)
        form.addRow("Elbow positions", self.in_elbow_positions)
        form.addRow("Attached thermal masses", self.in_tmass_count)
        form.addRow("Mass factor per attach", self.in_tmass_factor)
        form.addRow("Mass positions", self.in_tmass_positions)
        form.addRow("Mass spread", self.in_tmass_spread)
        form.addRow("Dead-leg length (rough)", self.in_tmass_deadleg_len)
        form.addRow("Dead-leg diameter ratio", self.in_tmass_deadleg_d_ratio)
        form.addRow("", btn_apply_tmass_est)
        form.addRow("", self.lbl_tmass_help)
        form.addRow("", self.lbl_tmass_est)
        return g

    def _build_flow_group(self):
        g = QGroupBox("4) Flow / Inlet")
        form = QFormLayout(g)

        self.in_p = self._dbl(5.0e6, 1.0, 1.0e9, 1, " Pa")
        self.in_Tin = self._dbl(1000.0, 1.0, 5000.0, 2, " K")
        self.in_tin_ramp = self._dbl(900.0, 0.0, 1.0e6, 1, " s")
        self.in_tin_ramp_model = QComboBox()
        self.in_tin_ramp_model.addItem("Logistic (S-curve)", "logistic")
        self.in_tin_ramp_model.addItem("Linear", "linear")
        self.in_tin_ramp_model.addItem("Exponential", "heater_exp")
        self.in_tin_ramp_model.setCurrentIndex(0)
        self.in_mdot = self._dbl(2.5, 0.0001, 1.0e4, 4, " kg/s")
        self.in_mdot_min = self._dbl(0.2, 0.0001, 1.0e4, 4, " kg/s")
        self.in_mdot_max = self._dbl(5.0, 0.0001, 1.0e4, 4, " kg/s")

        form.addRow("Pressure p", self.in_p)
        form.addRow("Heater setpoint Tin", self.in_Tin)
        form.addRow("Heater ramp profile", self.in_tin_ramp_model)
        form.addRow("Heater rise time\nto setpoint", self.in_tin_ramp)
        form.addRow("Mass flow m_dot", self.in_mdot)
        form.addRow("Search m_dot min", self.in_mdot_min)
        form.addRow("Search m_dot max", self.in_mdot_max)
        return g

    def _build_output_group(self):
        g = QGroupBox("5) Outputs")
        form = QFormLayout(g)

        self.chk_make_plots = QCheckBox("Save plot images")
        self.chk_make_plots.setChecked(False)
        self.chk_save_results = QCheckBox("Save run artifacts")
        self.chk_save_results.setChecked(True)
        self.chk_append_ledger = QCheckBox("Append run-history\nledger")
        self.chk_append_ledger.setChecked(False)

        checks = QWidget()
        checks_l = QVBoxLayout(checks)
        checks_l.setContentsMargins(0, 0, 0, 0)
        checks_l.setSpacing(2)
        checks_l.addWidget(self.chk_make_plots)
        checks_l.addWidget(self.chk_save_results)
        checks_l.addWidget(self.chk_append_ledger)

        self.dir_label = QLabel("(auto)")
        self.dir_label.setMinimumWidth(0)
        self.dir_label.setSizePolicy(QSizePolicy.Policy.Ignored, QSizePolicy.Policy.Preferred)
        btn_dir = QPushButton("Choose save folder...")
        btn_dir.clicked.connect(self._choose_dir)

        out_row = QWidget()
        out_l = QHBoxLayout(out_row)
        out_l.setContentsMargins(0, 0, 0, 0)
        out_l.setSpacing(6)
        out_l.addWidget(btn_dir)
        out_l.addWidget(self.dir_label, 1)

        self.ledger_label = QLabel("(auto)")
        self.ledger_label.setMinimumWidth(0)
        self.ledger_label.setSizePolicy(QSizePolicy.Policy.Ignored, QSizePolicy.Policy.Preferred)
        self.ledger_label.setToolTip("Auto: run_history.csv")
        btn_ledger = QPushButton("Choose ledger...")
        btn_ledger.clicked.connect(self._choose_ledger_file)
        ledger_row = QWidget()
        ledger_l = QHBoxLayout(ledger_row)
        ledger_l.setContentsMargins(0, 0, 0, 0)
        ledger_l.setSpacing(6)
        ledger_l.addWidget(btn_ledger)
        ledger_l.addWidget(self.ledger_label, 1)

        form.addRow("Options", checks)
        form.addRow("Folder", out_row)
        form.addRow("Run ledger", ledger_row)
        return g

    def _build_advanced_group(self):
        g = QGroupBox("6) Advanced Numerics")
        form = QFormLayout(g)

        self.in_Nx = QSpinBox()
        self.in_Nx.setRange(50, 30000)
        self.in_Nx.setValue(1300)

        self.in_save_frames = QSpinBox()
        self.in_save_frames.setRange(2, 5000)
        self.in_save_frames.setValue(240)

        self.in_dt_max = self._dbl(2.5, 0.0001, 1.0e3, 4, " s")
        self.in_dt_min = self._dbl(1.0e-4, 1.0e-8, 100.0, 6, " s")
        self.in_update_props = QSpinBox()
        self.in_update_props.setRange(1, 1000)
        self.in_update_props.setValue(5)
        self.in_adv_scheme = QComboBox()
        self.in_adv_scheme.addItem("semi_lagrangian", "semi_lagrangian")
        self.in_adv_scheme.addItem("upwind", "upwind")
        self.in_adv_scheme.setCurrentIndex(0)
        self.in_semi_lag_cmax = self._dbl(1.2, 0.1, 20.0, 2, "")
        self.in_ins_mass_mode = QComboBox()
        self.in_ins_mass_mode.addItem("penetration", "penetration")
        self.in_ins_mass_mode.addItem("full", "full")
        self.in_ins_mass_mode.setCurrentIndex(0)
        self.in_ins_mass_min_frac = self._dbl(0.25, 0.01, 1.0, 2, "")

        self.in_nr_wall = QSpinBox()
        self.in_nr_wall.setRange(3, 80)
        self.in_nr_wall.setValue(12)
        self.in_axial_restraint = self._dbl(0.00, 0.0, 1.0, 2, "")
        self.in_ignore_inlet_cells = QSpinBox()
        self.in_ignore_inlet_cells.setRange(0, 200)
        self.in_ignore_inlet_cells.setValue(2)

        self.progress = QComboBox()
        self.progress.addItems(["none", "basic"])
        self.progress.setCurrentText("basic")

        self.chk_use_float32 = QCheckBox("use_float32")
        self.chk_use_float32.setChecked(True)
        self.chk_log_to_file = QCheckBox("write run.log")
        self.chk_log_to_file.setChecked(True)
        self.chk_write_trace = QCheckBox("write runtime trace")
        self.chk_write_trace.setChecked(True)
        self.chk_show_plots = QCheckBox("open results popup\nafter run")
        self.chk_show_plots.setChecked(False)
        self.chk_include_pressure = QCheckBox("include pressure in\ntotal stress")
        self.chk_include_pressure.setChecked(True)
        self.chk_convergence_diag = QCheckBox("stress sensitivity\ndiagnostics")
        self.chk_convergence_diag.setChecked(True)
        self.chk_target_asymptote = QCheckBox("target asymptote stop")
        self.chk_target_asymptote.setChecked(True)

        flags = QWidget()
        flags_l = QVBoxLayout(flags)
        flags_l.setContentsMargins(0, 0, 0, 0)
        flags_l.setSpacing(2)
        flags_l.addWidget(self.chk_use_float32)
        flags_l.addWidget(self.chk_log_to_file)
        flags_l.addWidget(self.chk_write_trace)
        flags_l.addWidget(self.chk_show_plots)
        flags_l.addWidget(self.chk_include_pressure)
        flags_l.addWidget(self.chk_convergence_diag)
        flags_l.addWidget(self.chk_target_asymptote)

        form.addRow("Nx", self.in_Nx)
        form.addRow("Save frames", self.in_save_frames)
        form.addRow("dt max", self.in_dt_max)
        form.addRow("dt min", self.in_dt_min)
        form.addRow("Props update step", self.in_update_props)
        form.addRow("Advection", self.in_adv_scheme)
        form.addRow("Semi-Lag Courant max", self.in_semi_lag_cmax)
        form.addRow("Insulation mass model", self.in_ins_mass_mode)
        form.addRow("Insulation min frac", self.in_ins_mass_min_frac)
        form.addRow("Wall Nr", self.in_nr_wall)
        form.addRow("Axial restraint", self.in_axial_restraint)
        form.addRow("Ignore inlet cells", self.in_ignore_inlet_cells)
        form.addRow("Progress", self.progress)
        form.addRow("Flags", flags)
        return g

    def _build_live_tab(self):
        tab = QWidget()
        layout = QVBoxLayout(tab)

        controls = QHBoxLayout()
        self.btn_run_anim = QPushButton("Run Animation")
        self.btn_pause_anim = QPushButton("Pause")
        self.slider_time = QSlider(Qt.Orientation.Horizontal)
        self.slider_time.setEnabled(False)
        self.lbl_time_cursor = QLabel("t = -- s / -- s")

        self.btn_run_anim.clicked.connect(self._run_animation)
        self.btn_pause_anim.clicked.connect(self._pause_animation)
        self.slider_time.valueChanged.connect(self._on_slider_changed)

        if not HAS_MPL:
            msg = QLabel("Matplotlib is required for embedded live plots.\nInstall with: pip install matplotlib")
            msg.setWordWrap(True)
            layout.addWidget(msg)
            self.btn_run_anim.setEnabled(False)
            self.btn_pause_anim.setEnabled(False)
            controls.addWidget(self.btn_run_anim)
            controls.addWidget(self.btn_pause_anim)
            controls.addWidget(self.slider_time, 1)
            controls.addWidget(self.lbl_time_cursor)
            layout.addLayout(controls)
            return tab

        split = QSplitter(Qt.Orientation.Vertical)
        split.addWidget(self._build_profile_canvas())
        split.addWidget(self._build_heatmap_canvas())
        split.setStretchFactor(0, 1)
        split.setStretchFactor(1, 1)
        split.setSizes([430, 360])
        layout.addWidget(split, 1)

        controls.addWidget(self.btn_run_anim)
        controls.addWidget(self.btn_pause_anim)
        controls.addWidget(self.slider_time, 1)
        controls.addWidget(self.lbl_time_cursor)
        layout.addLayout(controls)
        return tab

    def _build_profile_canvas(self):
        w = QWidget()
        l = QVBoxLayout(w)
        l.setContentsMargins(0, 0, 0, 0)

        self.live_fig = Figure(figsize=(8, 6), constrained_layout=True)
        self.ax_outlet = self.live_fig.add_subplot(2, 1, 1)
        self.ax_profile = self.live_fig.add_subplot(2, 1, 2)
        self.live_canvas = FigureCanvas(self.live_fig)
        l.addWidget(self.live_canvas, 1)
        return w

    def _build_heatmap_canvas(self):
        w = QWidget()
        l = QVBoxLayout(w)
        l.setContentsMargins(0, 0, 0, 0)

        self.heat_fig = Figure(figsize=(8, 4), constrained_layout=True)
        self.ax_heat = self.heat_fig.add_subplot(1, 1, 1)
        self.heat_canvas = FigureCanvas(self.heat_fig)
        l.addWidget(self.heat_canvas, 1)
        return w

    def _build_results_tab(self):
        tab = QWidget()
        layout = QVBoxLayout(tab)
        if not HAS_MPL:
            msg = QLabel("Matplotlib is required to render results plots.")
            msg.setWordWrap(True)
            layout.addWidget(msg)
            return tab

        self.stats_box = QPlainTextEdit()
        self.stats_box.setReadOnly(True)
        self.stats_box.setMaximumBlockCount(500)
        self.stats_box.setFixedHeight(140)
        layout.addWidget(self.stats_box)

        self.warning_box = QPlainTextEdit()
        self.warning_box.setReadOnly(True)
        self.warning_box.setMaximumBlockCount(500)
        self.warning_box.setFixedHeight(120)
        self.warning_box.setStyleSheet("QPlainTextEdit { background: #1f1f1f; color: #ffb347; }")
        layout.addWidget(self.warning_box)

        tip = QLabel("Tip: click any results plot to open a larger popup view.")
        tip.setStyleSheet("color: #8c8c8c;")
        layout.addWidget(tip)

        self.results_fig = Figure(figsize=(12, 8), constrained_layout=True)
        self.results_canvas = FigureCanvas(self.results_fig)
        self.results_canvas.mpl_connect("button_press_event", self._on_results_plot_click)
        layout.addWidget(self.results_canvas, 1)
        return tab

    def _build_material_library_tab(self):
        tab = QWidget()
        layout = QHBoxLayout(tab)

        left = QVBoxLayout()
        right = QVBoxLayout()
        preset_col = QVBoxLayout()

        pipe_group = QGroupBox("Add / Update Pipe Material")
        pipe_form = QFormLayout(pipe_group)
        self.new_pipe_name = QLineEdit()
        self.new_pipe_rho = self._dbl(8000.0, 1.0, 50000.0, 2, " kg/m^3")
        self.new_pipe_cp = self._dbl(500.0, 1.0, 5000.0, 2, " J/kgK")
        self.new_pipe_k = self._dbl(16.0, 0.01, 2000.0, 3, " W/mK")
        self.new_pipe_E = self._dbl(200.0, 1.0, 500.0, 3, " GPa")
        self.new_pipe_alpha = self._dbl(14.0, 0.1, 50.0, 3, " um/mK")
        self.new_pipe_nu = self._dbl(0.30, 0.01, 0.49, 3, "")
        self.new_pipe_sy = self._dbl(300.0, 1.0, 3000.0, 1, " MPa")
        self.new_pipe_eps = self._dbl(0.70, 0.0, 1.0, 3, "")
        btn_add_pipe = QPushButton("Add / Update Pipe Material")
        btn_add_pipe.clicked.connect(self._add_pipe_material)

        pipe_form.addRow("Name", self.new_pipe_name)
        pipe_form.addRow("Density", self.new_pipe_rho)
        pipe_form.addRow("cp", self.new_pipe_cp)
        pipe_form.addRow("k", self.new_pipe_k)
        pipe_form.addRow("E", self.new_pipe_E)
        pipe_form.addRow("alpha", self.new_pipe_alpha)
        pipe_form.addRow("nu", self.new_pipe_nu)
        pipe_form.addRow("Yield strength", self.new_pipe_sy)
        pipe_form.addRow("Default emissivity", self.new_pipe_eps)
        pipe_form.addRow(btn_add_pipe)

        ins_group = QGroupBox("Add / Update Insulation Material")
        ins_form = QFormLayout(ins_group)
        self.new_ins_name = QLineEdit()
        self.new_ins_rho = self._dbl(128.0, 1.0, 5000.0, 2, " kg/m^3")
        self.new_ins_cp = self._dbl(840.0, 1.0, 5000.0, 2, " J/kgK")
        self.new_ins_k = self._dbl(0.045, 0.001, 10.0, 4, " W/mK")
        btn_add_ins = QPushButton("Add / Update Insulation Material")
        btn_add_ins.clicked.connect(self._add_ins_material)

        ins_form.addRow("Name", self.new_ins_name)
        ins_form.addRow("Density", self.new_ins_rho)
        ins_form.addRow("cp", self.new_ins_cp)
        ins_form.addRow("k", self.new_ins_k)
        ins_form.addRow(btn_add_ins)

        self.pipe_list = QListWidget()
        self.ins_list = QListWidget()

        pipe_list_group = QGroupBox("Pipe Materials")
        pipe_list_l = QVBoxLayout(pipe_list_group)
        pipe_list_l.addWidget(self.pipe_list)

        ins_list_group = QGroupBox("Insulation Materials")
        ins_list_l = QVBoxLayout(ins_list_group)
        ins_list_l.addWidget(self.ins_list)

        temp_table_group = QGroupBox("Temperature-Dependent Tables (cp(T), k(T))")
        temp_form = QFormLayout(temp_table_group)
        self.temp_table_type = QComboBox()
        self.temp_table_type.addItems(["Pipe", "Insulation"])
        self.temp_table_material = QComboBox()
        self.temp_table_type.currentTextChanged.connect(self._refresh_temp_prop_materials)
        self.temp_table_material.currentTextChanged.connect(self._load_temp_prop_editor)
        self.temp_table_T = QLineEdit()
        self.temp_table_cp = QLineEdit()
        self.temp_table_k = QLineEdit()
        self.temp_table_T.setPlaceholderText("300, 500, 700, 900, 1100")
        self.temp_table_cp.setPlaceholderText("e.g. 500, 530, 560, 600, 640")
        self.temp_table_k.setPlaceholderText("e.g. 14.5, 16.0, 17.8, 19.5, 21.2")

        temp_btns = QWidget()
        temp_btns_l = QHBoxLayout(temp_btns)
        temp_btns_l.setContentsMargins(0, 0, 0, 0)
        btn_load_table = QPushButton("Load")
        btn_save_table = QPushButton("Save Table")
        btn_flat_table = QPushButton("Reset Flat")
        btn_load_table.clicked.connect(self._load_temp_prop_editor)
        btn_save_table.clicked.connect(self._save_temp_prop_editor)
        btn_flat_table.clicked.connect(self._reset_temp_prop_editor_flat)
        temp_btns_l.addWidget(btn_load_table)
        temp_btns_l.addWidget(btn_save_table)
        temp_btns_l.addWidget(btn_flat_table)

        temp_hint = QLabel(
            "Enter comma-separated values with matching lengths. "
            "T must be strictly increasing and cp/k must stay positive."
        )
        temp_hint.setWordWrap(True)
        temp_hint.setStyleSheet("color: #8c8c8c;")

        temp_form.addRow("Material type", self.temp_table_type)
        temp_form.addRow("Material", self.temp_table_material)
        temp_form.addRow("T [K]", self.temp_table_T)
        temp_form.addRow("cp [J/kgK]", self.temp_table_cp)
        temp_form.addRow("k [W/mK]", self.temp_table_k)
        temp_form.addRow("", temp_btns)
        temp_form.addRow("", temp_hint)

        preset_group = QGroupBox("Pipe Presets")
        preset_form = QFormLayout(preset_group)
        self.new_preset_name = QLineEdit()
        self.new_preset_name.setPlaceholderText("e.g. Branch A startup")
        btn_add_preset = QPushButton("Save Current as Preset")
        btn_add_preset.clicked.connect(self._save_current_pipe_preset)
        preset_form.addRow("Preset name", self.new_preset_name)
        preset_form.addRow("", btn_add_preset)

        preset_list_group = QGroupBox("Preset Library")
        preset_list_l = QVBoxLayout(preset_list_group)
        self.pipe_preset_list = QListWidget()
        self.pipe_preset_list.itemDoubleClicked.connect(lambda _item: self._apply_selected_pipe_preset())
        preset_buttons = QHBoxLayout()
        btn_apply_preset = QPushButton("Apply")
        btn_apply_preset.clicked.connect(self._apply_selected_pipe_preset)
        btn_delete_preset = QPushButton("Delete")
        btn_delete_preset.clicked.connect(self._delete_selected_pipe_preset)
        preset_buttons.addWidget(btn_apply_preset)
        preset_buttons.addWidget(btn_delete_preset)
        preset_list_l.addWidget(self.pipe_preset_list, 1)
        preset_list_l.addLayout(preset_buttons)

        left.addWidget(pipe_group)
        left.addWidget(ins_group)
        right.addWidget(pipe_list_group)
        right.addWidget(ins_list_group)
        right.addWidget(temp_table_group)
        preset_col.addWidget(preset_group)
        preset_col.addWidget(preset_list_group, 1)

        layout.addLayout(left, 1)
        layout.addLayout(right, 1)
        layout.addLayout(preset_col, 1)
        return tab

    def _build_ledger_tab(self):
        tab = QWidget()
        layout = QVBoxLayout(tab)

        controls = QHBoxLayout()
        self.btn_ledger_refresh = QPushButton("Refresh")
        self.btn_ledger_refresh.clicked.connect(self._refresh_ledger_preview)
        self.btn_ledger_append = QPushButton("Append Current Config")
        self.btn_ledger_append.clicked.connect(self._append_current_config_to_ledger)
        self.btn_ledger_delete = QPushButton("Delete Selected Row(s)")
        self.btn_ledger_delete.clicked.connect(self._delete_selected_ledger_rows)
        self.lbl_ledger_status = QLabel("Ledger: --")
        self.lbl_ledger_status.setMinimumWidth(0)
        self.lbl_ledger_status.setSizePolicy(QSizePolicy.Policy.Ignored, QSizePolicy.Policy.Preferred)
        controls.addWidget(self.btn_ledger_refresh)
        controls.addWidget(self.btn_ledger_append)
        controls.addWidget(self.btn_ledger_delete)
        controls.addStretch(1)
        controls.addWidget(self.lbl_ledger_status)
        layout.addLayout(controls)

        self.ledger_table = CompactTableWidget()
        self.ledger_table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.ledger_table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.ledger_table.setSelectionMode(QAbstractItemView.SelectionMode.ExtendedSelection)
        self.ledger_table.setSizeAdjustPolicy(QAbstractScrollArea.SizeAdjustPolicy.AdjustIgnored)
        self.ledger_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Interactive)
        self.ledger_table.horizontalHeader().setStretchLastSection(False)
        self.ledger_table.setHorizontalScrollMode(QAbstractItemView.ScrollMode.ScrollPerPixel)
        self.ledger_table.setWordWrap(False)
        self.ledger_table.setMinimumWidth(0)
        self.ledger_table.setMinimumSize(0, 0)
        self.ledger_table.setSizePolicy(QSizePolicy.Policy.Ignored, QSizePolicy.Policy.Expanding)
        layout.addWidget(self.ledger_table, 1)
        return tab

    def _build_readme_tab(self):
        tab = QWidget()
        layout = QVBoxLayout(tab)
        controls = QHBoxLayout()
        btn_reload = QPushButton("Reload README")
        btn_reload.clicked.connect(self._refresh_readme_view)
        self.lbl_readme_path = QLabel(str(self._readme_path))
        controls.addWidget(btn_reload)
        controls.addWidget(self.lbl_readme_path, 1)
        layout.addLayout(controls)

        self.readme_view = QTextBrowser()
        self.readme_view.setOpenExternalLinks(True)
        layout.addWidget(self.readme_view, 1)
        return tab

    def _build_log_tab(self):
        tab = QWidget()
        layout = QVBoxLayout(tab)
        self.log_view = QPlainTextEdit()
        self.log_view.setReadOnly(True)
        layout.addWidget(self.log_view, 1)
        return tab

    def _build_run_bar(self):
        row = QHBoxLayout()
        self.btn_run = QPushButton("Run Simulation")
        self.btn_run.clicked.connect(self._run_clicked)
        self.btn_cancel = QPushButton("Cancel Simulation")
        self.btn_cancel.setEnabled(False)
        self.btn_cancel.clicked.connect(self._cancel_clicked)

        self.progress_bar = QProgressBar()
        self.progress_bar.setMinimum(0)
        self.progress_bar.setMaximum(1)
        self.progress_bar.setValue(0)
        self.progress_bar.setTextVisible(False)
        self.progress_bar.setFixedHeight(14)

        self.status = QLabel("Idle")
        self.status.setFrameShape(QFrame.Shape.NoFrame)
        self.btn_export = QPushButton("Export Bundle")
        self.btn_export.setEnabled(False)
        self.btn_export.clicked.connect(self._export_bundle_clicked)

        row.addWidget(self.btn_run)
        row.addWidget(self.btn_cancel)
        row.addWidget(self.btn_export)
        row.addWidget(self.progress_bar, 1)
        row.addWidget(self.status, 1)
        return row

    def _dbl(self, value, minimum, maximum, decimals, suffix):
        box = QDoubleSpinBox()
        box.setDecimals(decimals)
        box.setRange(minimum, maximum)
        box.setValue(value)
        box.setSuffix(suffix)
        return box

    def _setup_logging(self):
        root_logger = logging.getLogger()
        root_logger.setLevel(logging.INFO)
        for handler in list(root_logger.handlers):
            if getattr(handler, "_thermal_pipe_qt_handler", False):
                root_logger.removeHandler(handler)
                try:
                    handler.close()
                except Exception:
                    pass

        self.log_emitter = LogEmitter()
        self.log_emitter.message.connect(self.log_view.appendPlainText)
        self.log_handler = QtLogHandler(self.log_emitter)
        self.log_handler.setFormatter(
            logging.Formatter("%(asctime)s | %(levelname)s | %(message)s", datefmt="%H:%M:%S")
        )
        root_logger.addHandler(self.log_handler)
        if not HAS_NUMBA:
            logging.warning("Numba acceleration not available; runs may be slower. Install 'numba' for best performance.")
        elif sys.platform.startswith("win"):
            logging.warning(
                "Windows stability mode active: Numba kernels are disabled by default for runs "
                "to avoid native-process crashes."
            )

    def _refresh_material_lists(self):
        if hasattr(self, "pipe_list"):
            self.pipe_list.clear()
            for name, m in PIPE_MATERIALS.items():
                sy_mpa = float(m.get("Sy", 250e6)) / 1.0e6
                t_table = PIPE_TEMP_PROPS.get(name)
                t_tag = ""
                if isinstance(t_table, dict):
                    t_vals = [float(v) for v in t_table.get("T", [])]
                    if len(t_vals) >= 2:
                        t_tag = f", Tdep={len(t_vals)}pts [{min(t_vals):.0f}-{max(t_vals):.0f} K]"
                self.pipe_list.addItem(
                    f"{name}: rho={m['rho_w']:.0f}, cp={m['cp_w']:.1f}, k={m['k_w']:.2f}, "
                    f"E={m['E']/1e9:.1f} GPa, alpha={m['alpha']*1e6:.2f}e-6, nu={m['nu']:.2f}, Sy={sy_mpa:.1f} MPa{t_tag}"
                )
        if hasattr(self, "ins_list"):
            self.ins_list.clear()
            for name, m in INSULATION_MATERIALS.items():
                t_table = INSULATION_TEMP_PROPS.get(name)
                t_tag = ""
                if isinstance(t_table, dict):
                    t_vals = [float(v) for v in t_table.get("T", [])]
                    if len(t_vals) >= 2:
                        t_tag = f", Tdep={len(t_vals)}pts [{min(t_vals):.0f}-{max(t_vals):.0f} K]"
                self.ins_list.addItem(f"{name}: rho={m['rho_i']:.0f}, cp={m['cp_i']:.1f}, k={m['k_i']:.4f}{t_tag}")

    def _refresh_material_combos(self):
        cur_pipe = self.pipe_material.currentText() if hasattr(self, "pipe_material") else None
        cur_ins = self.ins_material.currentText() if hasattr(self, "ins_material") else None
        if hasattr(self, "pipe_material"):
            self.pipe_material.blockSignals(True)
            self.pipe_material.clear()
            self.pipe_material.addItems(list(PIPE_MATERIALS))
            if cur_pipe in PIPE_MATERIALS:
                self.pipe_material.setCurrentText(cur_pipe)
            self.pipe_material.blockSignals(False)
        if hasattr(self, "ins_material"):
            self.ins_material.blockSignals(True)
            self.ins_material.clear()
            self.ins_material.addItems(list(INSULATION_MATERIALS))
            if cur_ins in INSULATION_MATERIALS:
                self.ins_material.setCurrentText(cur_ins)
            self.ins_material.blockSignals(False)
        self._refresh_temp_prop_materials()

    def _estimate_tmass_factor_from_deadleg(self) -> float:
        length_main = max(float(self._length_from_display(self.in_L.value())), 1.0e-9)
        deadleg_len = max(float(self._length_from_display(self.in_tmass_deadleg_len.value())), 0.0)
        spread_frac = max(1.0e-4, 0.01 * float(self.in_tmass_spread.value()))
        diam_ratio = max(0.1, float(self.in_tmass_deadleg_d_ratio.value()))
        # Triangular spread integrates to ~3*spread_frac of line mass for factor=1.0.
        return float((diam_ratio * (deadleg_len / length_main)) / (3.0 * spread_frac))

    def _update_tmass_estimate_label(self, *_args):
        if not hasattr(self, "lbl_tmass_est"):
            return
        est = self._estimate_tmass_factor_from_deadleg()
        l_unit = "m" if self._units == "SI" else "ft"
        self.lbl_tmass_est.setText(
            "Dead-end T rough guide: "
            f"mass factor ~= {est:.2f} per attachment "
            f"(L_dead={self.in_tmass_deadleg_len.value():.3f} {l_unit}, "
            f"D_ratio={self.in_tmass_deadleg_d_ratio.value():.2f}, spread={self.in_tmass_spread.value():.1f}%L)."
        )

    def _apply_tmass_rough_estimate(self, *_args):
        est = self._estimate_tmass_factor_from_deadleg()
        est = max(float(self.in_tmass_factor.minimum()), min(float(self.in_tmass_factor.maximum()), est))
        self._set_box_safely(self.in_tmass_factor, est)
        self._update_tmass_estimate_label()

    @staticmethod
    def _parse_float_list(text: str) -> list[float]:
        raw = str(text or "").replace(";", ",").replace("\n", ",")
        vals: list[float] = []
        for token in raw.split(","):
            tok = token.strip()
            if not tok:
                continue
            vals.append(float(tok))
        return vals

    def _active_temp_prop_store(self) -> tuple[dict[str, dict[str, list[float]]], dict[str, dict[str, float]]]:
        if self.temp_table_type.currentText() == "Pipe":
            return PIPE_TEMP_PROPS, PIPE_MATERIALS
        return INSULATION_TEMP_PROPS, INSULATION_MATERIALS

    def _refresh_temp_prop_materials(self, *_args):
        if not hasattr(self, "temp_table_material"):
            return
        store, base = self._active_temp_prop_store()
        current = self.temp_table_material.currentText()
        names = list(base.keys())
        self.temp_table_material.blockSignals(True)
        self.temp_table_material.clear()
        self.temp_table_material.addItems(names)
        if current in names:
            self.temp_table_material.setCurrentText(current)
        elif names:
            self.temp_table_material.setCurrentIndex(0)
        self.temp_table_material.blockSignals(False)
        if names:
            self._load_temp_prop_editor()

    def _load_temp_prop_editor(self, *_args):
        if not hasattr(self, "temp_table_material"):
            return
        store, base = self._active_temp_prop_store()
        name = self.temp_table_material.currentText().strip()
        if not name:
            return
        table = store.get(name)
        if not isinstance(table, dict):
            if self.temp_table_type.currentText() == "Pipe":
                mat = base.get(name, {})
                table = self._flat_prop_table(cp=float(mat.get("cp_w", 500.0)), k=float(mat.get("k_w", 16.0)))
            else:
                mat = base.get(name, {})
                table = self._flat_prop_table(cp=float(mat.get("cp_i", 900.0)), k=float(mat.get("k_i", 0.06)))
            store[name] = table
        self.temp_table_T.setText(", ".join(f"{float(v):g}" for v in table.get("T", [])))
        self.temp_table_cp.setText(", ".join(f"{float(v):g}" for v in table.get("cp", [])))
        self.temp_table_k.setText(", ".join(f"{float(v):g}" for v in table.get("k", [])))

    def _save_temp_prop_editor(self, *_args):
        store, _base = self._active_temp_prop_store()
        name = self.temp_table_material.currentText().strip()
        if not name:
            QMessageBox.warning(self, "Temp Property Table", "Select a material first.")
            return
        try:
            t_vals = self._parse_float_list(self.temp_table_T.text())
            cp_vals = self._parse_float_list(self.temp_table_cp.text())
            k_vals = self._parse_float_list(self.temp_table_k.text())
        except Exception:
            QMessageBox.warning(self, "Temp Property Table", "Could not parse T/cp/k lists. Use comma-separated numbers.")
            return
        if len(t_vals) < 2:
            QMessageBox.warning(self, "Temp Property Table", "At least two temperature points are required.")
            return
        if len(cp_vals) != len(t_vals) or len(k_vals) != len(t_vals):
            QMessageBox.warning(self, "Temp Property Table", "T, cp, and k lists must have matching lengths.")
            return
        if any((t_vals[i] >= t_vals[i + 1]) for i in range(len(t_vals) - 1)):
            QMessageBox.warning(self, "Temp Property Table", "T values must be strictly increasing.")
            return
        if any(v <= 0.0 for v in cp_vals) or any(v <= 0.0 for v in k_vals):
            QMessageBox.warning(self, "Temp Property Table", "cp and k values must be > 0.")
            return
        store[name] = {
            "T": [float(v) for v in t_vals],
            "cp": [float(v) for v in cp_vals],
            "k": [float(v) for v in k_vals],
        }
        self._refresh_material_lists()
        QMessageBox.information(self, "Temp Property Table", f"Saved table for {name}.")

    def _reset_temp_prop_editor_flat(self, *_args):
        store, base = self._active_temp_prop_store()
        name = self.temp_table_material.currentText().strip()
        if not name:
            return
        if self.temp_table_type.currentText() == "Pipe":
            mat = base.get(name, {})
            table = self._flat_prop_table(cp=float(mat.get("cp_w", 500.0)), k=float(mat.get("k_w", 16.0)))
        else:
            mat = base.get(name, {})
            table = self._flat_prop_table(cp=float(mat.get("cp_i", 900.0)), k=float(mat.get("k_i", 0.06)))
        store[name] = table
        self._load_temp_prop_editor()
        self._refresh_material_lists()

    def _all_pipe_defaults(self) -> Dict[str, Dict[str, Any]]:
        merged = dict(self.PIPE_DEFAULTS)
        merged.update(self._custom_pipe_presets)
        return merged

    def _refresh_pipe_default_combo(self):
        if not hasattr(self, "pipe_default"):
            return
        current = self.pipe_default.currentText()
        names = list(self.PIPE_DEFAULTS.keys()) + sorted(
            [n for n in self._custom_pipe_presets.keys() if n not in self.PIPE_DEFAULTS]
        )
        self.pipe_default.blockSignals(True)
        self.pipe_default.clear()
        self.pipe_default.addItems(names)
        if current in names:
            self.pipe_default.setCurrentText(current)
        elif names:
            self.pipe_default.setCurrentIndex(0)
        self.pipe_default.blockSignals(False)

    def _refresh_pipe_preset_list(self):
        if not hasattr(self, "pipe_preset_list"):
            return
        self.pipe_preset_list.clear()
        for name in list(self.PIPE_DEFAULTS.keys()) + sorted(
            [n for n in self._custom_pipe_presets.keys() if n not in self.PIPE_DEFAULTS]
        ):
            self.pipe_preset_list.addItem(name)

    def _load_custom_pipe_presets(self):
        self._custom_pipe_presets = {}
        if not self._pipe_preset_path.exists():
            return
        try:
            raw = json.loads(self._pipe_preset_path.read_text(encoding="utf-8"))
            if isinstance(raw, dict):
                for name, cfg in raw.items():
                    if not isinstance(name, str) or not isinstance(cfg, dict):
                        continue
                    if name in self.PIPE_DEFAULTS:
                        continue
                    self._custom_pipe_presets[name] = cfg
        except Exception as exc:
            logging.warning("Failed to load pipe presets from %s: %s", self._pipe_preset_path, exc)

    def _save_custom_pipe_presets(self):
        try:
            self._pipe_preset_path.write_text(
                json.dumps(self._custom_pipe_presets, indent=2, sort_keys=True),
                encoding="utf-8",
            )
        except Exception as exc:
            QMessageBox.warning(self, "Preset Library", f"Failed to save presets: {exc}")

    def _capture_current_pipe_preset(self) -> Dict[str, Any]:
        L_si = float(self._length_from_display(self.in_L.value()))
        n_elbows = int(self.in_elbows.value())
        elbow_positions_frac = self._parse_elbow_positions(
            self.in_elbow_positions.text(),
            n_elbows,
            L_si,
        )
        n_tmass = int(self.in_tmass_count.value())
        tmass_positions_frac = self._parse_fractional_positions(
            self.in_tmass_positions.text(),
            n_tmass,
            L_si,
        )
        elbow_positions_text = ", ".join(f"{100.0 * v:.1f}%" for v in elbow_positions_frac) if n_elbows > 0 else ""
        tmass_positions_text = ", ".join(f"{100.0 * v:.1f}%" for v in tmass_positions_frac) if n_tmass > 0 else ""
        return {
            "pipe_material": self.pipe_material.currentText(),
            "insulation": bool(self.chk_use_insulation.isChecked()),
            "ins_material": self.ins_material.currentText(),
            "L": L_si,
            "Di": float(self._diam_from_display(self.in_Di.value())),
            "t_wall": float(self._diam_from_display(self.in_t_wall.value())),
            "t_ins": float(self._diam_from_display(self.in_t_ins.value())),
            "Tamb": float(self._temp_from_display(self.in_ambient.value())),
            "Tin": float(self._temp_from_display(self.in_Tin.value())),
            "Tin_ramp_s": float(self.in_tin_ramp.value()),
            "Tin_ramp_model": str(self.in_tin_ramp_model.currentData() or "logistic"),
            "p": float(self._pressure_from_display(self.in_p.value())),
            "m_dot": float(self._mdot_from_display(self.in_mdot.value())),
            "n_elbows": n_elbows,
            "elbow_sif": float(self.in_elbow_sif.value()),
            "elbow_positions": elbow_positions_text,
            "n_tmass": n_tmass,
            "tmass_factor": float(self.in_tmass_factor.value()),
            "tmass_positions": tmass_positions_text,
            "tmass_spread_pct": float(self.in_tmass_spread.value()),
            "tmass_deadleg_len_m": float(self._length_from_display(self.in_tmass_deadleg_len.value())),
            "tmass_deadleg_d_ratio": float(self.in_tmass_deadleg_d_ratio.value()),
        }

    def _save_current_pipe_preset(self):
        name = (self.new_preset_name.text() if hasattr(self, "new_preset_name") else "").strip()
        if not name:
            QMessageBox.warning(self, "Preset Library", "Preset name is required.")
            return
        if name in self.PIPE_DEFAULTS:
            QMessageBox.warning(self, "Preset Library", "Choose a new name (cannot overwrite built-in presets).")
            return
        self._custom_pipe_presets[name] = self._capture_current_pipe_preset()
        self._save_custom_pipe_presets()
        self._refresh_pipe_default_combo()
        self._refresh_pipe_preset_list()
        self.pipe_default.setCurrentText(name)
        self.new_preset_name.clear()
        QMessageBox.information(self, "Preset Library", f"Saved preset: {name}")

    def _selected_pipe_preset_name(self) -> str | None:
        if not hasattr(self, "pipe_preset_list"):
            return None
        item = self.pipe_preset_list.currentItem()
        if item is None:
            return None
        name = item.text().strip()
        return name or None

    def _apply_selected_pipe_preset(self):
        name = self._selected_pipe_preset_name()
        if not name:
            return
        self.pipe_default.setCurrentText(name)
        self._apply_pipe_default()

    def _delete_selected_pipe_preset(self):
        name = self._selected_pipe_preset_name()
        if not name:
            return
        if name in self.PIPE_DEFAULTS:
            QMessageBox.warning(self, "Preset Library", "Built-in presets cannot be deleted.")
            return
        if name in self._custom_pipe_presets:
            del self._custom_pipe_presets[name]
            self._save_custom_pipe_presets()
            self._refresh_pipe_default_combo()
            self._refresh_pipe_preset_list()
            QMessageBox.information(self, "Preset Library", f"Deleted preset: {name}")

    def _add_pipe_material(self):
        name = self.new_pipe_name.text().strip()
        if not name:
            QMessageBox.warning(self, "Material", "Pipe material name is required.")
            return
        PIPE_MATERIALS[name] = {
            "rho_w": self.new_pipe_rho.value(),
            "cp_w": self.new_pipe_cp.value(),
            "k_w": self.new_pipe_k.value(),
            "E": self.new_pipe_E.value() * 1.0e9,
            "alpha": self.new_pipe_alpha.value() * 1.0e-6,
            "nu": self.new_pipe_nu.value(),
            "Sy": self.new_pipe_sy.value() * 1.0e6,
            "eps_default": self.new_pipe_eps.value(),
        }
        PIPE_TEMP_PROPS[name] = self._flat_prop_table(cp=float(self.new_pipe_cp.value()), k=float(self.new_pipe_k.value()))
        self._refresh_material_combos()
        self.pipe_material.setCurrentText(name)
        self._refresh_material_lists()
        self._refresh_temp_prop_materials()
        QMessageBox.information(self, "Material", f"Saved pipe material: {name}")

    def _add_ins_material(self):
        name = self.new_ins_name.text().strip()
        if not name:
            QMessageBox.warning(self, "Material", "Insulation material name is required.")
            return
        INSULATION_MATERIALS[name] = {
            "rho_i": self.new_ins_rho.value(),
            "cp_i": self.new_ins_cp.value(),
            "k_i": self.new_ins_k.value(),
        }
        INSULATION_TEMP_PROPS[name] = self._flat_prop_table(cp=float(self.new_ins_cp.value()), k=float(self.new_ins_k.value()))
        self._refresh_material_combos()
        self.ins_material.setCurrentText(name)
        self._refresh_material_lists()
        self._refresh_temp_prop_materials()
        QMessageBox.information(self, "Material", f"Saved insulation material: {name}")

    def _sync_mode_widgets(self):
        mode_txt = self.mode.currentText()
        is_target_based = mode_txt in ("Outlet target", "Heatup-time optimize", "Stress-limit optimize")
        is_heatup_opt = mode_txt == "Heatup-time optimize"
        is_stress_opt = mode_txt == "Stress-limit optimize"
        self.in_target.setEnabled(is_target_based)
        self.target_metric.setEnabled(is_target_based)
        self.stop_dir.setEnabled(mode_txt == "Outlet target")
        self.in_heatup_target.setEnabled(is_heatup_opt)
        self.in_heatup_tol.setEnabled(is_heatup_opt)
        self.in_stress_limit.setEnabled(is_heatup_opt or is_stress_opt)
        self.in_mdot_min.setEnabled(is_heatup_opt or is_stress_opt)
        self.in_mdot_max.setEnabled(is_heatup_opt or is_stress_opt)

    def _sync_insulation_widgets(self):
        enabled = self.chk_use_insulation.isChecked()
        self.ins_material.setEnabled(enabled)
        self.in_t_ins.setEnabled(enabled)

    def _on_pipe_material_changed(self, name: str):
        mat = PIPE_MATERIALS.get(name)
        if not mat:
            return
        self.in_eps.setValue(float(mat["eps_default"]))

    def _temp_to_display(self, value):
        if self._units == "SI":
            return value
        return (np.asarray(value) - 273.15) * 9.0 / 5.0 + 32.0

    def _temp_from_display(self, value):
        if self._units == "SI":
            return value
        return (np.asarray(value) - 32.0) * 5.0 / 9.0 + 273.15

    def _inlet_temp_eff_series_si(self, times_s: np.ndarray) -> np.ndarray:
        t = np.asarray(times_s, dtype=float)
        if t.size == 0:
            return np.empty((0,), dtype=float)
        tin_target = float(self._last_run_si.get("Tin", self._temp_from_display(self.in_Tin.value())))
        tin_start = float(self._last_run_si.get("T_init_gas", self._last_run_si.get("Tamb", self._ambient_temp)))
        ramp_s = max(0.0, float(self._last_run_si.get("Tin_ramp_s", self.in_tin_ramp.value())))
        ramp_model = str(self._last_run_si.get("Tin_ramp_model", "logistic")).strip().lower()
        if ramp_s <= 1.0e-12:
            return np.full_like(t, tin_target, dtype=float)
        if ramp_model == "linear":
            frac = np.clip(t / ramp_s, 0.0, 1.0)
        elif ramp_model == "logistic":
            k = 8.0
            u = np.clip(t / ramp_s, 0.0, 1.0)
            lo = 1.0 / (1.0 + np.exp(0.5 * k))
            hi = 1.0 / (1.0 + np.exp(-0.5 * k))
            sig = 1.0 / (1.0 + np.exp(-k * (u - 0.5)))
            frac = np.clip((sig - lo) / np.maximum(hi - lo, 1.0e-12), 0.0, 1.0)
        else:
            frac = 1.0 - np.exp(-np.log(100.0) * np.maximum(t, 0.0) / ramp_s)
            frac = np.clip(frac, 0.0, 1.0)
        return tin_start + (tin_target - tin_start) * frac

    def _inlet_temp_eff_at_si(self, t_s: float) -> float:
        vals = self._inlet_temp_eff_series_si(np.asarray([float(t_s)], dtype=float))
        if vals.size == 0:
            return float(self._last_run_si.get("Tin", self._temp_from_display(self.in_Tin.value())))
        return float(vals[0])

    def _length_to_display(self, value_m):
        if self._units == "SI":
            return value_m
        return value_m * M2FT

    def _length_from_display(self, value_disp):
        if self._units == "SI":
            return value_disp
        return value_disp * FT2M

    def _diam_to_display(self, value_m):
        if self._units == "SI":
            return value_m
        return value_m * M2IN

    def _diam_from_display(self, value_disp):
        if self._units == "SI":
            return value_disp
        return value_disp * IN2M

    def _pressure_to_display(self, value_pa):
        if self._units == "SI":
            return value_pa
        return value_pa * PA2PSI

    def _pressure_from_display(self, value_disp):
        if self._units == "SI":
            return value_disp
        return value_disp * PSI2PA

    def _mdot_to_display(self, value_kgs):
        if self._units == "SI":
            return value_kgs
        return value_kgs * KG_S__TO__LBM_S

    def _mdot_from_display(self, value_disp):
        if self._units == "SI":
            return value_disp
        return value_disp * LBM_S__TO__KG_S

    def _stress_to_display(self, value_mpa):
        if self._units == "SI":
            return value_mpa
        return value_mpa * MPA_TO_KSI

    def _stress_from_display(self, value_disp):
        if self._units == "SI":
            return value_disp
        return value_disp * KSI_TO_MPA

    @staticmethod
    def _fmt_num(value: Any, decimals: int = 1, fallback: str = "--") -> str:
        try:
            v = float(value)
        except Exception:
            return fallback
        if not np.isfinite(v):
            return fallback
        return f"{v:.{max(0, int(decimals))}f}"

    def _temp_unit_label(self) -> str:
        return "K" if self._units == "SI" else "F"

    def _length_unit_label(self) -> str:
        return "m" if self._units == "SI" else "ft"

    def _stress_unit_label(self) -> str:
        return "MPa" if self._units == "SI" else "ksi"

    def _fmt_temp_si(self, value_k: Any, decimals: int = 1) -> str:
        disp = self._temp_to_display(float(value_k))
        return self._fmt_num(disp, decimals=decimals)

    def _fmt_length_si(self, value_m: Any, decimals: int = 3) -> str:
        disp = self._length_to_display(float(value_m))
        return self._fmt_num(disp, decimals=decimals)

    def _fmt_stress_si(self, value_mpa: Any, decimals: int = 2) -> str:
        disp = self._stress_to_display(float(value_mpa))
        return self._fmt_num(disp, decimals=decimals)

    def _fmt_mdot_si(self, value_kg_s: Any, decimals: int = 3) -> str:
        disp = self._mdot_to_display(float(value_kg_s))
        return self._fmt_num(disp, decimals=decimals)

    def _fmt_time_s(self, value_s: Any, decimals: int = 1) -> str:
        return self._fmt_num(value_s, decimals=decimals)

    def _update_live_readout(
        self,
        *,
        sim_time_s: Any | None = None,
        tg_out_k: Any | None = None,
        tw_out_k: Any | None = None,
        ti_out_k: Any | None = None,
        tin_eff_k: Any | None = None,
        tg_in_k: Any | None = None,
    ):
        unit = self._temp_unit_label()
        tg_out = self._fmt_temp_si(tg_out_k, 1) if tg_out_k is not None else "--"
        tw_out = self._fmt_temp_si(tw_out_k, 1) if tw_out_k is not None else "--"
        ti_out = self._fmt_temp_si(ti_out_k, 1) if ti_out_k is not None else "--"
        tin_eff = self._fmt_temp_si(tin_eff_k, 1) if tin_eff_k is not None else "--"
        tg_in = self._fmt_temp_si(tg_in_k, 1) if tg_in_k is not None else "--"
        t_txt = self._fmt_time_s(sim_time_s, 1) if sim_time_s is not None else "--"
        if hasattr(self, "lbl_live_readout"):
            self.lbl_live_readout.setText(
                f"Time: {t_txt} s || Tg_out {tg_out} | Tw_out {tw_out} | Ti_out {ti_out} | Tin_eff {tin_eff} | Tg_in {tg_in} {unit}"
            )

    def _target_series_from_outlet_components(
        self,
        tg_out_k: np.ndarray,
        tw_out_k: np.ndarray,
        ti_out_k: np.ndarray,
    ) -> np.ndarray:
        metric = _sanitize_target_metric(str(self._last_run_si.get("target_metric", "gas_outlet")))
        tg = np.asarray(tg_out_k, dtype=float)
        tw = np.asarray(tw_out_k, dtype=float)
        ti = np.asarray(ti_out_k, dtype=float)
        if metric == "gas_outlet":
            return tg
        if metric == "wall_inner_outlet":
            return tw
        if metric == "insulation_outlet":
            return ti
        wall_frac = _wall_fraction_from_geom(self._last_geom if self._last_geom else {"Di": 0.13, "t_wall": 0.018, "t_ins": 0.0, "k_w": 15.0, "k_i": 0.05})
        return tw - wall_frac * (tw - ti)

    def _update_target_time_readout(
        self,
        *,
        times_s: np.ndarray | None = None,
        tg_out_k: np.ndarray | None = None,
        tw_out_k: np.ndarray | None = None,
        ti_out_k: np.ndarray | None = None,
    ):
        if not hasattr(self, "lbl_target_time"):
            return
        mode_target = bool(float(self._last_run_si.get("mode_target", 0.0)) >= 0.5)
        if not mode_target:
            self.lbl_target_time.setText("Target time: n/a (fixed)")
            return
        target = float(self._last_run_si.get("target", np.nan))
        if not np.isfinite(target):
            self.lbl_target_time.setText("Target time: --")
            return

        t = np.asarray(times_s if times_s is not None else self._snap_t, dtype=float)
        tg = np.asarray(tg_out_k if tg_out_k is not None else self._snap_outlet, dtype=float)
        tw = np.asarray(tw_out_k if tw_out_k is not None else self._snap_tw_out, dtype=float)
        ti = np.asarray(ti_out_k if ti_out_k is not None else self._snap_ti_out, dtype=float)
        n = int(min(t.size, tg.size, tw.size, ti.size))
        if n == 0:
            self.lbl_target_time.setText("Target time: --")
            return
        t = t[:n]
        series = self._target_series_from_outlet_components(tg[:n], tw[:n], ti[:n])
        mode_le = bool(target <= float(series[0]))
        t_hit = self._target_crossing_time(t, series, target, mode_le=mode_le)
        if t_hit is not None:
            eta = max(0.0, float(t_hit) - float(t[-1]))
            self.lbl_target_time.setText(f"Target time: {self._fmt_time_s(t_hit)} s (ETA {self._fmt_time_s(eta)} s)")
            return

        if n < 3:
            self.lbl_target_time.setText("Target time: --")
            return
        k = min(8, n)
        tt = t[-k:]
        yy = series[-k:]
        dt = float(tt[-1] - tt[0])
        if dt <= 1.0e-9:
            self.lbl_target_time.setText("Target time: --")
            return
        slope = float((yy[-1] - yy[0]) / dt)
        if (mode_le and slope >= -1.0e-12) or ((not mode_le) and slope <= 1.0e-12):
            self.lbl_target_time.setText("Target time: --")
            return
        t_pred = float(tt[-1] + (target - yy[-1]) / slope)
        if not np.isfinite(t_pred) or t_pred < float(tt[-1]):
            self.lbl_target_time.setText("Target time: --")
            return
        eta = max(0.0, t_pred - float(tt[-1]))
        self.lbl_target_time.setText(f"Target time: ~{self._fmt_time_s(t_pred)} s (ETA {self._fmt_time_s(eta)} s)")

    @staticmethod
    def _round_ledger_value(key: str, value: Any) -> Any:
        if value is None:
            return None
        if isinstance(value, (int, np.integer, bool)):
            return int(value)
        if not isinstance(value, (float, np.floating)):
            return value
        v = float(value)
        if not np.isfinite(v):
            return ""
        k = key.lower()
        if k.endswith("_k") or "_temp_" in k or k.startswith(("tg_", "tw_", "ti_", "tin_", "tamb_", "target_")):
            return round(v, 1)
        if k.endswith("_s"):
            return round(v, 1)
        if k.endswith("_pa"):
            return round(v, 0)
        if k.endswith("_mpa") or k.endswith("_ksi"):
            return round(v, 2)
        if "mdot" in k or k.endswith("_kg_s"):
            return round(v, 3)
        if k.endswith("_m"):
            return round(v, 4)
        if "factor" in k or "ratio" in k:
            return round(v, 3)
        return round(v, 4)

    def _format_ledger_cell(self, key: str, value: Any) -> str:
        if value is None:
            return ""
        if isinstance(value, (int, np.integer, bool)):
            return str(int(value))
        if isinstance(value, (float, np.floating)):
            rounded = self._round_ledger_value(key, float(value))
            if rounded == "":
                return ""
            if isinstance(rounded, float) and rounded.is_integer() and key.lower().endswith("_pa"):
                return f"{int(rounded)}"
            return str(rounded)
        return str(value)

    @staticmethod
    def _set_box_safely(box: QDoubleSpinBox, value: float):
        box.blockSignals(True)
        box.setValue(float(value))
        box.blockSignals(False)

    def _apply_unit_ranges(self):
        if self._units == "SI":
            self.in_L.setRange(0.01, 1.0e6)
            self.in_Di.setRange(0.001, 1.0e3)
            self.in_t_wall.setRange(0.0001, 1.0)
            self.in_t_ins.setRange(0.0, 2.0)
            self.in_ambient.setRange(100.0, 2000.0)
            self.in_Tin.setRange(1.0, 5000.0)
            self.in_target.setRange(1.0, 5000.0)
            self.in_p.setRange(1.0, 1.0e9)
            self.in_mdot.setRange(0.0001, 1.0e4)
            self.in_mdot_min.setRange(0.0001, 1.0e4)
            self.in_mdot_max.setRange(0.0001, 1.0e4)
            self.in_stress_limit.setRange(1.0, 1.0e5)
            self.in_tmass_deadleg_len.setRange(0.0, 1.0e6)
        else:
            self.in_L.setRange(0.01 * M2FT, 1.0e6 * M2FT)
            self.in_Di.setRange(0.001 * M2IN, 1.0e3 * M2IN)
            self.in_t_wall.setRange(0.0001 * M2IN, 1.0 * M2IN)
            self.in_t_ins.setRange(0.0, 2.0 * M2IN)
            self.in_ambient.setRange(K_to_F(100.0), K_to_F(2000.0))
            self.in_Tin.setRange(K_to_F(1.0), K_to_F(5000.0))
            self.in_target.setRange(K_to_F(1.0), K_to_F(5000.0))
            self.in_p.setRange(1.0 * PA2PSI, 1.0e9 * PA2PSI)
            self.in_mdot.setRange(0.0001 * KG_S__TO__LBM_S, 1.0e4 * KG_S__TO__LBM_S)
            self.in_mdot_min.setRange(0.0001 * KG_S__TO__LBM_S, 1.0e4 * KG_S__TO__LBM_S)
            self.in_mdot_max.setRange(0.0001 * KG_S__TO__LBM_S, 1.0e4 * KG_S__TO__LBM_S)
            self.in_stress_limit.setRange(1.0 * MPA_TO_KSI, 1.0e5 * MPA_TO_KSI)
            self.in_tmass_deadleg_len.setRange(0.0, 1.0e6 * M2FT)

    def _apply_unit_labels(self):
        if self._units == "SI":
            self.in_L.setSuffix(" m")
            self.in_Di.setSuffix(" m")
            self.in_t_wall.setSuffix(" m")
            self.in_t_ins.setSuffix(" m")
            self.in_ambient.setSuffix(" K")
            self.in_Tin.setSuffix(" K")
            self.in_target.setSuffix(" K")
            self.in_p.setSuffix(" Pa")
            self.in_mdot.setSuffix(" kg/s")
            self.in_mdot_min.setSuffix(" kg/s")
            self.in_mdot_max.setSuffix(" kg/s")
            self.in_stress_limit.setSuffix(" MPa")
            self.in_tmass_deadleg_len.setSuffix(" m")
        else:
            self.in_L.setSuffix(" ft")
            self.in_Di.setSuffix(" in")
            self.in_t_wall.setSuffix(" in")
            self.in_t_ins.setSuffix(" in")
            self.in_ambient.setSuffix(" F")
            self.in_Tin.setSuffix(" F")
            self.in_target.setSuffix(" F")
            self.in_p.setSuffix(" psi")
            self.in_mdot.setSuffix(" lbm/s")
            self.in_mdot_min.setSuffix(" lbm/s")
            self.in_mdot_max.setSuffix(" lbm/s")
            self.in_stress_limit.setSuffix(" ksi")
            self.in_tmass_deadleg_len.setSuffix(" ft")
        # Keep temperature inputs and targets at engineering precision.
        for tbox in (self.in_ambient, self.in_Tin, self.in_target):
            tbox.setDecimals(1)
        if hasattr(self, "in_elbow_positions"):
            if self._units == "SI":
                self.in_elbow_positions.setPlaceholderText("auto or e.g. 45, 50 (m) or 20%, 60%")
            else:
                self.in_elbow_positions.setPlaceholderText("auto or e.g. 150, 180 (ft) or 20%, 60%")
        if hasattr(self, "in_tmass_positions"):
            if self._units == "SI":
                self.in_tmass_positions.setPlaceholderText("auto or e.g. 12, 35 (m) or 20%, 60%")
            else:
                self.in_tmass_positions.setPlaceholderText("auto or e.g. 40, 120 (ft) or 20%, 60%")
        if hasattr(self, "lbl_outlet"):
            self.lbl_outlet.setText(f"Outlet Tg: -- {'K' if self._units == 'SI' else 'F'}")
        if hasattr(self, "lbl_wall_out"):
            self.lbl_wall_out.setText(f"Outlet Tw: -- {'K' if self._units == 'SI' else 'F'}")
        if hasattr(self, "lbl_ins_out"):
            self.lbl_ins_out.setText(f"Outlet Ti: -- {'K' if self._units == 'SI' else 'F'}")
        if hasattr(self, "lbl_inlet"):
            self.lbl_inlet.setText(f"Tin_eff(bc): -- {'K' if self._units == 'SI' else 'F'}")
        if hasattr(self, "lbl_inlet_cell"):
            self.lbl_inlet_cell.setText(f"Tg_in(c0): -- {'K' if self._units == 'SI' else 'F'}")
        self._update_live_readout()
        self._update_target_time_readout()
        self._update_tmass_estimate_label()

    def _on_units_changed(self, new_units: str):
        old = self._units
        if new_units == old:
            return

        values = {
            "L": self.in_L.value(),
            "Di": self.in_Di.value(),
            "t_wall": self.in_t_wall.value(),
            "t_ins": self.in_t_ins.value(),
            "ambient": self.in_ambient.value(),
            "Tin": self.in_Tin.value(),
            "target": self.in_target.value(),
            "p": self.in_p.value(),
            "mdot": self.in_mdot.value(),
            "mdot_min": self.in_mdot_min.value(),
            "mdot_max": self.in_mdot_max.value(),
            "stress_limit": self.in_stress_limit.value(),
            "tmass_deadleg": self.in_tmass_deadleg_len.value(),
        }

        if old == "SI" and new_units == "Imperial":
            converted = {
                "L": values["L"] * M2FT,
                "Di": values["Di"] * M2IN,
                "t_wall": values["t_wall"] * M2IN,
                "t_ins": values["t_ins"] * M2IN,
                "ambient": K_to_F(values["ambient"]),
                "Tin": K_to_F(values["Tin"]),
                "target": K_to_F(values["target"]),
                "p": values["p"] * PA2PSI,
                "mdot": values["mdot"] * KG_S__TO__LBM_S,
                "mdot_min": values["mdot_min"] * KG_S__TO__LBM_S,
                "mdot_max": values["mdot_max"] * KG_S__TO__LBM_S,
                "stress_limit": values["stress_limit"] * MPA_TO_KSI,
                "tmass_deadleg": values["tmass_deadleg"] * M2FT,
            }
        elif old == "Imperial" and new_units == "SI":
            converted = {
                "L": values["L"] * FT2M,
                "Di": values["Di"] * IN2M,
                "t_wall": values["t_wall"] * IN2M,
                "t_ins": values["t_ins"] * IN2M,
                "ambient": F_to_K(values["ambient"]),
                "Tin": F_to_K(values["Tin"]),
                "target": F_to_K(values["target"]),
                "p": values["p"] * PSI2PA,
                "mdot": values["mdot"] * LBM_S__TO__KG_S,
                "mdot_min": values["mdot_min"] * LBM_S__TO__KG_S,
                "mdot_max": values["mdot_max"] * LBM_S__TO__KG_S,
                "stress_limit": values["stress_limit"] * KSI_TO_MPA,
                "tmass_deadleg": values["tmass_deadleg"] * FT2M,
            }
        else:
            return

        self._units = new_units
        self._apply_unit_ranges()
        self._apply_unit_labels()

        self._set_box_safely(self.in_L, converted["L"])
        self._set_box_safely(self.in_Di, converted["Di"])
        self._set_box_safely(self.in_t_wall, converted["t_wall"])
        self._set_box_safely(self.in_t_ins, converted["t_ins"])
        self._set_box_safely(self.in_ambient, converted["ambient"])
        self._set_box_safely(self.in_Tin, converted["Tin"])
        self._set_box_safely(self.in_target, converted["target"])
        self._set_box_safely(self.in_p, converted["p"])
        self._set_box_safely(self.in_mdot, converted["mdot"])
        self._set_box_safely(self.in_mdot_min, converted["mdot_min"])
        self._set_box_safely(self.in_mdot_max, converted["mdot_max"])
        self._set_box_safely(self.in_stress_limit, converted["stress_limit"])
        self._set_box_safely(self.in_tmass_deadleg_len, converted["tmass_deadleg"])

        if HAS_MPL and self._play_times.size > 0:
            idx = self.slider_time.value() if self.slider_time.maximum() >= 0 else 0
            self._render_playback_frame(idx)
            if self._last_result is not None:
                self._render_static_results(self._last_result)
            self.lbl_outlet.setText(f"Outlet Tg: {self._fmt_temp_si(self._play_tg[idx, -1])} {self._temp_unit_label()}")
            self.lbl_wall_out.setText(f"Outlet Tw: {self._fmt_temp_si(self._play_tw[idx, -1])} {self._temp_unit_label()}")
            self.lbl_ins_out.setText(f"Outlet Ti: {self._fmt_temp_si(self._play_ti[idx, -1])} {self._temp_unit_label()}")
            if self._play_tin_eff.size > idx:
                self.lbl_inlet.setText(f"Tin_eff(bc): {self._fmt_temp_si(self._play_tin_eff[idx])} {self._temp_unit_label()}")
            if self._play_tg.shape[0] > idx:
                self.lbl_inlet_cell.setText(
                    f"Tg_in(c0): {self._fmt_temp_si(self._play_tg[idx, 0])} {self._temp_unit_label()}"
                )
            self._update_live_readout(
                sim_time_s=self._play_times[idx] if self._play_times.size > idx else None,
                tg_out_k=self._play_tg[idx, -1],
                tw_out_k=self._play_tw[idx, -1],
                ti_out_k=self._play_ti[idx, -1],
                tin_eff_k=self._play_tin_eff[idx] if self._play_tin_eff.size > idx else None,
                tg_in_k=self._play_tg[idx, 0],
            )
        else:
            self._reset_live_views()

    def _choose_dir(self):
        path = QFileDialog.getExistingDirectory(self, "Select save folder")
        if path:
            self._save_dir = Path(path)
            self.dir_label.setText(self._save_dir.name or str(self._save_dir))
            self.dir_label.setToolTip(str(self._save_dir))
        else:
            self._save_dir = None
            self.dir_label.setText("(auto)")
            self.dir_label.setToolTip("")
        self._refresh_ledger_preview()

    def _choose_ledger_file(self):
        path, _ = QFileDialog.getSaveFileName(
            self,
            "Select or create run ledger",
            str((self._save_dir or Path.cwd()) / "run_history.csv"),
            "CSV files (*.csv);;Excel Workbook (*.xlsx)",
        )
        if path:
            self._ledger_path = Path(path)
            self.ledger_label.setText(self._ledger_path.name)
            self.ledger_label.setToolTip(str(self._ledger_path))
        else:
            self._ledger_path = None
            self.ledger_label.setText("(auto)")
            self.ledger_label.setToolTip(str(self._resolve_ledger_path()))
        self._refresh_ledger_preview()

    def _resolve_ledger_path(self) -> Path:
        if self._ledger_path is not None:
            return self._ledger_path
        base = self._save_dir if self._save_dir is not None else Path.cwd()
        return base / "run_history.csv"

    def _refresh_readme_view(self):
        if not hasattr(self, "readme_view"):
            return
        self.lbl_readme_path.setText(str(self._readme_path))
        if not self._readme_path.exists():
            self.readme_view.setPlainText(f"README not found at: {self._readme_path}")
            return
        try:
            text = self._readme_path.read_text(encoding="utf-8")
        except Exception as exc:
            self.readme_view.setPlainText(f"Failed to read README: {exc}")
            return
        try:
            self.readme_view.setMarkdown(text)
        except Exception:
            self.readme_view.setPlainText(text)

    @staticmethod
    def _merge_ledger_headers(existing: list[str], incoming: list[str]) -> list[str]:
        merged: list[str] = [str(h).strip() for h in existing if str(h).strip()]
        for key in incoming:
            k = str(key).strip()
            if k and k not in merged:
                merged.append(k)
        return merged

    @staticmethod
    def _normalize_legacy_ledger_row(values: list[Any], header_len: int) -> list[Any]:
        # Legacy schema lacked asset_id/branch_id; newer rows may already contain both fields.
        if not values:
            return ["", "", ""]
        if len(values) >= header_len + 2:
            return [values[0], values[1], values[2]] + values[3:]
        return [values[0], "", ""] + values[1:]

    def _load_csv_ledger_rows(self, ledger_path: Path) -> tuple[list[str], list[dict[str, Any]]]:
        with ledger_path.open("r", encoding="utf-8", newline="") as f:
            raw = list(csv.reader(f))
        if not raw:
            return [], []

        header_base = [str(v).strip() for v in raw[0] if str(v).strip()]
        if not header_base:
            return [], []

        legacy_schema = (
            header_base[0] == "timestamp_local"
            and "asset_id" not in header_base
            and "branch_id" not in header_base
        )
        header_len_base = len(header_base)
        headers = [header_base[0], "asset_id", "branch_id"] + header_base[1:] if legacy_schema else list(header_base)

        rows: list[dict[str, Any]] = []
        extra_headers: list[str] = []
        for raw_row in raw[1:]:
            vals: list[Any] = list(raw_row)
            if not any(str(v).strip() for v in vals):
                continue
            if legacy_schema:
                vals = self._normalize_legacy_ledger_row(vals, header_len_base)
            if len(vals) < len(headers):
                vals.extend([""] * (len(headers) - len(vals)))

            row_map: dict[str, Any] = {h: vals[i] if i < len(vals) else "" for i, h in enumerate(headers)}
            if len(vals) > len(headers):
                for k, extra in enumerate(vals[len(headers):], start=1):
                    extra_name = f"extra_{k}"
                    if extra_name not in extra_headers:
                        extra_headers.append(extra_name)
                    row_map[extra_name] = extra
            rows.append(row_map)

        headers = headers + [h for h in extra_headers if h not in headers]
        return headers, rows

    @staticmethod
    def _write_csv_ledger_rows(ledger_path: Path, headers: list[str], rows: list[dict[str, Any]]) -> None:
        with ledger_path.open("w", encoding="utf-8", newline="") as f:
            writer = csv.DictWriter(f, fieldnames=headers)
            writer.writeheader()
            for row in rows:
                writer.writerow({h: row.get(h, "") for h in headers})

    def _load_xlsx_ledger_rows(self, ledger_path: Path) -> tuple[list[str], list[dict[str, Any]]]:
        try:
            from openpyxl import load_workbook  # type: ignore
        except Exception as exc:  # pragma: no cover - optional dependency
            raise RuntimeError(f"openpyxl required for XLSX support: {exc}") from exc

        wb = load_workbook(ledger_path)
        ws = wb.active
        data = list(ws.iter_rows(values_only=True))
        if not data:
            return [], []

        header_base = [str(v).strip() if v is not None else "" for v in data[0]]
        while header_base and not header_base[-1]:
            header_base.pop()
        if not header_base:
            return [], []

        legacy_schema = (
            header_base[0] == "timestamp_local"
            and "asset_id" not in header_base
            and "branch_id" not in header_base
        )
        header_len_base = len(header_base)
        headers = [header_base[0], "asset_id", "branch_id"] + header_base[1:] if legacy_schema else list(header_base)

        rows: list[dict[str, Any]] = []
        extra_headers: list[str] = []
        for raw_row in data[1:]:
            vals: list[Any] = [v for v in raw_row]
            while vals and vals[-1] is None:
                vals.pop()
            if not any(v is not None and str(v).strip() for v in vals):
                continue
            if legacy_schema:
                vals = self._normalize_legacy_ledger_row(vals, header_len_base)
            if len(vals) < len(headers):
                vals.extend([""] * (len(headers) - len(vals)))

            row_map: dict[str, Any] = {h: vals[i] if i < len(vals) else "" for i, h in enumerate(headers)}
            if len(vals) > len(headers):
                for k, extra in enumerate(vals[len(headers):], start=1):
                    extra_name = f"extra_{k}"
                    if extra_name not in extra_headers:
                        extra_headers.append(extra_name)
                    row_map[extra_name] = extra
            rows.append(row_map)

        headers = headers + [h for h in extra_headers if h not in headers]
        return headers, rows

    @staticmethod
    def _write_xlsx_ledger_rows(ledger_path: Path, headers: list[str], rows: list[dict[str, Any]]) -> None:
        try:
            from openpyxl import Workbook  # type: ignore
        except Exception as exc:  # pragma: no cover - optional dependency
            raise RuntimeError(f"openpyxl required for XLSX support: {exc}") from exc

        wb = Workbook()
        ws = wb.active
        ws.title = "run_history"
        ws.append(headers)
        for row in rows:
            ws.append([row.get(h, "") for h in headers])
        wb.save(ledger_path)

    def _refresh_ledger_preview(self):
        if not hasattr(self, "ledger_table"):
            return
        ledger_path = self._resolve_ledger_path()
        self.lbl_ledger_status.setText(f"Ledger: {ledger_path.name}")
        self.lbl_ledger_status.setToolTip(str(ledger_path))
        self.ledger_table.clear()
        self.ledger_table.setRowCount(0)
        self.ledger_table.setColumnCount(0)

        if not ledger_path.exists():
            self.ledger_table.setRowCount(1)
            self.ledger_table.setColumnCount(1)
            self.ledger_table.setHorizontalHeaderLabels(["Info"])
            self.ledger_table.setItem(0, 0, QTableWidgetItem("Ledger file does not exist yet."))
            return

        try:
            if ledger_path.suffix.lower() == ".xlsx":
                headers, rows = self._load_xlsx_ledger_rows(ledger_path)
            else:
                headers, rows = self._load_csv_ledger_rows(ledger_path)
        except Exception as exc:
            self.ledger_table.setRowCount(1)
            self.ledger_table.setColumnCount(1)
            self.ledger_table.setHorizontalHeaderLabels(["Info"])
            self.ledger_table.setItem(0, 0, QTableWidgetItem(str(exc)))
            return

        self._ledger_rows = rows
        self._ledger_fieldnames = headers
        self.lbl_ledger_status.setText(f"Ledger: {ledger_path.name} ({len(rows)} rows)")
        display_headers = self._select_ledger_display_headers(headers)
        self.ledger_table.setColumnCount(len(display_headers))
        self.ledger_table.setHorizontalHeaderLabels(display_headers)
        self.ledger_table.setRowCount(len(rows))
        for i, row in enumerate(rows):
            for j, h in enumerate(display_headers):
                self.ledger_table.setItem(i, j, QTableWidgetItem(self._format_ledger_cell(h, row.get(h, ""))))
        self._set_ledger_column_widths(display_headers)

    @staticmethod
    def _select_ledger_display_headers(headers: list[str]) -> list[str]:
        preferred = [
            "timestamp_local",
            "asset_id",
            "branch_id",
            "mode",
            "target_metric",
            "stop_reason",
            "reached_target",
            "m_dot_kg_s",
            "Tg_out_final_K",
            "time_to_target_s",
            "max_vm_elbow_mpa",
            "t_max_vm_s",
            "x_max_vm_m",
            "warnings",
        ]
        out = [h for h in preferred if h in headers]
        if out:
            return out
        return headers[: min(12, len(headers))]

    def _set_ledger_column_widths(self, headers: list[str]):
        if not hasattr(self, "ledger_table"):
            return
        for j, h in enumerate(headers):
            key = str(h).strip().lower()
            if key in ("timestamp_local", "mode", "stop_reason", "asset_id", "branch_id"):
                w = 170
            elif "warnings" in key:
                w = 320
            elif key.endswith("_id"):
                w = 140
            else:
                w = 130
            self.ledger_table.setColumnWidth(j, w)

    def _append_current_config_to_ledger(self):
        try:
            _ = self._collect_spec()
        except Exception:
            pass

        class _Result:
            pass

        r = _Result()
        r.stop_reason = "manual_config_append"
        r.reached_Tg_target = False
        r.Tg_outlet_final = float(self._last_stats.get("Tg_outlet_final_K", np.nan))
        self._append_run_ledger(r)
        self._refresh_ledger_preview()

    def _delete_selected_ledger_rows(self):
        ledger_path = self._resolve_ledger_path()
        if not ledger_path.exists():
            return
        selected = sorted({idx.row() for idx in self.ledger_table.selectionModel().selectedRows()})
        if not selected:
            return
        try:
            if ledger_path.suffix.lower() == ".xlsx":
                headers, rows = self._load_xlsx_ledger_rows(ledger_path)
                keep = [row for i, row in enumerate(rows) if i not in set(selected)]
                self._write_xlsx_ledger_rows(ledger_path, headers, keep)
            else:
                headers, rows = self._load_csv_ledger_rows(ledger_path)
                keep = [row for i, row in enumerate(rows) if i not in set(selected)]
                self._write_csv_ledger_rows(ledger_path, headers, keep)
        except Exception as exc:
            QMessageBox.warning(self, "Ledger", f"Could not edit ledger: {exc}")
            return
        self._refresh_ledger_preview()

    def _parse_fractional_positions(self, text: str, count: int, length_si: float) -> list[float]:
        n = int(max(0, count))
        if n == 0:
            return []
        raw = (text or "").strip().lower()
        if raw in ("", "auto", "default"):
            return [float(v) for v in np.linspace(0.15, 0.85, n)]
        length_disp = float(self._length_to_display(max(length_si, 1.0e-12)))
        vals: list[float] = []
        for token in raw.replace(";", ",").split(","):
            tok = token.strip()
            if not tok:
                continue
            is_percent = "%" in tok
            tok = tok.replace("%", "")
            try:
                v = float(tok)
            except ValueError:
                continue

            if is_percent:
                frac = v / 100.0
            elif abs(v) <= 1.0:
                frac = v
            else:
                # Plain numeric entries are interpreted as absolute distance in current length units.
                frac = v / max(length_disp, 1.0e-12)

            vals.append(max(0.0, min(1.0, float(frac))))
        if len(vals) < n:
            auto = [float(v) for v in np.linspace(0.15, 0.85, n)]
            vals.extend(auto[len(vals):])
        return vals[:n]

    def _parse_elbow_positions(self, text: str, n_elbows: int, length_si: float) -> list[float]:
        return self._parse_fractional_positions(text, n_elbows, length_si)

    def _current_target_metric(self) -> str:
        idx = self.target_metric.currentIndex()
        data = self.target_metric.itemData(idx)
        return _sanitize_target_metric(str(data))

    @staticmethod
    def _flat_prop_table(cp: float, k: float) -> Dict[str, list[float]]:
        return {"T": [300.0, 600.0, 900.0, 1200.0], "cp": [float(cp)] * 4, "k": [float(k)] * 4}

    def _resolve_pipe_prop_table(self, mat_name: str, mat: Dict[str, float]) -> Dict[str, list[float]]:
        table = PIPE_TEMP_PROPS.get(mat_name)
        if isinstance(table, dict):
            return table
        return self._flat_prop_table(cp=float(mat["cp_w"]), k=float(mat["k_w"]))

    def _resolve_ins_prop_table(self, mat_name: str, mat: Dict[str, float]) -> Dict[str, list[float]]:
        table = INSULATION_TEMP_PROPS.get(mat_name)
        if isinstance(table, dict):
            return table
        return self._flat_prop_table(cp=float(mat["cp_i"]), k=float(mat["k_i"]))

    def _apply_pipe_default(self):
        name = self.pipe_default.currentText()
        cfg = self._all_pipe_defaults().get(name, {})
        if not cfg:
            return
        if "pipe_material" in cfg and cfg["pipe_material"] in PIPE_MATERIALS:
            self.pipe_material.setCurrentText(str(cfg["pipe_material"]))
        if "ins_material" in cfg and cfg["ins_material"] in INSULATION_MATERIALS:
            self.ins_material.setCurrentText(str(cfg["ins_material"]))
        if "insulation" in cfg:
            self.chk_use_insulation.setChecked(bool(cfg["insulation"]))

        if "L" in cfg:
            self._set_box_safely(self.in_L, float(self._length_to_display(float(cfg["L"]))))
        if "Di" in cfg:
            self._set_box_safely(self.in_Di, float(self._diam_to_display(float(cfg["Di"]))))
        if "t_wall" in cfg:
            self._set_box_safely(self.in_t_wall, float(self._diam_to_display(float(cfg["t_wall"]))))
        if "t_ins" in cfg:
            self._set_box_safely(self.in_t_ins, float(self._diam_to_display(float(cfg["t_ins"]))))
        if "Tamb" in cfg:
            self._set_box_safely(self.in_ambient, float(self._temp_to_display(float(cfg["Tamb"]))))
        if "Tin" in cfg:
            self._set_box_safely(self.in_Tin, float(self._temp_to_display(float(cfg["Tin"]))))
        if "Tin_ramp_s" in cfg:
            self._set_box_safely(self.in_tin_ramp, float(cfg["Tin_ramp_s"]))
        if "Tin_ramp_model" in cfg:
            key = str(cfg["Tin_ramp_model"]).strip().lower()
            idx = self.in_tin_ramp_model.findData(key)
            if idx >= 0:
                self.in_tin_ramp_model.setCurrentIndex(idx)
        if "p" in cfg:
            self._set_box_safely(self.in_p, float(self._pressure_to_display(float(cfg["p"]))))
        if "m_dot" in cfg:
            self._set_box_safely(self.in_mdot, float(self._mdot_to_display(float(cfg["m_dot"]))))
        if "n_elbows" in cfg:
            self.in_elbows.setValue(int(cfg["n_elbows"]))
        if "elbow_sif" in cfg:
            self._set_box_safely(self.in_elbow_sif, float(cfg["elbow_sif"]))
        self.in_elbow_positions.setText(str(cfg.get("elbow_positions", cfg.get("elbow_positions_pct", ""))))
        if "n_tmass" in cfg:
            self.in_tmass_count.setValue(int(cfg["n_tmass"]))
        if "tmass_factor" in cfg:
            self._set_box_safely(self.in_tmass_factor, float(cfg["tmass_factor"]))
        if "tmass_spread_pct" in cfg:
            self._set_box_safely(self.in_tmass_spread, float(cfg["tmass_spread_pct"]))
        if "tmass_deadleg_len_m" in cfg:
            self._set_box_safely(self.in_tmass_deadleg_len, float(self._length_to_display(float(cfg["tmass_deadleg_len_m"]))))
        if "tmass_deadleg_d_ratio" in cfg:
            self._set_box_safely(self.in_tmass_deadleg_d_ratio, float(cfg["tmass_deadleg_d_ratio"]))
        self.in_tmass_positions.setText(str(cfg.get("tmass_positions", cfg.get("tmass_positions_pct", ""))))
        self._sync_insulation_widgets()
        self._update_tmass_estimate_label()

    def _apply_preset(self, name: str):
        cfg = self.PRESETS.get(name)
        if not cfg:
            return
        self.in_Nx.setValue(cfg["Nx"])
        self.in_save_frames.setValue(cfg["save_frames"])
        self.in_dt_max.setValue(cfg["dt_max"])
        self.in_dt_min.setValue(cfg["dt_min"])
        self.in_update_props.setValue(cfg["update_props_every"])
        if "adv_scheme" in cfg and hasattr(self, "in_adv_scheme"):
            idx = self.in_adv_scheme.findData(cfg["adv_scheme"])
            if idx >= 0:
                self.in_adv_scheme.setCurrentIndex(idx)
        if "semi_lag_courant_max" in cfg and hasattr(self, "in_semi_lag_cmax"):
            self._set_box_safely(self.in_semi_lag_cmax, float(cfg["semi_lag_courant_max"]))
        if "use_float32" in cfg and hasattr(self, "chk_use_float32"):
            self.chk_use_float32.setChecked(bool(cfg["use_float32"]))

    def _collect_spec(self) -> RunSpec:
        mode_ui = self.mode.currentText()
        mode = "time" if mode_ui == "Fixed time" else "target"

        pipe_mat = PIPE_MATERIALS[self.pipe_material.currentText()]
        if self.chk_use_insulation.isChecked():
            ins_mat = INSULATION_MATERIALS[self.ins_material.currentText()]
            t_ins_disp = self.in_t_ins.value()
        else:
            ins_mat = {"rho_i": 128.0, "cp_i": 840.0, "k_i": 0.045}
            t_ins_disp = 0.0

        L = self._length_from_display(self.in_L.value())
        Di = self._diam_from_display(self.in_Di.value())
        t_wall = self._diam_from_display(self.in_t_wall.value())
        t_ins = self._diam_from_display(t_ins_disp)
        ambient = float(self._temp_from_display(self.in_ambient.value()))
        Tin = float(self._temp_from_display(self.in_Tin.value()))
        target = float(self._temp_from_display(self.in_target.value()))
        target_metric = self._current_target_metric()
        p = float(self._pressure_from_display(self.in_p.value()))
        m_dot = float(self._mdot_from_display(self.in_mdot.value()))

        self._ambient_temp = ambient
        self._last_pipe_material = self.pipe_material.currentText()
        elbow_positions_frac = self._parse_elbow_positions(
            self.in_elbow_positions.text(),
            int(self.in_elbows.value()),
            L,
        )
        tmass_count = int(self.in_tmass_count.value())
        tmass_positions_frac = self._parse_fractional_positions(
            self.in_tmass_positions.text(),
            tmass_count,
            L,
        )
        self._last_mech = {
            "E": pipe_mat["E"],
            "alpha": pipe_mat["alpha"],
            "nu": pipe_mat["nu"],
            "Sy": float(pipe_mat.get("Sy", 250.0e6)),
        }
        self._last_geom = {
            "Di": Di,
            "t_wall": t_wall,
            "t_ins": t_ins,
            "k_w": float(pipe_mat["k_w"]),
            "k_i": float(ins_mat["k_i"]),
            "n_elbows": int(self.in_elbows.value()),
            "elbow_sif": float(self.in_elbow_sif.value()),
            "elbow_positions_frac": elbow_positions_frac,
            "elbow_positions_input": self.in_elbow_positions.text().strip(),
            "n_tmass": tmass_count,
            "tmass_factor": float(self.in_tmass_factor.value()),
            "tmass_positions_frac": tmass_positions_frac,
            "tmass_positions_input": self.in_tmass_positions.text().strip(),
            "tmass_spread_frac": 0.01 * float(self.in_tmass_spread.value()),
            "tmass_deadleg_len_m": float(self._length_from_display(self.in_tmass_deadleg_len.value())),
            "tmass_deadleg_d_ratio": float(self.in_tmass_deadleg_d_ratio.value()),
        }
        self._last_run_si = {
            "p": p,
            "Tin": Tin,
            "Tamb": ambient,
            "m_dot": m_dot,
            "T_init_gas": ambient,
            "T_init_wall": ambient,
            "T_init_ins": ambient,
            "L": L,
            "Di": Di,
            "t_wall": t_wall,
            "t_ins": t_ins,
            "Nx": float(self.in_Nx.value()),
            "nr_wall": float(self.in_nr_wall.value()),
            "axial_restraint": float(self.in_axial_restraint.value()),
            "ignore_inlet_cells": float(self.in_ignore_inlet_cells.value()),
            "Tin_ramp_s": float(self.in_tin_ramp.value()),
            "Tin_ramp_model": str(self.in_tin_ramp_model.currentData() or "logistic"),
            "target": target,
            "target_metric": target_metric,
            "target_metric_label": TARGET_METRIC_LABELS.get(target_metric, target_metric),
            "mode_target": 1.0 if mode == "target" else 0.0,
            "asset_id": self.asset_id.text().strip(),
            "branch_id": self.branch_id.text().strip(),
        }

        hardware = HardwareConfig(
            L=L,
            Di=Di,
            t_wall=t_wall,
            t_ins=t_ins,
            h_out=0.0,
            eps_rad=self.in_eps.value(),
        )

        run_inputs = RunInputs(
            p=p,
            Tin=Tin,
            m_dot=m_dot,
            mode=mode,
            t_end=self.in_t_end.value(),
            Tg_out_target=target if mode == "target" else None,
            stop_dir=None if self.stop_dir.currentText() == "auto" else self.stop_dir.currentText(),
        )

        overrides = {
            "Nx": self.in_Nx.value(),
            "save_frames": self.in_save_frames.value(),
            "dt_max": self.in_dt_max.value(),
            "dt_min": self.in_dt_min.value(),
            "Tin_ramp_s": self.in_tin_ramp.value(),
            "Tin_ramp_model": str(self.in_tin_ramp_model.currentData() or "logistic"),
            "update_props_every": self.in_update_props.value(),
            "adv_scheme": str(self.in_adv_scheme.currentData() or "semi_lagrangian"),
            "semi_lag_courant_max": float(self.in_semi_lag_cmax.value()),
            "progress": self.progress.currentText(),
            "use_float32": self.chk_use_float32.isChecked(),
            "enable_numba": (False if sys.platform.startswith("win") else bool(HAS_NUMBA)),
            "log_to_file": self.chk_log_to_file.isChecked(),
            "write_trace_csv": self.chk_write_trace.isChecked(),
            "target_asymptote_check": self.chk_target_asymptote.isChecked(),
            # GUI worker-thread runs must not call pyplot popups directly.
            "show_plots": False,
            "rho_w": pipe_mat["rho_w"],
            "cp_w": pipe_mat["cp_w"],
            "k_w": pipe_mat["k_w"],
            "rho_i": ins_mat["rho_i"],
            "cp_i": ins_mat["cp_i"],
            "k_i": ins_mat["k_i"],
            "pipe_prop_table": self._resolve_pipe_prop_table(self.pipe_material.currentText(), pipe_mat),
            "ins_prop_table": self._resolve_ins_prop_table(self.ins_material.currentText(), ins_mat),
            "use_temp_dependent_props": bool(self.chk_temp_dep_props.isChecked()),
            "target_metric": target_metric,
            "thermal_mass_count": tmass_count,
            "thermal_mass_factor": float(self.in_tmass_factor.value()),
            "thermal_mass_positions_frac": tmass_positions_frac,
            "thermal_mass_spread_frac": 0.01 * float(self.in_tmass_spread.value()),
            "Tamb": ambient,
            "eps_rad": self.in_eps.value(),
            "T_init_wall": ambient,
            "T_init_ins": ambient,
            "T_init_gas": ambient,
            "h_out_mode": "auto",
            "h_out": 0.0,
            "insulation_mass_mode": str(self.in_ins_mass_mode.currentData() or "penetration"),
            "insulation_mass_min_frac": float(self.in_ins_mass_min_frac.value()),
        }

        return RunSpec(
            hardware=hardware,
            run=run_inputs,
            overrides=overrides,
            save_dir=self._save_dir,
            # Save/export images is handled in the Qt main thread after completion.
            make_plots=False,
            save_results=self.chk_save_results.isChecked(),
        )

    def _collect_optimization_config(self) -> Dict[str, Any] | None:
        mode_ui = self.mode.currentText()
        if mode_ui not in ("Heatup-time optimize", "Stress-limit optimize"):
            return None

        mdot_min = float(self._mdot_from_display(self.in_mdot_min.value()))
        mdot_max = float(self._mdot_from_display(self.in_mdot_max.value()))
        if mdot_min <= 0.0 or mdot_max <= 0.0 or mdot_max <= mdot_min:
            raise ValueError("Search mass flow bounds must satisfy 0 < m_dot_min < m_dot_max.")

        stress_limit_mpa = float(self._stress_from_display(self.in_stress_limit.value()))
        if stress_limit_mpa <= 0.0:
            raise ValueError("Stress limit must be > 0.")

        target_temp = float(self._temp_from_display(self.in_target.value()))
        target_metric = self._current_target_metric()
        cfg: Dict[str, Any] = {
            "mode": "heatup_time_opt" if mode_ui == "Heatup-time optimize" else "stress_limit_opt",
            "target_temp_K": target_temp,
            "target_metric": target_metric,
            "mdot_min_kg_s": mdot_min,
            "mdot_max_kg_s": mdot_max,
            "stress_limit_mpa": stress_limit_mpa,
            "include_pressure": bool(self.chk_include_pressure.isChecked()),
            "nr_wall": int(self.in_nr_wall.value()),
            "geom": dict(self._last_geom),
            "run_si": dict(self._last_run_si),
            "mech": dict(self._last_mech),
            "search_nx": int(max(80, min(int(self.in_Nx.value()), max(120, int(0.35 * self.in_Nx.value()))))),
            "opt_t_end_s": float(self.in_t_end.value()),
            "coarse_points": 9,
            "refine_iters": 6,
            "bisection_iters": 12,
            "stress_tol_mpa": 2.0,
        }
        if cfg["mode"] == "heatup_time_opt":
            cfg["heatup_target_s"] = float(self.in_heatup_target.value())
            cfg["heatup_tol_s"] = float(self.in_heatup_tol.value())
            cfg["opt_t_end_s"] = max(
                float(self.in_t_end.value()),
                float(self.in_heatup_target.value()) + max(60.0, 3.0 * float(self.in_heatup_tol.value())),
            )
        return cfg

    def _clear_heat_colorbar(self):
        if self._heat_cbar is not None:
            try:
                self._heat_cbar.remove()
            except Exception:
                pass
            self._heat_cbar = None

        if HAS_MPL and hasattr(self, "heat_fig") and hasattr(self, "ax_heat"):
            # Colorbars add extra axes; keep only the main heatmap axis.
            for ax in list(self.heat_fig.axes):
                if ax is not self.ax_heat:
                    try:
                        self.heat_fig.delaxes(ax)
                    except Exception:
                        pass

    def _update_live_heatmap(self, heat: np.ndarray, extent: list[float]):
        if not HAS_MPL:
            return
        if heat.size == 0:
            return

        if self._heat_im is None or self._heat_im.axes is not self.ax_heat:
            self._heat_im = self.ax_heat.imshow(heat, aspect="auto", extent=extent, origin="upper")
            self._clear_heat_colorbar()
            self._heat_cbar = self.heat_fig.colorbar(
                self._heat_im,
                ax=self.ax_heat,
                label=("K" if self._units == "SI" else "F"),
            )
        else:
            self._heat_im.set_data(heat)
            self._heat_im.set_extent(extent)
            vmin = float(np.nanmin(heat))
            vmax = float(np.nanmax(heat))
            if vmax <= vmin:
                vmax = vmin + 1.0e-6
            self._heat_im.set_clim(vmin, vmax)
            if self._heat_cbar is not None:
                self._heat_cbar.set_label("K" if self._units == "SI" else "F")
                self._heat_cbar.update_normal(self._heat_im)
        self.heat_canvas.draw_idle()

    def _reset_live_views(self):
        self._snap_t.clear()
        self._snap_outlet.clear()
        self._snap_tw_out.clear()
        self._snap_ti_out.clear()
        self._snap_inlet.clear()
        self._snap_inlet_cell.clear()
        self._snap_tg_rows.clear()
        self._snapshot_counter = 0
        self.lbl_runtime.setText("Sim time: 0.0 s")
        self.lbl_outlet.setText(f"Outlet Tg: -- {'K' if self._units == 'SI' else 'F'}")
        self.lbl_wall_out.setText(f"Outlet Tw: -- {'K' if self._units == 'SI' else 'F'}")
        self.lbl_ins_out.setText(f"Outlet Ti: -- {'K' if self._units == 'SI' else 'F'}")
        self.lbl_inlet.setText(f"Tin_eff(bc): -- {'K' if self._units == 'SI' else 'F'}")
        self.lbl_inlet_cell.setText(f"Tg_in(c0): -- {'K' if self._units == 'SI' else 'F'}")
        self._update_live_readout()
        if hasattr(self, "lbl_target_time"):
            self.lbl_target_time.setText("Target time: --")
        self.lbl_frames.setText("Frames: 0")
        self._play_timer.stop()

        self._play_times = np.array([], dtype=float)
        self._play_tg = np.empty((0, 0), dtype=float)
        self._play_tw = np.empty((0, 0), dtype=float)
        self._play_ti = np.empty((0, 0), dtype=float)
        self._play_tin_eff = np.array([], dtype=float)
        self.slider_time.setEnabled(False)
        self.slider_time.setRange(0, 0)
        self.lbl_time_cursor.setText("t = -- s / -- s")

        if not HAS_MPL:
            return

        self._clear_heat_colorbar()

        self.ax_outlet.clear()
        self.ax_outlet.set_title("Outlet and Inlet Setpoint vs Simulation Time")
        self.ax_outlet.set_xlabel("time [s]")
        self.ax_outlet.set_ylabel(f"Temperature [{'K' if self._units == 'SI' else 'F'}]")
        (self.line_outlet,) = self.ax_outlet.plot([], [], color="#2a9d8f", linewidth=2.0, label="Tg_out")
        (self.line_inlet,) = self.ax_outlet.plot([], [], color="#6c757d", linewidth=1.4, linestyle="--", label="Tin_eff")
        (self.line_inlet_cell,) = self.ax_outlet.plot([], [], color="#264653", linewidth=1.2, linestyle=":", label="Tg_in")
        self.ax_outlet.grid(alpha=0.25)
        self.ax_outlet.legend(loc="best")

        self.ax_profile.clear()
        self.ax_profile.set_title("Current Axial Temperature Profile")
        self.ax_profile.set_xlabel(f"x [{'m' if self._units == 'SI' else 'ft'}]")
        self.ax_profile.set_ylabel(f"Temperature [{'K' if self._units == 'SI' else 'F'}]")
        (self.line_tg,) = self.ax_profile.plot([], [], "--", linewidth=1.5, label="Gas Tg")
        (self.line_tw,) = self.ax_profile.plot([], [], "-", linewidth=1.5, label="Wall Tw")
        (self.line_ti,) = self.ax_profile.plot([], [], ":", linewidth=1.5, label="Insulation Ti")
        self.ax_profile.grid(alpha=0.25)
        self.ax_profile.legend(loc="upper right")

        self.ax_heat.clear()
        self.ax_heat.set_title("Live Heatmap (Gas Temperature)")
        self.ax_heat.set_xlabel(f"x [{'m' if self._units == 'SI' else 'ft'}]")
        self.ax_heat.set_ylabel("time [s]")
        self._heat_im = None

        self.live_canvas.draw_idle()
        self.heat_canvas.draw_idle()

        if hasattr(self, "results_fig"):
            self.results_fig.clear()
            self._results_axes = []
            self._results_cbar_parent = {}
            self.results_canvas.draw_idle()
        if hasattr(self, "stats_box"):
            self.stats_box.clear()
        if hasattr(self, "warning_box"):
            self.warning_box.clear()

    def _set_playback_data(self, result):
        self._play_times = np.asarray(result.times, dtype=float)
        self._play_tg = np.asarray(result.Tg_hist, dtype=float)
        self._play_tw = np.asarray(result.Tw_hist, dtype=float)
        self._play_ti = np.asarray(result.Ti_hist, dtype=float)
        self._play_tin_eff = self._inlet_temp_eff_series_si(self._play_times)

        if self._play_times.size == 0:
            self.slider_time.setEnabled(False)
            self.slider_time.setRange(0, 0)
            self.lbl_time_cursor.setText("t = -- s / -- s")
            return

        self.slider_time.setEnabled(True)
        self.slider_time.setRange(0, self._play_times.size - 1)
        self.slider_time.setValue(self._play_times.size - 1)
        self._render_playback_frame(self._play_times.size - 1)

    def _run_clicked(self):
        if self._thread is not None and self._thread.isRunning():
            return

        try:
            spec = self._collect_spec()
            opt_cfg = self._collect_optimization_config()
        except Exception as exc:
            QMessageBox.critical(self, "Input error", str(exc))
            return

        self._current_L = float(spec.hardware.L)
        self._last_result = None
        self._last_stats = {}
        self._last_warnings = []
        self._reset_live_views()
        self._update_target_time_readout()
        if hasattr(self, "btn_export"):
            self.btn_export.setEnabled(False)
        self._cancel_pending = False

        self.btn_run.setEnabled(False)
        self.btn_cancel.setEnabled(True)
        self.status.setText("Running...")
        self.progress_bar.setRange(0, 1000)
        self.progress_bar.setValue(0)

        self._thread = QThread(self)
        self._worker = SimulationWorker(spec, optimization=opt_cfg)
        self._worker.moveToThread(self._thread)

        self._thread.started.connect(self._worker.run)
        self._worker.snapshot.connect(self._on_snapshot)
        self._worker.finished.connect(self._on_finished)
        self._worker.cancelled.connect(self._on_cancelled)
        self._worker.failed.connect(self._on_failed)
        self._worker.finished.connect(self._thread.quit)
        self._worker.cancelled.connect(self._thread.quit)
        self._worker.failed.connect(self._thread.quit)
        self._thread.finished.connect(self._on_thread_finished)
        self._thread.finished.connect(self._thread.deleteLater)
        self._thread.start()

    def _cancel_clicked(self):
        if self._worker is None or self._thread is None or not self._thread.isRunning():
            return
        self._cancel_pending = True
        try:
            self._worker.request_cancel()
        except Exception:
            pass
        self.status.setText("Cancelling...")
        self.btn_cancel.setEnabled(False)

    @pyqtSlot(float, object, object, object)
    def _on_snapshot(self, t_s: float, tg_obj, tw_obj, ti_obj):
        self._snapshot_counter += 1
        tg = np.asarray(tg_obj, dtype=float)
        tw = np.asarray(tw_obj, dtype=float)
        ti = np.asarray(ti_obj, dtype=float)
        if tg.size == 0:
            return

        self._snap_t.append(float(t_s))
        self._snap_outlet.append(float(tg[-1]))
        self._snap_tw_out.append(float(tw[-1]))
        self._snap_ti_out.append(float(ti[-1]))
        tin_eff_si = self._inlet_temp_eff_at_si(float(t_s))
        self._snap_inlet.append(tin_eff_si)
        self._snap_inlet_cell.append(float(tg[0]))
        self._snap_tg_rows.append(tg.astype(np.float32, copy=True))

        self.lbl_runtime.setText(f"Sim time: {self._fmt_time_s(t_s)} s")
        self.lbl_outlet.setText(f"Outlet Tg: {self._fmt_temp_si(float(tg[-1]))} {self._temp_unit_label()}")
        self.lbl_wall_out.setText(f"Outlet Tw: {self._fmt_temp_si(float(tw[-1]))} {self._temp_unit_label()}")
        self.lbl_ins_out.setText(f"Outlet Ti: {self._fmt_temp_si(float(ti[-1]))} {self._temp_unit_label()}")
        self.lbl_inlet.setText(f"Tin_eff(bc): {self._fmt_temp_si(float(tin_eff_si))} {self._temp_unit_label()}")
        self.lbl_inlet_cell.setText(f"Tg_in(c0): {self._fmt_temp_si(float(tg[0]))} {self._temp_unit_label()}")
        self._update_live_readout(
            sim_time_s=float(t_s),
            tg_out_k=float(tg[-1]),
            tw_out_k=float(tw[-1]),
            ti_out_k=float(ti[-1]),
            tin_eff_k=float(tin_eff_si),
            tg_in_k=float(tg[0]),
        )
        self._update_target_time_readout()
        self.lbl_frames.setText(f"Frames: {len(self._snap_t)}")
        t_cap = max(float(self.in_t_end.value()), 1.0e-9)
        frac = max(0.0, min(1.0, float(t_s) / t_cap))
        self.progress_bar.setValue(int(round(1000.0 * frac)))

        if not HAS_MPL:
            return

        x = self._length_to_display(np.linspace(0.0, self._current_L, tg.size))
        self.line_outlet.set_data(self._snap_t, self._temp_to_display(np.asarray(self._snap_outlet)))
        self.line_inlet.set_data(self._snap_t, self._temp_to_display(np.asarray(self._snap_inlet)))
        self.line_inlet_cell.set_data(self._snap_t, self._temp_to_display(np.asarray(self._snap_inlet_cell)))
        self.ax_outlet.relim()
        self.ax_outlet.autoscale_view()

        self.line_tg.set_data(x, self._temp_to_display(tg))
        self.line_tw.set_data(x, self._temp_to_display(tw))
        self.line_ti.set_data(x, self._temp_to_display(ti))
        self.ax_profile.relim()
        self.ax_profile.autoscale_view()
        self.live_canvas.draw_idle()

        if self._snap_tg_rows:
            heat = self._temp_to_display(np.vstack(self._snap_tg_rows))
            x_max = float(self._length_to_display(self._current_L))
            extent = [0.0, x_max, self._snap_t[-1], self._snap_t[0]]
            self._update_live_heatmap(np.asarray(heat, dtype=float), extent)

    @pyqtSlot(object)
    def _on_finished(self, result):
        if str(getattr(result, "stop_reason", "")) == "aborted_by_user" or self._cancel_pending:
            self._on_cancelled(result)
            return
        self._last_result = result
        self.btn_run.setEnabled(True)
        self.btn_cancel.setEnabled(False)
        if hasattr(self, "btn_export"):
            self.btn_export.setEnabled(True)
        self.progress_bar.setRange(0, 1000)
        self.progress_bar.setValue(1000)

        msg = f"Done: {result.stop_reason}; Tg_out={self._fmt_temp_si(result.Tg_outlet_final)} {self._temp_unit_label()}"
        opt_summary = getattr(result, "opt_summary", None)
        if isinstance(opt_summary, dict):
            md_disp = self._fmt_mdot_si(float(opt_summary.get("m_dot_kg_s", self._last_run_si.get("m_dot", 0.0))))
            md_unit = "kg/s" if self._units == "SI" else "lbm/s"
            stress_ok = bool(opt_summary.get("meets_stress_limit", False))
            reached_ok = bool(opt_summary.get("target_reached_final", False))
            if opt_summary.get("mode") == "heatup_time_opt":
                t_final = opt_summary.get("time_to_target_final_s")
                time_ok = bool(opt_summary.get("meets_heatup_tolerance", False))
                if t_final is not None:
                    msg += f" | opt m_dot={md_disp} {md_unit}, t_hit={self._fmt_time_s(float(t_final))}s"
                else:
                    msg += f" | opt m_dot={md_disp} {md_unit}"
                msg += f" | ok: target={reached_ok}, time={time_ok}, stress={stress_ok}"
            elif opt_summary.get("mode") == "stress_limit_opt":
                t_final = opt_summary.get("time_to_target_final_s")
                if t_final is not None:
                    msg += f" | stress-opt m_dot={md_disp} {md_unit}, t_hit={self._fmt_time_s(float(t_final))}s"
                else:
                    msg += f" | stress-opt m_dot={md_disp} {md_unit}"
                msg += f" | ok: target={reached_ok}, stress={stress_ok}"
        if result.outdir:
            msg += f" | saved -> {result.outdir}"
        self.status.setText(msg)
        logging.info(msg)
        if isinstance(opt_summary, dict) and result.outdir:
            try:
                outdir = Path(result.outdir)
                outdir.mkdir(parents=True, exist_ok=True)
                (outdir / "optimization_summary.json").write_text(
                    json.dumps(opt_summary, indent=2, default=float),
                    encoding="utf-8",
                )
            except Exception as exc:
                logging.warning("Could not save optimization_summary.json: %s", exc)

        self._set_playback_data(result)
        if self.chk_make_plots.isChecked():
            self._save_plot_images_from_result(result)
        try:
            self._render_static_results(result)
        except Exception as exc:
            self._results_axes = []
            logging.exception("Failed to render embedded result plots: %s", exc)
        if self.chk_show_plots.isChecked() and HAS_MPL and self._results_axes:
            try:
                self._show_results_popup(self._results_axes[0])
            except Exception as exc:
                logging.warning("Could not open post-run results popup: %s", exc)
        if self.chk_append_ledger.isChecked():
            try:
                self._append_run_ledger(result)
            except Exception as exc:
                logging.warning("Failed to append run ledger: %s", exc)

        if HAS_MPL:
            self.live_canvas.draw_idle()
            self.heat_canvas.draw_idle()

    @pyqtSlot(str)
    def _on_failed(self, tb_text: str):
        self.btn_run.setEnabled(True)
        self.btn_cancel.setEnabled(False)
        self.progress_bar.setRange(0, 1000)
        self.progress_bar.setValue(0)
        self._cancel_pending = False
        self.status.setText("Error")
        QMessageBox.critical(self, "Simulation error", tb_text)

    @pyqtSlot()
    def _on_cancelled(self, result=None):
        self.btn_run.setEnabled(True)
        self.btn_cancel.setEnabled(False)
        self.progress_bar.setRange(0, 1000)
        self.progress_bar.setValue(0)
        self.status.setText("Cancelled by user")
        self._last_result = None
        self._last_stats = {}
        self._last_warnings = []
        self._cancel_pending = False
        # Discard all aborted run data/visuals.
        self._reset_live_views()
        if hasattr(self, "btn_export"):
            self.btn_export.setEnabled(False)
        outdir = getattr(result, "outdir", None) if result is not None else None
        if outdir:
            try:
                shutil.rmtree(Path(outdir), ignore_errors=True)
            except Exception as exc:
                logging.warning("Could not remove aborted run directory: %s", exc)
        logging.info("Simulation cancelled by user; partial data discarded.")

    @pyqtSlot()
    def _on_thread_finished(self):
        self.btn_cancel.setEnabled(False)
        self._cancel_pending = False
        self._worker = None
        self._thread = None

    def _export_bundle_clicked(self):
        if self._last_result is None:
            QMessageBox.information(self, "Export", "No run available yet. Run a simulation first.")
            return
        default_name = f"thermal_pipe_bundle_{datetime.now().strftime('%Y%m%d_%H%M%S')}.zip"
        base = self._save_dir if self._save_dir is not None else Path.cwd()
        path, _ = QFileDialog.getSaveFileName(self, "Export run bundle", str(base / default_name), "Zip archive (*.zip)")
        if not path:
            return
        zpath = Path(path)
        if zpath.suffix.lower() != ".zip":
            zpath = zpath.with_suffix(".zip")

        result = self._last_result
        try:
            with zipfile.ZipFile(zpath, "w", compression=zipfile.ZIP_DEFLATED) as zf:
                manifest = {
                    "created_local": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    "units": self._units,
                    "status": self.status.text(),
                    "has_saved_outdir": bool(result.outdir),
                }

                run_dir = result.outdir if result.outdir is not None else None
                if run_dir is not None:
                    run_dir = Path(run_dir)
                if run_dir is not None and run_dir.exists():
                    for fp in run_dir.rglob("*"):
                        if fp.is_file():
                            zf.write(fp, arcname=f"run/{fp.relative_to(run_dir)}")
                    manifest["run_source"] = str(run_dir)
                else:
                    buf = io.BytesIO()
                    np.savez_compressed(
                        buf,
                        x=np.asarray(result.x, dtype=float),
                        times=np.asarray(result.times, dtype=float),
                        Tg=np.asarray(result.Tg_hist, dtype=float),
                        Tw=np.asarray(result.Tw_hist, dtype=float),
                        Ti=np.asarray(result.Ti_hist, dtype=float),
                    )
                    zf.writestr("run/fields.npz", buf.getvalue())

                setup = {
                    "run_si": self._last_run_si,
                    "geom": self._last_geom,
                    "mech": self._last_mech,
                    "pipe_material_selected": self._last_pipe_material,
                    "mode_ui": self.mode.currentText(),
                }
                zf.writestr("session/current_setup.json", json.dumps(setup, indent=2))
                zf.writestr("session/stats.json", json.dumps(self._last_stats, indent=2, default=float))
                zf.writestr("session/warnings.txt", "\n".join(self._last_warnings) if self._last_warnings else "")

                opt_summary = getattr(result, "opt_summary", None)
                if isinstance(opt_summary, dict):
                    zf.writestr("session/optimization_summary.json", json.dumps(opt_summary, indent=2, default=float))
                    manifest["optimization_mode"] = str(opt_summary.get("mode", ""))

                if HAS_MPL and hasattr(self, "results_fig"):
                    pbuf = io.BytesIO()
                    self.results_fig.savefig(pbuf, format="png", dpi=180)
                    zf.writestr("session/results_plot.png", pbuf.getvalue())

                ledger_path = self._resolve_ledger_path()
                if ledger_path.exists() and ledger_path.is_file():
                    zf.write(ledger_path, arcname=f"ledger/{ledger_path.name}")
                if self._pipe_preset_path.exists():
                    zf.write(self._pipe_preset_path, arcname=f"library/{self._pipe_preset_path.name}")
                zf.writestr(
                    "library/materials_snapshot.json",
                    json.dumps(
                        {
                            "pipe_materials": PIPE_MATERIALS,
                            "insulation_materials": INSULATION_MATERIALS,
                            "pipe_temp_props": PIPE_TEMP_PROPS,
                            "insulation_temp_props": INSULATION_TEMP_PROPS,
                        },
                        indent=2,
                    ),
                )
                if self._readme_path.exists():
                    zf.write(self._readme_path, arcname="docs/README.md")

                zf.writestr("manifest.json", json.dumps(manifest, indent=2))
        except Exception as exc:
            QMessageBox.critical(self, "Export failed", f"Could not create bundle:\n{exc}")
            return

        QMessageBox.information(self, "Export", f"Bundle written to:\n{zpath}")

    def _run_animation(self):
        if self._play_times.size == 0:
            return
        if self.slider_time.value() >= self.slider_time.maximum():
            self.slider_time.setValue(0)
        self._play_timer.start()

    def _pause_animation(self):
        self._play_timer.stop()

    def _playback_tick(self):
        if self._play_times.size == 0:
            self._play_timer.stop()
            return
        idx = self.slider_time.value() + 1
        if idx > self.slider_time.maximum():
            self._play_timer.stop()
            return
        self.slider_time.setValue(idx)

    def _on_slider_changed(self, idx: int):
        self._render_playback_frame(int(idx))

    def _render_playback_frame(self, idx: int):
        if self._play_times.size == 0 or not HAS_MPL:
            return
        idx = max(0, min(idx, self._play_times.size - 1))
        temp_unit = "K" if self._units == "SI" else "F"
        x_unit = "m" if self._units == "SI" else "ft"

        t = float(self._play_times[idx])
        t_end = float(self._play_times[-1])
        self.lbl_runtime.setText(f"Sim time: {self._fmt_time_s(t)} s")
        self.lbl_time_cursor.setText(f"t = {self._fmt_time_s(t)} s / {self._fmt_time_s(t_end)} s")
        self.lbl_outlet.setText(f"Outlet Tg: {self._fmt_temp_si(self._play_tg[idx, -1])} {temp_unit}")
        self.lbl_wall_out.setText(f"Outlet Tw: {self._fmt_temp_si(self._play_tw[idx, -1])} {temp_unit}")
        self.lbl_ins_out.setText(f"Outlet Ti: {self._fmt_temp_si(self._play_ti[idx, -1])} {temp_unit}")
        if self._play_tin_eff.size > idx:
            self.lbl_inlet.setText(f"Tin_eff(bc): {self._fmt_temp_si(self._play_tin_eff[idx])} {temp_unit}")
        self.lbl_inlet_cell.setText(f"Tg_in(c0): {self._fmt_temp_si(self._play_tg[idx, 0])} {temp_unit}")
        self._update_live_readout(
            sim_time_s=t,
            tg_out_k=self._play_tg[idx, -1],
            tw_out_k=self._play_tw[idx, -1],
            ti_out_k=self._play_ti[idx, -1],
            tin_eff_k=self._play_tin_eff[idx] if self._play_tin_eff.size > idx else None,
            tg_in_k=self._play_tg[idx, 0],
        )
        self._update_target_time_readout(
            times_s=self._play_times[: idx + 1],
            tg_out_k=self._play_tg[: idx + 1, -1],
            tw_out_k=self._play_tw[: idx + 1, -1],
            ti_out_k=self._play_ti[: idx + 1, -1],
        )

        x = self._length_to_display(np.linspace(0.0, self._current_L, self._play_tg.shape[1]))
        tg = self._play_tg[idx]
        tw = self._play_tw[idx]
        ti = self._play_ti[idx]

        self.line_outlet.set_data(self._play_times[: idx + 1], self._temp_to_display(self._play_tg[: idx + 1, -1]))
        if self._play_tin_eff.size > 0:
            self.line_inlet.set_data(self._play_times[: idx + 1], self._temp_to_display(self._play_tin_eff[: idx + 1]))
        self.line_inlet_cell.set_data(self._play_times[: idx + 1], self._temp_to_display(self._play_tg[: idx + 1, 0]))
        self.ax_outlet.set_ylabel(f"Temperature [{temp_unit}]")
        self.ax_outlet.relim()
        self.ax_outlet.autoscale_view()

        self.line_tg.set_data(x, self._temp_to_display(tg))
        self.line_tw.set_data(x, self._temp_to_display(tw))
        self.line_ti.set_data(x, self._temp_to_display(ti))
        self.ax_profile.set_xlabel(f"x [{x_unit}]")
        self.ax_profile.set_ylabel(f"Temperature [{temp_unit}]")
        self.ax_profile.relim()
        self.ax_profile.autoscale_view()
        self.live_canvas.draw_idle()

        heat = self._temp_to_display(self._play_tg[: idx + 1])
        x_max = float(self._length_to_display(self._current_L))
        extent = [0.0, x_max, self._play_times[idx], self._play_times[0]]
        self.ax_heat.set_xlabel(f"x [{x_unit}]")
        self._update_live_heatmap(np.asarray(heat, dtype=float), extent)

    def _compute_stress_v2(
        self,
        tw_si: np.ndarray,
        ti_si: np.ndarray,
        nr_wall: int,
    ) -> Dict[str, np.ndarray | float]:
        nr = int(max(3, nr_wall))
        E = float(self._last_mech.get("E", 210.0e9))
        alpha = float(self._last_mech.get("alpha", 12.0e-6))
        nu = float(self._last_mech.get("nu", 0.30))
        Di_m = float(self._last_geom.get("Di", max(self._diam_from_display(self.in_Di.value()), 1.0e-6)))
        t_wall_m = float(self._last_geom.get("t_wall", max(self._diam_from_display(self.in_t_wall.value()), 1.0e-6)))
        t_ins_m = float(self._last_geom.get("t_ins", 0.0))
        k_w = float(self._last_geom.get("k_w", 15.0))
        k_i = float(self._last_geom.get("k_i", 0.04))
        p_si = float(self._last_run_si.get("p", self._pressure_from_display(self.in_p.value())))
        Tamb = float(self._last_run_si.get("Tamb", self._ambient_temp))
        axial_restraint = float(self._last_run_si.get("axial_restraint", self.in_axial_restraint.value()))
        include_pressure = bool(self.chk_include_pressure.isChecked())
        n_elbows = int(self._last_geom.get("n_elbows", self.in_elbows.value()))
        elbow_sif = float(self._last_geom.get("elbow_sif", self.in_elbow_sif.value()))
        elbow_positions_frac = list(self._last_geom.get("elbow_positions_frac", []))

        ri = max(0.5 * Di_m, 1.0e-9)
        ro = ri + max(t_wall_m, 1.0e-9)
        x_len = max(float(self._current_L), 1.0e-6)
        nx = max(1, tw_si.shape[1])
        dx_si = x_len / max(1, nx - 1)

        R_wall = np.log(max(ro / ri, 1.0 + 1.0e-12)) / (2.0 * np.pi * max(k_w, 1.0e-12) * max(dx_si, 1.0e-12))
        if t_ins_m > 1.0e-12:
            r_ins_o = ro + t_ins_m
            R_ins = np.log(max(r_ins_o / ro, 1.0 + 1.0e-12)) / (2.0 * np.pi * max(k_i, 1.0e-12) * max(dx_si, 1.0e-12))
        else:
            R_ins = 0.0
        wall_frac = float(R_wall / max(R_wall + R_ins, 1.0e-12))

        deltaT_wall_si = np.abs(tw_si - ti_si) * wall_frac
        xi = np.linspace(0.0, 1.0, nr, dtype=float)
        r = ri + xi * max(t_wall_m, 1.0e-9)
        grad_term = (0.5 - xi)[:, None, None] * deltaT_wall_si[None, :, :]
        temp_r = tw_si[None, :, :] + grad_term

        thermo_coeff = E * alpha / max(1.0e-9, (1.0 - nu))
        sigma_theta_th = thermo_coeff * (tw_si[None, :, :] - temp_r)
        sigma_z_th = -axial_restraint * thermo_coeff * (tw_si - Tamb)[None, :, :]
        sigma_r_th = np.zeros_like(sigma_theta_th)

        if include_pressure:
            denom = max(ro * ro - ri * ri, 1.0e-12)
            r2 = np.maximum(r * r, 1.0e-12)[:, None, None]
            coeff = p_si * ri * ri / denom
            sigma_r_p = coeff * (1.0 - (ro * ro) / r2)
            sigma_theta_p = coeff * (1.0 + (ro * ro) / r2)
            sigma_z_p = np.full_like(sigma_r_p, coeff)
        else:
            sigma_r_p = np.zeros_like(temp_r)
            sigma_theta_p = np.zeros_like(temp_r)
            sigma_z_p = np.zeros_like(temp_r)

        sigma_vm_th = np.sqrt(
            0.5
            * (
                (sigma_theta_th - sigma_z_th) ** 2
                + (sigma_z_th - sigma_r_th) ** 2
                + (sigma_r_th - sigma_theta_th) ** 2
            )
        )

        sigma_theta = sigma_theta_th + sigma_theta_p
        sigma_r = sigma_r_p
        sigma_z = sigma_z_th + sigma_z_p

        sigma_vm_total = np.sqrt(
            0.5
            * (
                (sigma_theta - sigma_z) ** 2
                + (sigma_z - sigma_r) ** 2
                + (sigma_r - sigma_theta) ** 2
            )
        )

        vm_thermal_mpa = np.max(sigma_vm_th, axis=0) / 1.0e6
        vm_total_mpa = np.max(sigma_vm_total, axis=0) / 1.0e6
        vm_inner_mpa = sigma_vm_total[0, :, :] / 1.0e6
        vm_outer_mpa = sigma_vm_total[-1, :, :] / 1.0e6

        elbow_profile = np.ones(nx, dtype=float)
        if n_elbows > 0 and nx > 2:
            if len(elbow_positions_frac) != n_elbows:
                elbow_positions_frac = self._parse_elbow_positions(
                    str(self._last_geom.get("elbow_positions_input", "")),
                    n_elbows,
                    x_len,
                )
            spread = max(1, int(0.03 * nx))
            for frac in elbow_positions_frac:
                j0 = int(round(max(0.0, min(1.0, float(frac))) * (nx - 1)))
                for dj in range(-3 * spread, 3 * spread + 1):
                    j = j0 + dj
                    if j < 0 or j >= nx:
                        continue
                    w = max(0.0, 1.0 - abs(dj) / (3.0 * spread + 1.0e-9))
                    elbow_profile[j] = max(elbow_profile[j], 1.0 + (elbow_sif - 1.0) * w)

        vm_total_elbow_mpa = vm_total_mpa * elbow_profile[None, :]
        vm_thermal_elbow_mpa = vm_thermal_mpa * elbow_profile[None, :]
        elbow_factor = float(np.max(elbow_profile))

        return {
            "deltaT_wall_si": deltaT_wall_si,
            "vm_map_mpa": vm_total_mpa,
            "vm_map_elbow_mpa": vm_total_elbow_mpa,
            "vm_thermal_mpa": vm_thermal_mpa,
            "vm_thermal_elbow_mpa": vm_thermal_elbow_mpa,
            "vm_inner_mpa": vm_inner_mpa,
            "vm_outer_mpa": vm_outer_mpa,
            "nr_wall": float(nr),
            "wall_frac": wall_frac,
            "elbow_factor": elbow_factor,
            "elbow_profile": elbow_profile,
        }

    @staticmethod
    def _target_crossing_time(times: np.ndarray, values: np.ndarray, target: float, mode_le: bool) -> float | None:
        if values.size == 0 or times.size == 0:
            return None
        cond = values <= target if mode_le else values >= target
        hit = np.where(cond)[0]
        if hit.size == 0:
            return None
        i = int(hit[0])
        if i == 0:
            return float(times[0])
        t0, t1 = float(times[i - 1]), float(times[i])
        y0, y1 = float(values[i - 1]), float(values[i])
        if abs(y1 - y0) <= 1.0e-12:
            return t1
        frac = (target - y0) / (y1 - y0)
        frac = max(0.0, min(1.0, frac))
        return t0 + frac * (t1 - t0)

    def _build_health_warnings(
        self,
        tw_si: np.ndarray,
        vm_map_elbow_mpa: np.ndarray,
        times: np.ndarray,
        ignore_inlet_cells: int,
        j_max: int,
    ) -> list[str]:
        warnings: list[str] = []
        Sy_mpa = float(self._last_mech.get("Sy", 250.0e6)) / 1.0e6
        alpha = float(self._last_mech.get("alpha", 12.0e-6))
        n_elbows = int(self._last_geom.get("n_elbows", self.in_elbows.value()))
        elbow_factor = float(self._last_geom.get("elbow_sif", self.in_elbow_sif.value())) if n_elbows > 0 else 1.0

        vm_max = float(np.nanmax(vm_map_elbow_mpa))
        tw_max = float(np.nanmax(tw_si))
        tw_min = float(np.nanmin(tw_si))
        thermal_strain_range = alpha * max(0.0, tw_max - tw_min)

        if vm_max >= Sy_mpa:
            warnings.append(
                f"High: max stress indicator {vm_max:.1f} MPa exceeds nominal yield {Sy_mpa:.1f} MPa (screening)."
            )
        elif vm_max >= 0.8 * Sy_mpa:
            warnings.append(
                f"Warning: max stress indicator {vm_max:.1f} MPa is above 80% of nominal yield {Sy_mpa:.1f} MPa."
            )

        if thermal_strain_range >= 2.0e-3:
            warnings.append(
                f"High cycle-thermal strain range ({thermal_strain_range:.4f}) may drive low-cycle fatigue/ratcheting."
            )
        elif thermal_strain_range >= 1.0e-3:
            warnings.append(
                f"Moderate thermal strain range ({thermal_strain_range:.4f}); monitor cyclic loading and hold times."
            )

        if tw_max >= 1050.0:
            warnings.append(
                f"High wall temperature {tw_max:.1f} K: creep/oxidation mechanisms can accelerate."
            )
        elif tw_max >= 900.0:
            warnings.append(
                f"Elevated wall temperature {tw_max:.1f} K: include high-temperature degradation checks."
            )

        dwell_high = float(times[-1] - times[0]) if times.size > 1 and tw_max >= 900.0 else 0.0
        ratchet_index = (vm_max / max(Sy_mpa, 1.0e-6)) * (thermal_strain_range / 1.0e-3) * (1.0 + dwell_high / 3600.0)
        if ratchet_index >= 1.5:
            warnings.append(
                f"Ratcheting screening index {ratchet_index:.2f} is high; consider inelastic assessment for cyclic service."
            )
        elif ratchet_index >= 1.0:
            warnings.append(
                f"Ratcheting screening index {ratchet_index:.2f} is moderate; track cycle accumulation in run ledger."
            )

        if ignore_inlet_cells > 0 and j_max < ignore_inlet_cells:
            warnings.append(
                "Peak stress occurs in early inlet cells; verify with inlet ramp and mesh sensitivity diagnostics."
            )

        if n_elbows > 0:
            elbow_input = str(self._last_geom.get("elbow_positions_input", self.in_elbow_positions.text())).strip()
            if not elbow_input:
                warnings.append(
                    "Elbow positions are auto-distributed; enter explicit positions for branch/layout-specific screening."
                )
            warnings.append(
                f"Elbow screening active: applied stress amplification factor {elbow_factor:.2f} (count={n_elbows})."
            )

        if not warnings:
            warnings.append("No immediate screening warnings for this run. Continue monitoring with logged history.")
        return warnings

    def _append_run_ledger(self, result):
        now = datetime.now()
        ledger_path = self._ledger_path
        if ledger_path is None:
            base = self._save_dir if self._save_dir is not None else Path.cwd()
            ledger_path = base / "run_history.csv"
        stats = self._last_stats if self._last_stats else {}
        opt_summary = getattr(result, "opt_summary", None)

        row = {
            "timestamp_local": now.strftime("%Y-%m-%d %H:%M:%S"),
            "asset_id": str(self._last_run_si.get("asset_id", self.asset_id.text().strip())),
            "branch_id": str(self._last_run_si.get("branch_id", self.branch_id.text().strip())),
            "material": self._last_pipe_material,
            "mode": self.mode.currentText(),
            "stop_reason": str(result.stop_reason),
            "reached_target": int(bool(result.reached_Tg_target)),
            "target_metric": str(self._last_run_si.get("target_metric", "gas_outlet")),
            "target_temp_K": float(self._last_run_si.get("target", np.nan)),
            "target_outlet_final_K": float(getattr(result, "target_outlet_final", np.nan)),
            "L_m": float(self._last_run_si.get("L", 0.0)),
            "Di_m": float(self._last_run_si.get("Di", 0.0)),
            "t_wall_m": float(self._last_run_si.get("t_wall", 0.0)),
            "t_ins_m": float(self._last_run_si.get("t_ins", 0.0)),
            "p_Pa": float(self._last_run_si.get("p", 0.0)),
            "Tin_K": float(self._last_run_si.get("Tin", 0.0)),
            "Tamb_K": float(self._last_run_si.get("Tamb", 0.0)),
            "m_dot_kg_s": float(self._last_run_si.get("m_dot", 0.0)),
            "Nx": int(self._last_run_si.get("Nx", 0)),
            "Nr_wall": int(self._last_run_si.get("nr_wall", 0)),
            "inlet_ramp_s": float(self._last_run_si.get("Tin_ramp_s", 0.0)),
            "inlet_ramp_model": str(self._last_run_si.get("Tin_ramp_model", "logistic")),
            "axial_restraint": float(self._last_run_si.get("axial_restraint", 0.0)),
            "elbows": int(self._last_geom.get("n_elbows", 0)),
            "elbow_sif": float(self._last_geom.get("elbow_sif", 1.0)),
            "thermal_masses": int(self._last_geom.get("n_tmass", 0)),
            "thermal_mass_factor": float(self._last_geom.get("tmass_factor", 0.0)),
            "thermal_mass_spread_pctL": float(self._last_geom.get("tmass_spread_frac", 0.0)) * 100.0,
            "thermal_mass_deadleg_len_m": float(self._last_geom.get("tmass_deadleg_len_m", 0.0)),
            "thermal_mass_deadleg_d_ratio": float(self._last_geom.get("tmass_deadleg_d_ratio", 1.0)),
            "Tg_out_final_K": float(result.Tg_outlet_final),
            "Tw_out_final_K": float(stats.get("Tw_outlet_final_K", np.nan)),
            "Tw_outer_surface_out_final_K": float(stats.get("Tw_outer_surface_outlet_K", np.nan)),
            "Ti_out_final_K": float(stats.get("Ti_outlet_final_K", np.nan)),
            "Tin_eff_final_K": float(stats.get("Tin_eff_final_K", np.nan)),
            "time_to_target_s": stats.get("time_to_target_s"),
            "max_vm_mpa": float(stats.get("sigma_max_mpa", np.nan)),
            "max_vm_elbow_mpa": float(stats.get("sigma_max_elbow_mpa", np.nan)),
            "max_vm_thermal_mpa": float(stats.get("sigma_max_thermal_mpa", np.nan)),
            "max_vm_thermal_elbow_mpa": float(stats.get("sigma_max_thermal_elbow_mpa", np.nan)),
            "t_max_vm_s": float(stats.get("t_sigma_max", np.nan)),
            "x_max_vm_m": float(stats.get("x_sigma_max_m", np.nan)),
            "max_wall_dT_K": float(stats.get("max_wall_dT_K", np.nan)),
            "warnings": " | ".join(self._last_warnings),
        }
        if isinstance(opt_summary, dict):
            row.update(
                {
                    "opt_mode": str(opt_summary.get("mode", "")),
                    "opt_selected_mdot_kg_s": float(opt_summary.get("m_dot_kg_s", np.nan)),
                    "opt_t_hit_final_s": opt_summary.get("time_to_target_final_s"),
                    "opt_sigma_final_mpa": float(opt_summary.get("sigma_final_mpa", np.nan)),
                    "opt_stress_limit_mpa": float(opt_summary.get("stress_limit_mpa", np.nan)),
                    "opt_heatup_target_s": float(opt_summary.get("heatup_target_s", np.nan)),
                    "opt_heatup_tol_s": float(opt_summary.get("heatup_tol_s", np.nan)),
                    "opt_meets_stress_limit": int(bool(opt_summary.get("meets_stress_limit", False))),
                    "opt_meets_heatup_tolerance": int(bool(opt_summary.get("meets_heatup_tolerance", False))),
                    "opt_target_reached_final": int(bool(opt_summary.get("target_reached_final", False))),
                }
            )
        row = {k: self._round_ledger_value(k, v) for k, v in row.items()}

        incoming_headers = list(row.keys())
        if ledger_path.suffix.lower() == ".xlsx":
            try:
                headers_existing, rows_existing = self._load_xlsx_ledger_rows(ledger_path) if ledger_path.exists() else ([], [])
                headers = self._merge_ledger_headers(headers_existing, incoming_headers)
                rows_existing.append(row)
                self._write_xlsx_ledger_rows(ledger_path, headers, rows_existing)
                self._refresh_ledger_preview()
                return
            except Exception as exc:
                logging.warning("XLSX ledger append failed; falling back to CSV ledger. Reason: %s", exc)
                ledger_path = ledger_path.with_suffix(".csv")

        headers_existing, rows_existing = self._load_csv_ledger_rows(ledger_path) if ledger_path.exists() else ([], [])
        headers = self._merge_ledger_headers(headers_existing, incoming_headers)
        rows_existing.append(row)
        self._write_csv_ledger_rows(ledger_path, headers, rows_existing)
        self._refresh_ledger_preview()

    def _on_results_plot_click(self, event):
        if not HAS_MPL:
            return
        if event is None or event.inaxes is None:
            return
        try:
            if event.button is not None and int(event.button) != 1:
                return
        except Exception:
            pass
        source_ax = self._results_cbar_parent.get(event.inaxes, event.inaxes)
        if source_ax not in self._results_axes:
            return
        self._show_results_popup(source_ax)

    def _show_results_popup(self, source_ax):
        dlg = QDialog(self)
        dlg.setWindowTitle(source_ax.get_title() or "Plot Detail")
        dlg.resize(1000, 700)
        layout = QVBoxLayout(dlg)
        fig = Figure(figsize=(10, 7), constrained_layout=True)
        canvas = FigureCanvas(fig)
        layout.addWidget(canvas, 1)

        ax = fig.add_subplot(1, 1, 1)
        self._copy_axis_content(source_ax, ax, fig)
        canvas.draw_idle()
        dlg.exec()

    @staticmethod
    def _copy_axis_content(source_ax, target_ax, target_fig):
        axis_aspect = source_ax.get_aspect() if hasattr(source_ax, "get_aspect") else "auto"
        for line in source_ax.get_lines():
            label = line.get_label()
            if label.startswith("_"):
                label = "_nolegend_"
            target_ax.plot(
                line.get_xdata(),
                line.get_ydata(),
                color=line.get_color(),
                linestyle=line.get_linestyle(),
                linewidth=line.get_linewidth(),
                marker=line.get_marker(),
                markersize=line.get_markersize(),
                alpha=line.get_alpha(),
                label=label,
            )
        for image in source_ax.images:
            data = np.asarray(image.get_array())
            copied = target_ax.imshow(
                data,
                aspect=axis_aspect,
                extent=image.get_extent(),
                origin=getattr(image, "origin", "upper"),
                cmap=image.get_cmap(),
                vmin=image.get_clim()[0],
                vmax=image.get_clim()[1],
            )
            cbar_label = ""
            colorbar = getattr(image, "colorbar", None)
            if colorbar is not None:
                try:
                    cbar_label = str(colorbar.ax.get_ylabel() or "")
                except Exception:
                    cbar_label = ""
            target_fig.colorbar(copied, ax=target_ax, label=cbar_label)

        target_ax.set_title(source_ax.get_title(), fontsize=13, pad=10)
        target_ax.set_xlabel(source_ax.get_xlabel(), fontsize=11)
        target_ax.set_ylabel(source_ax.get_ylabel(), fontsize=11)
        target_ax.tick_params(labelsize=10)
        target_ax.set_xlim(source_ax.get_xlim())
        target_ax.set_ylim(source_ax.get_ylim())
        if any(gl.get_visible() for gl in source_ax.get_xgridlines() + source_ax.get_ygridlines()):
            target_ax.grid(alpha=0.25)

        handles, labels = target_ax.get_legend_handles_labels()
        if any(lbl and lbl != "_nolegend_" for lbl in labels):
            target_ax.legend(loc="best", fontsize=10, framealpha=0.9)

    def _save_plot_images_from_result(self, result):
        if not HAS_MPL:
            return
        outdir = getattr(result, "outdir", None)
        if outdir is None:
            logging.warning("Plot image save requested, but no run output directory is available.")
            return
        try:
            outdir = Path(outdir)
            outdir.mkdir(parents=True, exist_ok=True)
            x_si = np.asarray(result.x, dtype=float)
            times = np.asarray(result.times, dtype=float)
            tw_si = np.asarray(result.Tw_hist, dtype=float)
            tg_si = np.asarray(result.Tg_hist, dtype=float)
            ti_si = np.asarray(result.Ti_hist, dtype=float)
            if times.size == 0 or x_si.size == 0:
                return

            x = np.asarray(self._length_to_display(x_si), dtype=float)
            tw = np.asarray(self._temp_to_display(tw_si), dtype=float)
            tg = np.asarray(self._temp_to_display(tg_si), dtype=float)
            ti = np.asarray(self._temp_to_display(ti_si), dtype=float)
            t = np.asarray(times, dtype=float)
            temp_unit = self._temp_unit_label()
            x_unit = "m" if self._units == "SI" else "ft"

            heat_fig = Figure(figsize=(15, 4), constrained_layout=True)
            axw = heat_fig.add_subplot(1, 3, 1)
            axg = heat_fig.add_subplot(1, 3, 2)
            axi = heat_fig.add_subplot(1, 3, 3)
            ext = [float(x[0]), float(x[-1]), float(t[-1]), float(t[0])]
            imw = axw.imshow(tw, aspect="auto", extent=ext)
            img = axg.imshow(tg, aspect="auto", extent=ext)
            imi = axi.imshow(ti, aspect="auto", extent=ext)
            axw.set_title("Wall Tw(x,t)")
            axg.set_title("Gas Tg(x,t)")
            axi.set_title("Insulation Ti(x,t)")
            for ax in (axw, axg, axi):
                ax.set_xlabel(f"x [{x_unit}]")
                ax.set_ylabel("time [s]")
            heat_fig.colorbar(imw, ax=axw, label=temp_unit)
            heat_fig.colorbar(img, ax=axg, label=temp_unit)
            heat_fig.colorbar(imi, ax=axi, label=temp_unit)
            heat_fig.savefig(outdir / "heatmaps.png", dpi=200)

            prof_fig = Figure(figsize=(10, 4), constrained_layout=True)
            axp = prof_fig.add_subplot(1, 1, 1)
            nmax = 30
            idx = np.linspace(0, max(0, t.size - 1), min(nmax, max(1, t.size)), dtype=int)
            for i in idx:
                axp.plot(x, tw[i], linewidth=1.0, alpha=0.8)
            for i in idx:
                axp.plot(x, tg[i], "--", linewidth=1.0, alpha=0.8)
            for i in idx:
                axp.plot(x, ti[i], ":", linewidth=1.0, alpha=0.8)
            axp.set_xlabel(f"x [{x_unit}]")
            axp.set_ylabel(f"Temperature [{temp_unit}]")
            axp.set_title("Profiles Over Time")
            axp.grid(alpha=0.25)
            prof_fig.savefig(outdir / "profiles.png", dpi=200)
            logging.info("Saved plot images: heatmaps.png, profiles.png")
        except Exception as exc:
            logging.warning("Failed to save plot images: %s", exc)

    def _render_static_results(self, result):
        if not HAS_MPL or not hasattr(self, "results_fig"):
            return
        if result.times.size == 0:
            self.results_fig.clear()
            self._results_axes = []
            self._results_cbar_parent = {}
            self.results_canvas.draw_idle()
            if hasattr(self, "stats_box"):
                self.stats_box.setPlainText("No results available.")
            if hasattr(self, "warning_box"):
                self.warning_box.setPlainText("No warnings available.")
            self._last_stats = {}
            self._last_warnings = []
            return

        self.results_fig.clear()
        axs = self.results_fig.subplots(2, 3)
        self._results_axes = [axs[0, 0], axs[0, 1], axs[0, 2], axs[1, 0], axs[1, 1], axs[1, 2]]
        self._results_cbar_parent = {}

        x_si = np.asarray(result.x, dtype=float)
        times = np.asarray(result.times, dtype=float)
        tg_si = np.asarray(result.Tg_hist, dtype=float)
        tw_si = np.asarray(result.Tw_hist, dtype=float)
        ti_si = np.asarray(result.Ti_hist, dtype=float)

        temp_unit = self._temp_unit_label()
        x_unit = self._length_unit_label()
        stress_unit = self._stress_unit_label()

        x = np.asarray(self._length_to_display(x_si), dtype=float)
        tg = np.asarray(self._temp_to_display(tg_si), dtype=float)
        tw = np.asarray(self._temp_to_display(tw_si), dtype=float)
        ti = np.asarray(self._temp_to_display(ti_si), dtype=float)
        target_metric = _sanitize_target_metric(str(self._last_run_si.get("target_metric", "gas_outlet")))
        target_metric_label = TARGET_METRIC_LABELS.get(target_metric, target_metric)
        target_series_si = _outlet_series_for_metric(tg_si, tw_si, ti_si, target_metric, self._last_geom)
        target_series = np.asarray(self._temp_to_display(target_series_si), dtype=float)
        tin_eff_si = self._inlet_temp_eff_series_si(times)
        tin_eff = np.asarray(self._temp_to_display(tin_eff_si), dtype=float)
        ramp_model = str(self._last_run_si.get("Tin_ramp_model", "logistic"))
        ramp_s = float(self._last_run_si.get("Tin_ramp_s", 0.0))
        alpha = float(self._last_mech.get("alpha", 12.0e-6))
        nr_wall = int(self.in_nr_wall.value())
        ignore_inlet_cells = int(self.in_ignore_inlet_cells.value())
        stress_v2 = self._compute_stress_v2(tw_si, ti_si, nr_wall=nr_wall)
        vm_map_mpa = np.asarray(stress_v2["vm_map_mpa"], dtype=float)
        vm_map_elbow_mpa = np.asarray(stress_v2["vm_map_elbow_mpa"], dtype=float)
        deltaT_wall_si = np.asarray(stress_v2["deltaT_wall_si"], dtype=float)

        vm_map_disp = self._stress_to_display(vm_map_mpa)
        vm_map_elbow_disp = self._stress_to_display(vm_map_elbow_mpa)
        vm_time_disp = np.max(vm_map_disp, axis=1)
        vm_time_elbow_disp = np.max(vm_map_elbow_disp, axis=1)

        # 1) Time-distance heatmap
        ax = axs[0, 0]
        im = ax.imshow(tg, aspect="auto", extent=[x[0], x[-1], times[-1], times[0]], origin="upper")
        ax.set_title("Gas Heatmap Tg(x,t)", fontsize=12, pad=8)
        ax.set_xlabel(f"x [{x_unit}]")
        ax.set_ylabel("time [s]")
        cbar_t = self.results_fig.colorbar(im, ax=ax, label=temp_unit, fraction=0.046, pad=0.04)
        self._results_cbar_parent[cbar_t.ax] = ax

        # 2) Waterfall-style profiles
        ax = axs[0, 1]
        idx = np.linspace(0, times.size - 1, min(14, times.size), dtype=int)
        for i in idx:
            ax.plot(x, tg[i], linewidth=1.0, alpha=0.85)
        ax.set_title("Waterfall Tg(x)", fontsize=12, pad=8)
        ax.set_xlabel(f"x [{x_unit}]")
        ax.set_ylabel(f"Tg [{temp_unit}]")
        ax.grid(alpha=0.25)

        # 3) Outlet temperatures over time
        ax = axs[0, 2]
        ax.plot(times, tg[:, -1], label="Tg_out", linewidth=1.8)
        ax.plot(times, tg[:, 0], label="Tg_in(c0)", linewidth=1.2, linestyle=":")
        ax.plot(times, tw[:, -1], label="Tw_out", linewidth=1.2)
        ax.plot(times, ti[:, -1], label="Ti_out", linewidth=1.2)
        ax.plot(times, tin_eff, label="Tin_eff", linewidth=1.2, linestyle="--", color="#6c757d")
        if target_metric == "wall_outer_outlet":
            ax.plot(times, target_series, label="Tw_out(surface)", linewidth=1.2, linestyle="-.")
        if result.Tg_outlet_target is not None:
            target_disp = float(self._temp_to_display(float(result.Tg_outlet_target)))
            ax.axhline(target_disp, color="#666666", linestyle="--", linewidth=1.0, label="Target")
        ax.set_title("Outlet & Inlet vs Time", fontsize=12, pad=8)
        ax.set_xlabel("time [s]")
        ax.set_ylabel(f"Temperature [{temp_unit}]")
        ax.grid(alpha=0.25)
        ax.legend(loc="lower right", fontsize=8, framealpha=0.88)

        # 4) Final total stress profile (straight vs elbow-adjusted)
        ax = axs[1, 0]
        ax.plot(x, vm_map_disp[-1], color="#d55e00", linewidth=1.6, label="Total (straight)")
        if np.nanmax(vm_map_elbow_disp - vm_map_disp) > 1.0e-9:
            ax.plot(x, vm_map_elbow_disp[-1], color="#cc79a7", linewidth=1.2, linestyle="--", label="Total (elbow)")
        ax.set_title("Final Total Stress Profile", fontsize=12, pad=8)
        ax.set_xlabel(f"x [{x_unit}]")
        ax.set_ylabel(f"sigma [{stress_unit}]")
        ax.grid(alpha=0.25)
        ax.legend(loc="upper left", fontsize=8, framealpha=0.88)

        # 5) Max total stress vs time
        ax = axs[1, 1]
        ax.plot(times, vm_time_disp, color="#d55e00", linewidth=1.5, label="Total (straight)")
        if np.nanmax(vm_time_elbow_disp - vm_time_disp) > 1.0e-9:
            ax.plot(times, vm_time_elbow_disp, color="#cc79a7", linewidth=1.2, linestyle="--", label="Total (elbow)")
        ax.set_title("Max Total Stress vs Time", fontsize=12, pad=8)
        ax.set_xlabel("time [s]")
        ax.set_ylabel(f"max sigma [{stress_unit}]")
        ax.grid(alpha=0.25)
        ax.legend(loc="upper left", fontsize=8, framealpha=0.88)

        # 6) Stress map over time and distance
        ax = axs[1, 2]
        im_sigma = ax.imshow(
            vm_map_elbow_disp,
            aspect="auto",
            extent=[x[0], x[-1], times[-1], times[0]],
            origin="upper",
        )
        ax.set_title("Total Stress Heatmap |sigma_vm|", fontsize=12, pad=8)
        ax.set_xlabel(f"x [{x_unit}]")
        ax.set_ylabel("time [s]")
        cbar_s = self.results_fig.colorbar(im_sigma, ax=ax, label=stress_unit, fraction=0.046, pad=0.04)
        self._results_cbar_parent[cbar_s.ax] = ax

        for a in self._results_axes:
            a.tick_params(labelsize=9)

        # Keep explicit spacing modest; constrained_layout handles final fit.
        try:
            self.results_fig.set_constrained_layout_pads(w_pad=0.03, h_pad=0.04, wspace=0.04, hspace=0.05)
        except Exception:
            pass

        # Statistics text
        i_max, j_max = np.unravel_index(np.nanargmax(vm_map_elbow_mpa), vm_map_elbow_mpa.shape)
        sigma_max_mpa = float(vm_map_mpa[i_max, j_max])
        sigma_max_elbow_mpa = float(vm_map_elbow_mpa[i_max, j_max])
        t_sigma_max = float(times[i_max])
        x_sigma_max_m = float(x_si[j_max])
        x_sigma_max_disp = float(x[j_max])

        dL_m = float(alpha * np.trapz((tw_si[-1] - float(self._ambient_temp)), x_si))
        if self._units == "SI":
            dL_text = f"{self._fmt_num(dL_m * 1000.0, 2)} mm"
        else:
            dL_text = f"{self._fmt_num(dL_m * M2IN, 3)} in"

        t_hit_s = None
        if result.Tg_outlet_target is None:
            target_time_text = "N/A (fixed-time mode)"
        else:
            target_si = float(result.Tg_outlet_target)
            mode_le = bool(target_si <= float(target_series_si[0]))
            t_hit_s = self._target_crossing_time(times, target_series_si, target_si, mode_le=mode_le)
            target_time_text = f"{self._fmt_time_s(t_hit_s)} s" if t_hit_s is not None else "Not reached"

        i_dt, j_dt = np.unravel_index(np.nanargmax(deltaT_wall_si), deltaT_wall_si.shape)
        deltaT_wall_scale = 1.0 if self._units == "SI" else 9.0 / 5.0
        max_deltaT_wall = float(deltaT_wall_si[i_dt, j_dt] * deltaT_wall_scale)
        t_max_dT = float(times[i_dt])
        x_max_dT = float(x[j_dt])

        sigma_max_disp = self._fmt_stress_si(sigma_max_mpa, 2)
        sigma_max_elbow_disp = self._fmt_stress_si(sigma_max_elbow_mpa, 2)
        wall_frac = _wall_fraction_from_geom(self._last_geom)
        tw_outer_surface_series_si = tw_si[:, -1] - wall_frac * (tw_si[:, -1] - ti_si[:, -1])

        tg_out_final_disp = self._fmt_temp_si(tg_si[-1, -1], 1)
        tg_in_final_disp = self._fmt_temp_si(tg_si[-1, 0], 1)
        tw_out_final_disp = self._fmt_temp_si(tw_si[-1, -1], 1)
        tw_in_final_disp = self._fmt_temp_si(tw_si[-1, 0], 1)
        tw_outer_surface_final_disp = self._fmt_temp_si(tw_outer_surface_series_si[-1], 1)
        ti_out_final_disp = self._fmt_temp_si(ti_si[-1, -1], 1)
        ti_in_final_disp = self._fmt_temp_si(ti_si[-1, 0], 1)
        tin_eff_final_disp = self._fmt_temp_si(tin_eff_si[-1], 1) if tin_eff_si.size else "--"
        tg_mean_final_disp = self._fmt_temp_si(np.mean(tg_si[-1]), 1)
        tw_mean_final_disp = self._fmt_temp_si(np.mean(tw_si[-1]), 1)
        ti_mean_final_disp = self._fmt_temp_si(np.mean(ti_si[-1]), 1)
        target_final_disp = self._fmt_temp_si(target_series_si[-1], 1)

        radial_diag_text = "off"
        axial_diag_text = "off"
        artifact_ratio_text = "off"
        radial_delta = np.nan
        axial_delta = np.nan
        sigma_excluding_inlet_disp = np.nan
        hotspot_inlet_flag = bool(ignore_inlet_cells > 0 and j_max < ignore_inlet_cells)
        if self.chk_convergence_diag.isChecked():
            nr_ref = max(2 * nr_wall, nr_wall + 2)
            stress_refined = self._compute_stress_v2(tw_si, ti_si, nr_wall=nr_ref)
            vm_ref_mpa = np.asarray(stress_refined["vm_map_elbow_mpa"], dtype=float)
            sigma_ref_mpa = float(np.nanmax(vm_ref_mpa))
            radial_delta = abs(sigma_max_elbow_mpa - sigma_ref_mpa) / max(abs(sigma_ref_mpa), 1.0e-9)
            radial_diag_text = f"{self._fmt_num(100.0 * radial_delta, 1)}% (Nr={nr_wall}->{nr_ref})"

            vm_coarse = vm_map_elbow_mpa[:, ::2] if vm_map_elbow_mpa.shape[1] > 2 else vm_map_elbow_mpa
            sigma_coarse_mpa = float(np.nanmax(vm_coarse))
            axial_delta = abs(sigma_max_elbow_mpa - sigma_coarse_mpa) / max(abs(sigma_max_elbow_mpa), 1.0e-9)
            axial_diag_text = f"{self._fmt_num(100.0 * axial_delta, 1)}% (coarse-x)"

            if ignore_inlet_cells < vm_map_elbow_mpa.shape[1]:
                sigma_excluding_inlet_mpa = float(np.nanmax(vm_map_elbow_mpa[:, ignore_inlet_cells:]))
                sigma_excluding_inlet_disp = float(self._stress_to_display(sigma_excluding_inlet_mpa))
                artifact_ratio = sigma_max_elbow_mpa / max(sigma_excluding_inlet_mpa, 1.0e-9)
                artifact_ratio_text = f"{self._fmt_num(artifact_ratio, 2)}x"
            else:
                artifact_ratio_text = "n/a"

        tracking_error_disp = (
            float(self._temp_to_display(float(tin_eff_si[-1]))) - float(self._temp_to_display(float(tg_si[-1, 0])))
            if tin_eff_si.size
            else np.nan
        )

        stats_lines = [
            "Run Statistics",
            f"- Time to target condition ({target_metric_label}): {target_time_text}",
            f"- Max total stress indicator (straight): {sigma_max_disp} {stress_unit}",
            f"- Max total stress indicator (elbow-adjusted): {sigma_max_elbow_disp} {stress_unit}",
            f"- Time of max stress: {self._fmt_time_s(t_sigma_max)} s",
            f"- Location of max stress: {self._fmt_num(x_sigma_max_disp, 3)} {x_unit}",
            f"- Free thermal expansion estimate: {dL_text}",
            f"- Max estimated wall through-thickness dT: {self._fmt_num(max_deltaT_wall, 1)} {temp_unit}",
            f"- Time/location of max wall dT: {self._fmt_time_s(t_max_dT)} s @ {self._fmt_num(x_max_dT, 3)} {x_unit}",
            f"- Radial sensitivity diagnostic: {radial_diag_text}",
            f"- Axial sensitivity diagnostic: {axial_diag_text}",
            f"- Inlet-hotspot ratio (global / excluding first cells): {artifact_ratio_text}",
            f"- Final target variable ({target_metric_label}): {target_final_disp} {temp_unit}",
            f"- Final gas temperature: inlet={tg_in_final_disp}, outlet={tg_out_final_disp} {temp_unit}",
            f"- Final inlet tracking error (Tin_eff - Tg_in): {self._fmt_num(tracking_error_disp, 1)} {temp_unit}",
            f"- Final wall temperature (inner node): inlet={tw_in_final_disp}, outlet={tw_out_final_disp} {temp_unit}",
            f"- Final wall outer-surface estimate at outlet: {tw_outer_surface_final_disp} {temp_unit}",
            f"- Final insulation temperature: inlet={ti_in_final_disp}, outlet={ti_out_final_disp} {temp_unit}",
            f"- Final axial mean temperatures: Tg={tg_mean_final_disp}, Tw={tw_mean_final_disp}, Ti={ti_mean_final_disp} {temp_unit}",
            f"- Inlet heater setpoint at end: {tin_eff_final_disp} {temp_unit} (model={ramp_model}, warm-up={self._fmt_time_s(ramp_s)} s)",
            f"- Simulation runtime: sim={self._fmt_time_s(times[-1])} s, steps={int(result.n_steps)}, wall={self._fmt_num(result.wall_time_s, 2)} s",
        ]
        opt_summary = getattr(result, "opt_summary", None)
        if isinstance(opt_summary, dict):
            md_disp = self._fmt_mdot_si(float(opt_summary.get("m_dot_kg_s", np.nan)))
            md_unit = "kg/s" if self._units == "SI" else "lbm/s"
            mode_tag = str(opt_summary.get("mode", ""))
            sigma_lim_disp = self._fmt_stress_si(float(opt_summary.get("stress_limit_mpa", np.nan)), 2)
            sigma_final_disp = self._fmt_stress_si(float(opt_summary.get("sigma_final_mpa", np.nan)), 2)
            meets_stress = bool(opt_summary.get("meets_stress_limit", False))
            reached_final = bool(opt_summary.get("target_reached_final", False))
            if mode_tag == "heatup_time_opt":
                t_goal = float(opt_summary.get("heatup_target_s", np.nan))
                t_tol = float(opt_summary.get("heatup_tol_s", np.nan))
                t_hit = opt_summary.get("time_to_target_final_s")
                t_hit_txt = "not reached" if t_hit is None else f"{self._fmt_time_s(float(t_hit))} s"
                meets_heatup = bool(opt_summary.get("meets_heatup_tolerance", False))
                stats_lines.extend(
                    [
                        "- Optimization mode: Heatup-time optimize",
                        f"- Selected mass flow: {md_disp} {md_unit}",
                        f"- Heatup target: {self._fmt_time_s(t_goal)} s +/- {self._fmt_time_s(t_tol)} s",
                        f"- Final run time-to-target: {t_hit_txt}",
                        f"- Final total stress: {sigma_final_disp} {stress_unit} (limit {sigma_lim_disp} {stress_unit})",
                        f"- Constraint status: target_reached={reached_final}, heatup_tol={meets_heatup}, stress_limit={meets_stress}",
                    ]
                )
            elif mode_tag == "stress_limit_opt":
                t_hit = opt_summary.get("time_to_target_final_s")
                t_hit_txt = "not reached" if t_hit is None else f"{self._fmt_time_s(float(t_hit))} s"
                stats_lines.extend(
                    [
                        "- Optimization mode: Stress-limit optimize",
                        f"- Selected mass flow: {md_disp} {md_unit}",
                        f"- Final run time-to-target: {t_hit_txt}",
                        f"- Final total stress: {sigma_final_disp} {stress_unit} (limit {sigma_lim_disp} {stress_unit})",
                        f"- Constraint status: target_reached={reached_final}, stress_limit={meets_stress}",
                    ]
                )

        warnings = self._build_health_warnings(
            tw_si=tw_si,
            vm_map_elbow_mpa=vm_map_elbow_mpa,
            times=times,
            ignore_inlet_cells=ignore_inlet_cells,
            j_max=j_max,
        )
        if self.chk_convergence_diag.isChecked() and np.isfinite(radial_delta) and radial_delta > 0.15:
            warnings.append(
                "Stress changes by >15% under radial refinement; increase Nr_wall or treat result as low-confidence."
            )
        if self.chk_convergence_diag.isChecked() and np.isfinite(axial_delta) and axial_delta > 0.20:
            warnings.append(
                "Stress changes by >20% under coarse axial sensitivity check; consider higher Nx and inlet ramp."
            )
        if np.isfinite(sigma_excluding_inlet_disp) and hotspot_inlet_flag:
            warnings.append(
                f"Inlet-cell exclusion check: max stress away from inlet is {self._fmt_num(sigma_excluding_inlet_disp, 2)} {stress_unit}."
            )
        opt_summary = getattr(result, "opt_summary", None)
        if isinstance(opt_summary, dict):
            if not bool(opt_summary.get("target_reached_final", True)):
                warnings.append("Optimization final run did not reach outlet target within allowed simulation time.")
            if not bool(opt_summary.get("meets_stress_limit", True)):
                lim_disp = self._fmt_stress_si(float(opt_summary.get("stress_limit_mpa", np.nan)), 2)
                sig_disp = self._fmt_stress_si(float(opt_summary.get("sigma_final_mpa", np.nan)), 2)
                warnings.append(
                    f"Optimization final stress exceeds limit ({sig_disp} > {lim_disp} {stress_unit})."
                )
            if opt_summary.get("mode") == "heatup_time_opt" and not bool(opt_summary.get("meets_heatup_tolerance", True)):
                warnings.append("Heatup-time optimization did not satisfy requested time tolerance in final run.")

        if hasattr(self, "stats_box"):
            self.stats_box.setPlainText("\n".join(stats_lines))
        if hasattr(self, "warning_box"):
            self.warning_box.setPlainText("Health / Fatigue Warnings (screening)\n" + "\n".join(f"- {w}" for w in warnings))

        self._last_warnings = warnings
        self._last_stats = {
            "time_to_target_s": t_hit_s,
            "target_metric": target_metric,
            "target_metric_label": target_metric_label,
            "sigma_max_thermal_mpa": sigma_max_mpa,
            "sigma_max_thermal_elbow_mpa": sigma_max_elbow_mpa,
            "sigma_max_mpa": sigma_max_mpa,
            "sigma_max_elbow_mpa": sigma_max_elbow_mpa,
            "t_sigma_max": t_sigma_max,
            "x_sigma_max_m": x_sigma_max_m,
            "max_wall_dT_K": float(np.nanmax(deltaT_wall_si)),
            "Tg_outlet_final_K": float(result.Tg_outlet_final),
            "Tg_inlet_final_K": float(tg_si[-1, 0]),
            "Tw_outlet_final_K": float(tw_si[-1, -1]),
            "Tw_inlet_final_K": float(tw_si[-1, 0]),
            "Tw_outer_surface_outlet_K": float(tw_outer_surface_series_si[-1]),
            "Ti_outlet_final_K": float(ti_si[-1, -1]),
            "Ti_inlet_final_K": float(ti_si[-1, 0]),
            "Tin_eff_final_K": float(tin_eff_si[-1]) if tin_eff_si.size else float("nan"),
            "Tg_mean_final_K": float(np.mean(tg_si[-1])),
            "Tw_mean_final_K": float(np.mean(tw_si[-1])),
            "Ti_mean_final_K": float(np.mean(ti_si[-1])),
            "sim_time_s": float(times[-1]),
            "n_steps": int(result.n_steps),
            "wall_time_s": float(result.wall_time_s),
            "tin_ramp_model": ramp_model,
            "tin_ramp_s": ramp_s,
            "radial_diag": float(radial_delta) if np.isfinite(radial_delta) else None,
            "axial_diag": float(axial_delta) if np.isfinite(axial_delta) else None,
            "hotspot_inlet": bool(hotspot_inlet_flag),
        }

        self.results_canvas.draw_idle()

    def closeEvent(self, event):
        self._play_timer.stop()
        if hasattr(self, "_worker") and self._worker is not None:
            self._worker.cancel()
        if hasattr(self, "_thread") and self._thread is not None and self._thread.isRunning():
            self._thread.quit()
            self._thread.wait(2000)
        
        root_logger = logging.getLogger()
        if getattr(self, "log_handler", None) is not None:
            try:
                root_logger.removeHandler(self.log_handler)
            except Exception:
                pass
            try:
                self.log_handler.close()
            except Exception:
                pass
        super().closeEvent(event)


def main():
    app = QApplication(sys.argv)
    window = ThermalPipeWindow()
    window.show()
    sys.exit(app.exec())


if __name__ == "__main__":
    main()
